#!/usr/bin/env python3
"""
Chicken Valuation Engine
========================
Computes broiler carcass value from manually-entered cut prices.
No USDA API — prices come from manual_entry.py.

Usage:
  python3 chicken_valuation.py
  python3 chicken_valuation.py --live-weight 7.0 --save-db
"""

import argparse
import json
import os
from datetime import date, datetime
from dataclasses import dataclass, field

from config import (
    DEFAULT_CHICKEN_LIVE_WEIGHT, DEFAULT_CHICKEN_DRESS_PCT,
    CHICKEN_CUT_YIELDS, REPORTS_DIR,
)
from manual_entry import load_prices


@dataclass
class ChickenValuation:
    live_weight: float
    dress_pct: float
    dressed_weight: float
    report_date: str
    cut_values: list = field(default_factory=list)
    total_cut_value: float = 0.0
    value_per_lb_live: float = 0.0
    value_per_bird: float = 0.0
    missing_cuts: list = field(default_factory=list)


def compute_chicken_value(
    live_weight: float = DEFAULT_CHICKEN_LIVE_WEIGHT,
    dress_pct: float = DEFAULT_CHICKEN_DRESS_PCT,
) -> ChickenValuation:
    dressed_wt = live_weight * dress_pct
    prices = load_prices('chicken')
    price_date = prices.get('date', date.today().isoformat())

    cut_values = []
    total = 0.0
    missing = []

    for code, (desc, yield_pct, category) in CHICKEN_CUT_YIELDS.items():
        entry = prices.get('cuts', {}).get(code, {})
        price_per_lb = entry.get('price_per_lb', 0)

        if price_per_lb <= 0:
            missing.append(code)
            continue

        cut_weight = dressed_wt * (yield_pct / 100.0)
        value = cut_weight * price_per_lb
        total += value

        cut_values.append({
            'cut_code': code,
            'description': desc,
            'category': category,
            'yield_pct': yield_pct,
            'cut_weight_lbs': round(cut_weight, 3),
            'price_per_lb': price_per_lb,
            'cut_value': round(value, 2),
        })

    return ChickenValuation(
        live_weight=live_weight,
        dress_pct=dress_pct,
        dressed_weight=round(dressed_wt, 2),
        report_date=price_date,
        cut_values=sorted(cut_values, key=lambda x: -x['cut_value']),
        total_cut_value=round(total, 2),
        value_per_lb_live=round(total / live_weight, 4) if live_weight > 0 else 0,
        value_per_bird=round(total, 2),
        missing_cuts=missing,
    )


def print_chicken_valuation(v: ChickenValuation):
    print(f"\n{'=' * 75}")
    print(f"  CHICKEN (BROILER) VALUATION")
    print(f"{'=' * 75}")
    print(f"  Price Date:       {v.report_date}")
    print(f"  Live Weight:      {v.live_weight:.1f} lbs")
    print(f"  Dressing %:       {v.dress_pct:.1%}")
    print(f"  Dressed Weight:   {v.dressed_weight:.2f} lbs")

    if v.missing_cuts:
        print(f"\n  WARNING: {len(v.missing_cuts)} cuts missing prices: "
              f"{', '.join(v.missing_cuts)}")
        print(f"  Run: python3 manual_entry.py chicken set-all")

    print(f"\n  CUT DETAIL")
    print(f"  {'─' * 70}")
    print(f"  {'Cut':<25} {'Category':<12} {'Yld%':>5} {'Wt(lb)':>7} "
          f"{'$/lb':>7} {'Value':>8}")
    print(f"  {'─' * 70}")

    for cut in v.cut_values:
        print(f"  {cut['description'][:24]:<25} {cut['category']:<12} "
              f"{cut['yield_pct']:>4.1f}% {cut['cut_weight_lbs']:>6.3f} "
              f"${cut['price_per_lb']:>5.2f} ${cut['cut_value']:>7.2f}")

    print(f"  {'─' * 70}")
    print(f"  {'TOTAL':<25} {'':12} {'':>5} {v.dressed_weight:>7.2f} "
          f"{'':>7} ${v.total_cut_value:>7.2f}")
    print()
    print(f"  Value per lb live:  ${v.value_per_lb_live:.4f}")
    print(f"  Value per bird:     ${v.value_per_bird:.2f}")
    print(f"{'=' * 75}")


def write_chicken_excel(valuation, filename=None):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("openpyxl not installed — skipping Excel output")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Chicken Valuation"

    header_fill = PatternFill(start_color="E2A037", end_color="E2A037", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    currency_fmt = '#,##0.00'

    headers = ["Cut", "Category", "Yield %", "Weight (lb)", "$/lb", "Value ($)"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 14
    for col in 'CDEF':
        ws.column_dimensions[col].width = 12

    for r, cut in enumerate(valuation.cut_values, 2):
        ws.cell(row=r, column=1, value=cut['description'])
        ws.cell(row=r, column=2, value=cut['category'])
        ws.cell(row=r, column=3, value=cut['yield_pct']).number_format = '0.0'
        ws.cell(row=r, column=4, value=cut['cut_weight_lbs']).number_format = '0.000'
        ws.cell(row=r, column=5, value=cut['price_per_lb']).number_format = currency_fmt
        ws.cell(row=r, column=6, value=cut['cut_value']).number_format = currency_fmt

    if filename is None:
        filename = os.path.join(REPORTS_DIR, f"chicken_valuation_{datetime.now().strftime('%Y%m%d')}.xlsx")
    wb.save(filename)
    print(f"\nExcel workbook saved to: {filename}")


def main():
    parser = argparse.ArgumentParser(description="Chicken Valuation Engine")
    parser.add_argument('--live-weight', type=float, default=DEFAULT_CHICKEN_LIVE_WEIGHT)
    parser.add_argument('--dress-pct', type=float, default=DEFAULT_CHICKEN_DRESS_PCT)
    parser.add_argument('--output', type=str, default=None)
    parser.add_argument('--save-db', action='store_true')
    args = parser.parse_args()

    valuation = compute_chicken_value(args.live_weight, args.dress_pct)
    print_chicken_valuation(valuation)
    write_chicken_excel(valuation, args.output)

    if args.save_db:
        try:
            from db import init_schema, save_valuation
            init_schema()
            save_valuation('chicken', valuation.report_date, {
                'live_weight': valuation.live_weight,
                'dressing_pct': valuation.dress_pct,
                'hot_carcass_weight': valuation.dressed_weight,
                'total_cut_value': valuation.total_cut_value,
                'byproduct_value': 0,
                'gross_value': valuation.total_cut_value,
                'processing_cost': 0,
                'net_value': valuation.total_cut_value,
                'value_per_lb_live': valuation.value_per_lb_live,
                'cut_detail': valuation.cut_values,
            })
            print("\nChicken data saved to PostgreSQL.")
        except Exception as e:
            print(f"\nDB save failed: {e}")


if __name__ == '__main__':
    main()
