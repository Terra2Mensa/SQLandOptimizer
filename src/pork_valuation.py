#!/usr/bin/env python3
"""
Pork Valuation Engine
=====================
Auto-pulls USDA pork cutout prices and live hog prices, computes
carcass value from live weight.

Data sources:
  - USDA DataMart 2498 (National Daily Pork FOB Plant - cut-level)
  - USDA DataMart 2510 (National Daily Direct Purchased Swine)

Usage:
  python3 pork_valuation.py
  python3 pork_valuation.py --live-weight 280 --save-db
  python3 pork_valuation.py --output pork_valuation.xlsx
"""

import argparse
import json
import os
from datetime import datetime
from dataclasses import dataclass, field

from config import (
    DATAMART_BASE_URL,
    REPORT_PORK_DAILY, REPORT_PORK_LIVE,
    DEFAULT_PORK_LIVE_WEIGHT, DEFAULT_PORK_DRESS_PCT,
    PORK_PRIMAL_ORDER, PORK_CUT_YIELDS, PORK_PROCESSORS,
    REPORTS_DIR,
)
from shared import fetch_datamart, parse_number


# ---------------------------------------------------------------------------
# Data classes
# ---------------------------------------------------------------------------

@dataclass
class PorkCutPrice:
    section: str
    description: str
    weighted_average: float
    price_range_low: float
    price_range_high: float
    total_pounds: int

    @property
    def price_per_lb(self) -> float:
        return self.weighted_average / 100.0


@dataclass
class PorkCarcassValuation:
    live_weight: float
    dress_pct: float
    hot_carcass_weight: float
    report_date: str
    primal_values: dict = field(default_factory=dict)
    cut_values: list = field(default_factory=list)
    total_cut_value: float = 0.0
    carcass_value_cwt: float = 0.0
    net_value: float = 0.0
    value_per_cwt_live: float = 0.0


@dataclass
class PorkPurchasePriceResult:
    live_weight: float
    dress_pct: float
    processor_name: str
    processor_costs: dict = field(default_factory=dict)
    live_basis_cwt: float = 0.0
    carcass_basis_cwt: float = 0.0
    cutout_minus_margin_cwt: float = 0.0


# ---------------------------------------------------------------------------
# Data fetching
# ---------------------------------------------------------------------------

def fetch_pork_cutout() -> dict:
    """Fetch National Daily Pork FOB Plant (2498) — cut-level prices."""
    print("Fetching USDA pork cutout (report 2498)...")

    data = fetch_datamart(REPORT_PORK_DAILY)
    result = {
        'primals': {},
        'cuts_by_section': {},
        'report_date': '',
    }

    cut_sections = {'Loin Cuts', 'Butt Cuts', 'Picnic Cuts', 'Ham Cuts',
                    'Belly Cuts', 'Sparerib Cuts', 'Jowl Cuts',
                    'Trim Cuts', 'Variety Cuts'}

    for section in data:
        sec_name = section.get('reportSection', '')
        items = section.get('results', [])
        if not items:
            continue

        result['report_date'] = items[0].get('report_date', '')

        if sec_name == 'Cutout and Primal Values':
            item = items[0]
            result['primals'] = {
                'carcass': parse_number(item.get('pork_carcass', '0')),
                'loin': parse_number(item.get('pork_loin', '0')),
                'butt': parse_number(item.get('pork_butt', '0')),
                'picnic': parse_number(item.get('pork_picnic', '0')),
                'rib': parse_number(item.get('pork_rib', '0')),
                'ham': parse_number(item.get('pork_ham', '0')),
                'belly': parse_number(item.get('pork_belly', '0')),
            }

        elif sec_name in cut_sections:
            primal = sec_name.replace(' Cuts', '')
            cuts = []
            for item in items:
                avg = parse_number(item.get('weighted_average', '0'))
                if avg <= 0:
                    continue
                cuts.append(PorkCutPrice(
                    section=primal,
                    description=item.get('Item_Description', ''),
                    weighted_average=avg,
                    price_range_low=parse_number(item.get('price_range_low', '0')),
                    price_range_high=parse_number(item.get('price_range_high', '0')),
                    total_pounds=int(parse_number(item.get('total_pounds', '0'))),
                ))
            result['cuts_by_section'][primal] = cuts

    total_cuts = sum(len(c) for c in result['cuts_by_section'].values())
    print(f"  Report date: {result['report_date']}")
    print(f"  Carcass cutout: ${result['primals'].get('carcass', 0):.2f}/cwt")
    print(f"  {total_cuts} individual cuts with prices across "
          f"{len(result['cuts_by_section'])} primals")

    return result


def fetch_pork_live() -> dict:
    """Fetch National Daily Direct Purchased Swine (2510)."""
    print("Fetching USDA live hog prices (report 2510)...")

    data = fetch_datamart(REPORT_PORK_LIVE)
    result = {'rows': [], 'report_date': '', 'carcass_basis_185': 0.0}

    for section in data:
        sec_name = section.get('reportSection', '')
        items = section.get('results', [])
        if not items:
            continue

        result['report_date'] = items[0].get('report_date', '')

        if sec_name == 'Barrows/Gilts (producer/packer sold)':
            for item in items:
                price = parse_number(item.get('weighted_avg_price', '0'))
                if price <= 0:
                    continue
                result['rows'].append({
                    'purchase_type': item.get('purchase_type_name', ''),
                    'head_count': int(parse_number(item.get('head_count', '0'))),
                    'avg_weight': parse_number(item.get('avg_carcass_weight', '0')),
                    'price_low': parse_number(item.get('price_range_low', '0')),
                    'price_high': parse_number(item.get('price_range_high', '0')),
                    'avg_price': price,
                    'carcass_basis': parse_number(item.get('avg_net_price', '0')),
                })

        elif sec_name == '185 lb Carcass Basis':
            for item in items:
                val = parse_number(item.get('base_price', '0'))
                if val > 0:
                    result['carcass_basis_185'] = val
                    break

    print(f"  Report date: {result['report_date']}")
    print(f"  {len(result['rows'])} price rows")
    if result['carcass_basis_185'] > 0:
        print(f"  185 lb carcass basis: ${result['carcass_basis_185']:.2f}/cwt")

    return result


# ---------------------------------------------------------------------------
# Valuation
# ---------------------------------------------------------------------------

def compute_pork_value(
    cutout_data: dict,
    live_weight: float = DEFAULT_PORK_LIVE_WEIGHT,
    dress_pct: float = DEFAULT_PORK_DRESS_PCT,
) -> PorkCarcassValuation:
    hcw = live_weight * dress_pct

    # Use primal cutout values to compute total carcass value
    primals = cutout_data.get('primals', {})
    carcass_cwt = primals.get('carcass', 0)
    total_cut_value = carcass_cwt / 100.0 * hcw

    # Build cut detail from individual section cuts
    cut_values = []
    for section, cuts in cutout_data.get('cuts_by_section', {}).items():
        for cut in cuts:
            cut_values.append({
                'section': section,
                'description': cut.description,
                'price_cwt': round(cut.weighted_average, 2),
                'price_per_lb': round(cut.price_per_lb, 2),
                'total_pounds': cut.total_pounds,
                'price_low': round(cut.price_range_low, 2),
                'price_high': round(cut.price_range_high, 2),
            })

    net_value = total_cut_value
    value_per_cwt_live = (net_value / live_weight) * 100 if live_weight > 0 else 0

    return PorkCarcassValuation(
        live_weight=live_weight,
        dress_pct=dress_pct,
        hot_carcass_weight=round(hcw, 1),
        report_date=cutout_data['report_date'],
        primal_values=primals,
        cut_values=cut_values,
        total_cut_value=round(total_cut_value, 2),
        carcass_value_cwt=round(carcass_cwt, 2),
        net_value=round(net_value, 2),
        value_per_cwt_live=round(value_per_cwt_live, 2),
    )


def compute_pork_purchase_price(
    valuation: PorkCarcassValuation,
    live_data: dict,
    processor: dict,
    live_weight: float,
) -> PorkPurchasePriceResult:
    hcw = valuation.hot_carcass_weight
    dress_pct = valuation.dress_pct

    # Method 1: Live basis — weighted avg of negotiated live prices
    live_rows = [r for r in live_data.get('rows', []) if r['avg_price'] > 0]
    if live_rows:
        total_head = sum(r['head_count'] for r in live_rows)
        if total_head > 0:
            live_basis = sum(r['avg_price'] * r['head_count'] for r in live_rows) / total_head
        else:
            live_basis = sum(r['avg_price'] for r in live_rows) / len(live_rows)
    else:
        live_basis = 0.0

    # Method 2: Carcass basis converted to live
    carcass_basis_185 = live_data.get('carcass_basis_185', 0)
    carcass_to_live = carcass_basis_185 * dress_pct if carcass_basis_185 > 0 else 0

    # Method 3: Cutout minus margin
    kill_fee = processor.get('kill_fee', 0)
    fab_per_lb = processor.get('fab_cost_per_lb', 0)
    shrink_pct = processor.get('shrink_pct', 0)

    fab_total = fab_per_lb * hcw
    shrink_cost = shrink_pct * valuation.total_cut_value
    remainder = valuation.net_value - kill_fee - fab_total - shrink_cost
    cutout_minus = (remainder / live_weight) * 100 if live_weight > 0 else 0

    return PorkPurchasePriceResult(
        live_weight=live_weight,
        dress_pct=dress_pct,
        processor_name=processor.get('name', 'Unknown'),
        processor_costs={
            'kill_fee': kill_fee,
            'fab_cost_per_lb': fab_per_lb,
            'fab_total': round(fab_total, 2),
            'shrink_pct': shrink_pct,
            'shrink_cost': round(shrink_cost, 2),
            'payment_terms_days': processor.get('payment_terms_days', 0),
        },
        live_basis_cwt=round(live_basis, 2),
        carcass_basis_cwt=round(carcass_to_live, 2),
        cutout_minus_margin_cwt=round(cutout_minus, 2),
    )


# ---------------------------------------------------------------------------
# Console output
# ---------------------------------------------------------------------------

def print_pork_valuation(valuation: PorkCarcassValuation):
    v = valuation
    print(f"\n{'=' * 80}")
    print(f"  PORK CARCASS VALUATION")
    print(f"{'=' * 80}")
    print(f"  Report Date:        {v.report_date}")
    print(f"  Live Weight:        {v.live_weight:.0f} lbs")
    print(f"  Dressing %:         {v.dress_pct:.1%}")
    print(f"  Hot Carcass Weight:  {v.hot_carcass_weight:.1f} lbs")
    print()
    print(f"  PRIMAL CUTOUT VALUES ($/cwt)")
    print(f"  {'─' * 50}")
    print(f"  {'Carcass Composite:':<25} ${v.primal_values.get('carcass', 0):>10.2f}")
    for primal in ['loin', 'butt', 'picnic', 'rib', 'ham', 'belly']:
        label = primal.title()
        val = v.primal_values.get(primal, 0)
        print(f"  {label + ':':<25} ${val:>10.2f}")

    print()
    print(f"  CARCASS VALUE")
    print(f"  {'─' * 50}")
    print(f"  {'Carcass cutout $/cwt:':<25} ${v.carcass_value_cwt:>10.2f}")
    print(f"  {'Total carcass value:':<25} ${v.total_cut_value:>10.2f}")
    print(f"  {'Value $/cwt live:':<25} ${v.value_per_cwt_live:>10.2f}")

    # Cut detail by section
    print(f"\n  CUT DETAIL BY PRIMAL")
    print(f"  {'─' * 74}")
    print(f"  {'Description':<45} {'$/cwt':>8} {'$/lb':>7} {'Volume':>12}")
    print(f"  {'─' * 74}")

    current_section = ''
    for cut in sorted(valuation.cut_values, key=lambda x: (
            PORK_PRIMAL_ORDER.index(x['section']) if x['section'] in PORK_PRIMAL_ORDER else 99,
            -x['price_cwt'])):
        if cut['section'] != current_section:
            current_section = cut['section']
            print(f"\n  {current_section.upper()}")
        lbs_str = f"{cut['total_pounds']:,}" if cut['total_pounds'] else ''
        print(f"  {cut['description'][:44]:<45} ${cut['price_cwt']:>7.2f} "
              f"${cut['price_per_lb']:>5.2f} {lbs_str:>12}")

    print(f"{'=' * 80}")


def print_pork_purchase_price(pp: PorkPurchasePriceResult):
    print(f"\n{'=' * 80}")
    print(f"  PORK PURCHASE PRICE ANALYSIS | Processor: {pp.processor_name}")
    print(f"{'=' * 80}")
    print(f"  Processor Costs:")
    print(f"    Kill fee:       ${pp.processor_costs['kill_fee']:.2f}/head")
    print(f"    Fab cost:       ${pp.processor_costs['fab_cost_per_lb']:.2f}/lb "
          f"(${pp.processor_costs['fab_total']:.2f} total)")
    print(f"    Cooler shrink:  {pp.processor_costs['shrink_pct']:.1%} "
          f"(${pp.processor_costs['shrink_cost']:.2f})")
    print(f"    Payment terms:  Net {pp.processor_costs['payment_terms_days']} days")
    print()
    print(f"  {'METHOD':<35} {'$/cwt Live':>12} {'$/head':>12}")
    print(f"  {'─' * 60}")
    methods = [
        ('1. Negotiated Live Basis', pp.live_basis_cwt),
        ('2. Carcass Basis (-> live)', pp.carcass_basis_cwt),
        ('3. Cutout-Minus-Margin', pp.cutout_minus_margin_cwt),
    ]
    vals = [m[1] for m in methods if m[1] > 0]
    for label, cwt in methods:
        per_head = cwt / 100 * pp.live_weight if cwt > 0 else 0
        if cwt > 0:
            print(f"  {label:<35} ${cwt:>10.2f}   ${per_head:>10.2f}")
        else:
            print(f"  {label:<35} {'N/A':>12}   {'N/A':>12}")

    if vals:
        spread = max(vals) - min(vals)
        avg = sum(vals) / len(vals)
        avg_head = avg / 100 * pp.live_weight
        print(f"  {'─' * 60}")
        print(f"  {'Spread (high - low):':<35} ${spread:>10.2f}   "
              f"${spread / 100 * pp.live_weight:>10.2f}")
        print(f"  {'Average of methods:':<35} ${avg:>10.2f}   ${avg_head:>10.2f}")
    print(f"{'=' * 80}")


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------

def write_pork_excel(valuation, purchase_price, cutout_data, live_data,
                     filename=None):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("openpyxl not installed — skipping Excel output")
        return

    wb = Workbook()
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF", size=11)
    currency_fmt = '#,##0.00'
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # --- Sheet 1: Summary ---
    ws = wb.active
    ws.title = "Pork Summary"
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 18

    rows = [
        ("PORK CARCASS VALUATION", ""),
        ("Report Date", valuation.report_date),
        ("Live Weight (lbs)", valuation.live_weight),
        ("Dressing %", f"{valuation.dress_pct:.1%}"),
        ("Hot Carcass Weight (lbs)", valuation.hot_carcass_weight),
        ("", ""),
        ("PRIMAL CUTOUT VALUES ($/cwt)", ""),
        ("Carcass Composite", valuation.primal_values.get('carcass', 0)),
        ("Loin", valuation.primal_values.get('loin', 0)),
        ("Butt", valuation.primal_values.get('butt', 0)),
        ("Picnic", valuation.primal_values.get('picnic', 0)),
        ("Rib", valuation.primal_values.get('rib', 0)),
        ("Ham", valuation.primal_values.get('ham', 0)),
        ("Belly", valuation.primal_values.get('belly', 0)),
        ("", ""),
        ("CARCASS VALUE", ""),
        ("Total Carcass Value ($)", valuation.total_cut_value),
        ("Value $/cwt Live", valuation.value_per_cwt_live),
    ]

    for r_idx, (label, val) in enumerate(rows, 1):
        ws.cell(row=r_idx, column=1, value=label).font = header_font if label.isupper() else Font()
        cell = ws.cell(row=r_idx, column=2, value=val)
        if isinstance(val, float):
            cell.number_format = currency_fmt

    # --- Sheet 2: Cut Detail ---
    ws2 = wb.create_sheet("Cut Detail")
    headers = ["Primal", "Description", "$/cwt", "$/lb", "Low", "High", "Volume (lbs)"]
    for c, h in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    ws2.column_dimensions['A'].width = 12
    ws2.column_dimensions['B'].width = 40
    for col in 'CDEFG':
        ws2.column_dimensions[col].width = 14

    row = 2
    for cut in sorted(valuation.cut_values, key=lambda x: (
            PORK_PRIMAL_ORDER.index(x['section']) if x['section'] in PORK_PRIMAL_ORDER else 99,
            -x['price_cwt'])):
        ws2.cell(row=row, column=1, value=cut['section'])
        ws2.cell(row=row, column=2, value=cut['description'])
        ws2.cell(row=row, column=3, value=cut['price_cwt']).number_format = currency_fmt
        ws2.cell(row=row, column=4, value=cut['price_per_lb']).number_format = currency_fmt
        ws2.cell(row=row, column=5, value=cut['price_low']).number_format = currency_fmt
        ws2.cell(row=row, column=6, value=cut['price_high']).number_format = currency_fmt
        ws2.cell(row=row, column=7, value=cut['total_pounds']).number_format = '#,##0'
        row += 1

    # --- Sheet 3: Purchase Price ---
    ws3 = wb.create_sheet("Purchase Price")
    ws3.column_dimensions['A'].width = 35
    ws3.column_dimensions['B'].width = 15
    ws3.column_dimensions['C'].width = 15

    pp = purchase_price
    pp_rows = [
        ("PURCHASE PRICE ANALYSIS", "", ""),
        ("Processor", pp.processor_name, ""),
        ("Kill Fee", pp.processor_costs['kill_fee'], ""),
        ("Fab Cost $/lb", pp.processor_costs['fab_cost_per_lb'], ""),
        ("Shrink %", f"{pp.processor_costs['shrink_pct']:.1%}", ""),
        ("", "", ""),
        ("Method", "$/cwt Live", "$/head"),
        ("Negotiated Live Basis", pp.live_basis_cwt,
         pp.live_basis_cwt / 100 * pp.live_weight if pp.live_basis_cwt > 0 else 0),
        ("Carcass Basis (-> live)", pp.carcass_basis_cwt,
         pp.carcass_basis_cwt / 100 * pp.live_weight if pp.carcass_basis_cwt > 0 else 0),
        ("Cutout-Minus-Margin", pp.cutout_minus_margin_cwt,
         pp.cutout_minus_margin_cwt / 100 * pp.live_weight if pp.cutout_minus_margin_cwt > 0 else 0),
    ]
    for r_idx, (a, b, c) in enumerate(pp_rows, 1):
        ws3.cell(row=r_idx, column=1, value=a)
        cell_b = ws3.cell(row=r_idx, column=2, value=b)
        cell_c = ws3.cell(row=r_idx, column=3, value=c)
        if isinstance(b, float):
            cell_b.number_format = currency_fmt
        if isinstance(c, float):
            cell_c.number_format = currency_fmt

    if filename is None:
        filename = os.path.join(REPORTS_DIR, f"pork_valuation_{datetime.now().strftime('%Y%m%d')}.xlsx")
    wb.save(filename)
    print(f"\nExcel workbook saved to: {filename}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Pork Valuation Engine")
    parser.add_argument('--live-weight', type=float, default=DEFAULT_PORK_LIVE_WEIGHT,
                        help=f'Live weight in lbs (default: {DEFAULT_PORK_LIVE_WEIGHT})')
    parser.add_argument('--dress-pct', type=float, default=DEFAULT_PORK_DRESS_PCT,
                        help=f'Dressing percentage (default: {DEFAULT_PORK_DRESS_PCT})')
    parser.add_argument('--output', type=str, default=None,
                        help='Excel output filename')
    parser.add_argument('--save-db', action='store_true',
                        help='Save results to PostgreSQL')
    parser.add_argument('--no-live', action='store_true',
                        help='Skip live hog price fetch')
    args = parser.parse_args()

    # Fetch data
    cutout_data = fetch_pork_cutout()
    live_data = {} if args.no_live else fetch_pork_live()

    # Compute valuation
    valuation = compute_pork_value(cutout_data, args.live_weight, args.dress_pct)
    print_pork_valuation(valuation)

    # Purchase price analysis
    processor = PORK_PROCESSORS.get('processor_a', {})
    purchase_price = None
    if live_data:
        purchase_price = compute_pork_purchase_price(
            valuation, live_data, processor, args.live_weight)
        print_pork_purchase_price(purchase_price)

    # Excel output
    write_pork_excel(valuation, purchase_price, cutout_data, live_data,
                     args.output)

    # DB persistence
    if args.save_db:
        try:
            from db import (init_schema, save_pork_cutout, save_pork_primals,
                            save_pork_live, save_valuation)
            init_schema()

            # Save cut-level data
            for section, cuts in cutout_data.get('cuts_by_section', {}).items():
                cut_rows = [{'description': c.description,
                             'weighted_average': c.weighted_average,
                             'price_range_low': c.price_range_low,
                             'price_range_high': c.price_range_high,
                             'total_pounds': c.total_pounds}
                            for c in cuts]
                save_pork_cutout(cutout_data['report_date'], REPORT_PORK_DAILY,
                                 section, cut_rows)

            # Save primal values
            save_pork_primals(cutout_data['report_date'], cutout_data['primals'])

            # Save live hog prices
            if live_data and live_data.get('rows'):
                save_pork_live(live_data['report_date'], REPORT_PORK_LIVE,
                               live_data['rows'])

            # Save valuation
            save_valuation('pork', cutout_data['report_date'], {
                'live_weight': valuation.live_weight,
                'dressing_pct': valuation.dress_pct,
                'hot_carcass_weight': valuation.hot_carcass_weight,
                'total_cut_value': valuation.total_cut_value,
                'byproduct_value': 0,
                'gross_value': valuation.total_cut_value,
                'processing_cost': 0,
                'net_value': valuation.net_value,
                'value_per_lb_live': valuation.value_per_cwt_live / 100,
                'cut_detail': valuation.cut_values,
            })

            print("\nPork data saved to PostgreSQL.")
        except Exception as e:
            print(f"\nDB save failed: {e}")


if __name__ == '__main__':
    main()
