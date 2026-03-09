#!/usr/bin/env python3
"""
Lamb Valuation Engine
=====================
Auto-pulls USDA lamb carcass cutout prices (with IMPS codes) and computes
carcass value from live weight.

Data sources:
  - USDA DataMart 2649 (National Estimated Lamb Carcass Cutout)
  - USDA DataMart 2648 (National 5-Day Rolling Average Boxed Lamb Cuts)

Usage:
  python3 lamb_valuation.py
  python3 lamb_valuation.py --live-weight 140 --save-db
  python3 lamb_valuation.py --output lamb_valuation.xlsx
"""

import argparse
import json
import os
from datetime import datetime
from dataclasses import dataclass, field

from config import (
    DATAMART_BASE_URL,
    REPORT_LAMB_CUTOUT, REPORT_LAMB_BOXED,
    DEFAULT_LAMB_LIVE_WEIGHT, DEFAULT_LAMB_DRESS_PCT,
    LAMB_PRIMAL_ORDER, LAMB_SUBPRIMAL_YIELDS, LAMB_PROCESSORS,
    REPORTS_DIR,
)
from shared import fetch_datamart, parse_number


# ---------------------------------------------------------------------------
# Data classes
# ---------------------------------------------------------------------------

@dataclass
class LambCutPrice:
    imps_code: str
    description: str
    fob_price: float          # $/cwt
    percentage_carcass: float  # % of carcass weight
    cut_weight: float          # lbs (based on reference carcass)
    saddle: str                # Foresaddle or Hindsaddle

    @property
    def price_per_lb(self) -> float:
        return self.fob_price / 100.0


@dataclass
class LambCarcassValuation:
    live_weight: float
    dress_pct: float
    hot_carcass_weight: float
    report_date: str
    cut_values: list = field(default_factory=list)
    total_cut_value: float = 0.0
    gross_carcass_price: float = 0.0
    net_carcass_price: float = 0.0
    foresaddle_price: float = 0.0
    hindsaddle_price: float = 0.0
    processing_cost: float = 0.0
    value_per_cwt_carcass: float = 0.0
    value_per_cwt_live: float = 0.0


@dataclass
class LambPurchasePriceResult:
    live_weight: float
    dress_pct: float
    processor_name: str
    processor_costs: dict = field(default_factory=dict)
    net_carcass_cwt: float = 0.0
    cutout_minus_margin_cwt: float = 0.0


# ---------------------------------------------------------------------------
# Data fetching
# ---------------------------------------------------------------------------

def fetch_lamb_cutout() -> dict:
    """Fetch National Estimated Lamb Carcass Cutout (2649)."""
    print("Fetching USDA lamb carcass cutout (report 2649)...")

    data = fetch_datamart(REPORT_LAMB_CUTOUT)
    result = {
        'cuts': [],
        'gross_carcass_price': 0,
        'foresaddle_price': 0,
        'hindsaddle_price': 0,
        'net_carcass_price': 0,
        'processing_cost': 0,
        'report_date': '',
    }

    current_saddle = ''

    for section in data:
        sec_name = section.get('reportSection', '')
        items = section.get('results', [])
        if not items:
            continue

        result['report_date'] = items[0].get('report_date', '')

        if sec_name == 'DETAIL':
            for item in items:
                imps = item.get('imps_code') or ''
                desc = item.get('imps_description') or ''
                fob = parse_number(item.get('fob_price', '0'))
                pct = parse_number(item.get('percentage_carcass', '0'))
                wt = parse_number(item.get('cut_weight', '0'))

                # Track which saddle we're in
                desc_upper = desc.upper().strip()
                if desc_upper == 'FORESADDLE':
                    current_saddle = 'Foresaddle'
                    continue
                elif desc_upper == 'HINDSADDLE':
                    current_saddle = 'Hindsaddle'
                    continue

                if not imps and not fob:
                    # Non-priced items like "SHRINK"
                    if 'FAT' in desc_upper or 'BONE' in desc_upper:
                        result['cuts'].append(LambCutPrice(
                            imps_code='', description=desc.strip(),
                            fob_price=fob, percentage_carcass=pct,
                            cut_weight=wt, saddle=current_saddle))
                    continue

                result['cuts'].append(LambCutPrice(
                    imps_code=imps.strip(),
                    description=desc.strip(),
                    fob_price=fob,
                    percentage_carcass=pct,
                    cut_weight=wt,
                    saddle=current_saddle,
                ))

        elif sec_name == 'GROSS CARCASS VALUE':
            for item in items:
                v = parse_number(item.get('gross_carcass_price', '0'))
                if v > 0:
                    result['gross_carcass_price'] = v

        elif sec_name == 'FORESADDLE VALUE':
            for item in items:
                v = parse_number(item.get('foresaddle_price', '0'))
                if v > 0:
                    result['foresaddle_price'] = v

        elif sec_name == 'HINDSADDLE VALUE':
            for item in items:
                v = parse_number(item.get('hindsaddle_price', '0'))
                if v > 0:
                    result['hindsaddle_price'] = v

        elif sec_name == 'NET CARCASS VALUE':
            for item in items:
                v = parse_number(item.get('net_carcass_price', '0'))
                if v > 0:
                    result['net_carcass_price'] = v

        elif sec_name == 'OTHER':
            for item in items:
                v = parse_number(item.get('processing_cost', '0'))
                if v > 0:
                    result['processing_cost'] = v

    priced = [c for c in result['cuts'] if c.fob_price > 0]
    print(f"  Report date: {result['report_date']}")
    print(f"  {len(priced)} priced cuts")
    print(f"  Gross carcass: ${result['gross_carcass_price']:.2f}/cwt")
    print(f"  Net carcass:   ${result['net_carcass_price']:.2f}/cwt")
    print(f"  Processing:    ${result['processing_cost']:.2f}/cwt")

    return result


# ---------------------------------------------------------------------------
# Valuation
# ---------------------------------------------------------------------------

def compute_lamb_value(
    cutout_data: dict,
    live_weight: float = DEFAULT_LAMB_LIVE_WEIGHT,
    dress_pct: float = DEFAULT_LAMB_DRESS_PCT,
) -> LambCarcassValuation:
    hcw = live_weight * dress_pct

    # Build cut detail — use API-provided percentage_carcass for yields
    cut_values = []
    total_cut_value = 0.0

    for cut in cutout_data['cuts']:
        if cut.fob_price <= 0:
            continue

        pct_carcass = cut.percentage_carcass
        cut_weight = hcw * (pct_carcass / 100.0)
        value = cut_weight * cut.price_per_lb
        total_cut_value += value

        # Map to primal
        primal = cut.saddle
        imps = cut.imps_code
        if imps in LAMB_SUBPRIMAL_YIELDS:
            primal = LAMB_SUBPRIMAL_YIELDS[imps][2]

        cut_values.append({
            'imps_code': cut.imps_code,
            'description': cut.description,
            'primal': primal,
            'saddle': cut.saddle,
            'fob_price_cwt': round(cut.fob_price, 2),
            'price_per_lb': round(cut.price_per_lb, 2),
            'yield_pct': pct_carcass,
            'cut_weight_lbs': round(cut_weight, 2),
            'cut_value': round(value, 2),
        })

    gross_cwt = cutout_data.get('gross_carcass_price', 0)
    net_cwt = cutout_data.get('net_carcass_price', 0)
    processing = cutout_data.get('processing_cost', 0)

    value_per_cwt_carcass = (total_cut_value / hcw) * 100 if hcw > 0 else 0
    value_per_cwt_live = (total_cut_value / live_weight) * 100 if live_weight > 0 else 0

    return LambCarcassValuation(
        live_weight=live_weight,
        dress_pct=dress_pct,
        hot_carcass_weight=round(hcw, 1),
        report_date=cutout_data['report_date'],
        cut_values=sorted(cut_values, key=lambda x: -x['cut_value']),
        total_cut_value=round(total_cut_value, 2),
        gross_carcass_price=gross_cwt,
        net_carcass_price=net_cwt,
        foresaddle_price=cutout_data.get('foresaddle_price', 0),
        hindsaddle_price=cutout_data.get('hindsaddle_price', 0),
        processing_cost=processing,
        value_per_cwt_carcass=round(value_per_cwt_carcass, 2),
        value_per_cwt_live=round(value_per_cwt_live, 2),
    )


def compute_lamb_purchase_price(
    valuation: LambCarcassValuation,
    processor: dict,
    live_weight: float,
) -> LambPurchasePriceResult:
    hcw = valuation.hot_carcass_weight
    dress_pct = valuation.dress_pct

    # Net carcass value from USDA (already deducts processing)
    net_cwt = valuation.net_carcass_price
    net_to_live = net_cwt * dress_pct if net_cwt > 0 else 0

    # Cutout minus our processor's margin
    kill_fee = processor.get('kill_fee', 0)
    fab_per_lb = processor.get('fab_cost_per_lb', 0)
    shrink_pct = processor.get('shrink_pct', 0)

    fab_total = fab_per_lb * hcw
    shrink_cost = shrink_pct * valuation.total_cut_value
    remainder = valuation.total_cut_value - kill_fee - fab_total - shrink_cost
    cutout_minus = (remainder / live_weight) * 100 if live_weight > 0 else 0

    return LambPurchasePriceResult(
        live_weight=live_weight,
        dress_pct=dress_pct,
        processor_name=processor.get('name', 'Unknown'),
        processor_costs={
            'kill_fee': kill_fee,
            'fab_cost_per_lb': fab_per_lb,
            'fab_total': round(fab_total, 2),
            'shrink_pct': shrink_pct,
            'shrink_cost': round(shrink_cost, 2),
            'payment_terms_days': processor.get('payment_terms_days', 0),
        },
        net_carcass_cwt=round(net_to_live, 2),
        cutout_minus_margin_cwt=round(cutout_minus, 2),
    )


# ---------------------------------------------------------------------------
# Console output
# ---------------------------------------------------------------------------

def print_lamb_valuation(valuation: LambCarcassValuation):
    v = valuation
    print(f"\n{'=' * 90}")
    print(f"  LAMB CARCASS VALUATION")
    print(f"{'=' * 90}")
    print(f"  Report Date:        {v.report_date}")
    print(f"  Live Weight:        {v.live_weight:.0f} lbs")
    print(f"  Dressing %:         {v.dress_pct:.1%}")
    print(f"  Hot Carcass Weight:  {v.hot_carcass_weight:.1f} lbs")
    print()
    print(f"  USDA CARCASS VALUES ($/cwt)")
    print(f"  {'─' * 50}")
    print(f"  {'Gross Carcass:':<25} ${v.gross_carcass_price:>10.2f}")
    print(f"  {'Foresaddle:':<25} ${v.foresaddle_price:>10.2f}")
    print(f"  {'Hindsaddle:':<25} ${v.hindsaddle_price:>10.2f}")
    print(f"  {'Processing Cost:':<25} ${v.processing_cost:>10.2f}")
    print(f"  {'Net Carcass:':<25} ${v.net_carcass_price:>10.2f}")

    print()
    print(f"  COMPUTED VALUES")
    print(f"  {'─' * 50}")
    print(f"  {'Total cut value ($):':<25} ${v.total_cut_value:>10.2f}")
    print(f"  {'Value $/cwt carcass:':<25} ${v.value_per_cwt_carcass:>10.2f}")
    print(f"  {'Value $/cwt live:':<25} ${v.value_per_cwt_live:>10.2f}")

    # Cut detail
    print(f"\n  CUT-LEVEL DETAIL")
    print(f"  {'─' * 85}")
    print(f"  {'IMPS':<8} {'Description':<35} {'Saddle':<12} {'$/cwt':>8} "
          f"{'Yld%':>6} {'Wt(lb)':>7} {'Value':>9}")
    print(f"  {'─' * 85}")

    for cut in v.cut_values:
        code = cut['imps_code'] or '---'
        print(f"  {code:<8} {cut['description'][:34]:<35} {cut['saddle']:<12} "
              f"${cut['fob_price_cwt']:>7.2f} {cut['yield_pct']:>5.1f}% "
              f"{cut['cut_weight_lbs']:>6.1f} ${cut['cut_value']:>8.2f}")

    print(f"{'=' * 90}")


def print_lamb_purchase_price(pp: LambPurchasePriceResult):
    print(f"\n{'=' * 80}")
    print(f"  LAMB PURCHASE PRICE ANALYSIS | Processor: {pp.processor_name}")
    print(f"{'=' * 80}")
    print(f"  Processor Costs:")
    print(f"    Kill fee:       ${pp.processor_costs['kill_fee']:.2f}/head")
    print(f"    Fab cost:       ${pp.processor_costs['fab_cost_per_lb']:.2f}/lb "
          f"(${pp.processor_costs['fab_total']:.2f} total)")
    print(f"    Cooler shrink:  {pp.processor_costs['shrink_pct']:.1%} "
          f"(${pp.processor_costs['shrink_cost']:.2f})")
    print(f"    Payment terms:  Net {pp.processor_costs['payment_terms_days']} days")
    print()
    print(f"  {'METHOD':<35} {'$/cwt Live':>12} {'$/head':>12}")
    print(f"  {'─' * 60}")

    methods = [
        ('1. USDA Net Carcass (-> live)', pp.net_carcass_cwt),
        ('2. Cutout-Minus-Margin', pp.cutout_minus_margin_cwt),
    ]
    vals = [m[1] for m in methods if m[1] > 0]

    for label, cwt in methods:
        per_head = cwt / 100 * pp.live_weight if cwt > 0 else 0
        if cwt > 0:
            print(f"  {label:<35} ${cwt:>10.2f}   ${per_head:>10.2f}")
        else:
            print(f"  {label:<35} {'N/A':>12}   {'N/A':>12}")

    if vals:
        avg = sum(vals) / len(vals)
        print(f"  {'─' * 60}")
        print(f"  {'Average:':<35} ${avg:>10.2f}   "
              f"${avg / 100 * pp.live_weight:>10.2f}")
    print(f"{'=' * 80}")


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------

def write_lamb_excel(valuation, purchase_price, cutout_data, filename=None):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("openpyxl not installed — skipping Excel output")
        return

    wb = Workbook()
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF", size=11)
    currency_fmt = '#,##0.00'

    # --- Sheet 1: Summary ---
    ws = wb.active
    ws.title = "Lamb Summary"
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 18

    rows = [
        ("LAMB CARCASS VALUATION", ""),
        ("Report Date", valuation.report_date),
        ("Live Weight (lbs)", valuation.live_weight),
        ("Dressing %", f"{valuation.dress_pct:.1%}"),
        ("Hot Carcass Weight (lbs)", valuation.hot_carcass_weight),
        ("", ""),
        ("USDA CARCASS VALUES ($/cwt)", ""),
        ("Gross Carcass", valuation.gross_carcass_price),
        ("Foresaddle", valuation.foresaddle_price),
        ("Hindsaddle", valuation.hindsaddle_price),
        ("Processing Cost", valuation.processing_cost),
        ("Net Carcass", valuation.net_carcass_price),
        ("", ""),
        ("COMPUTED VALUES", ""),
        ("Total Cut Value ($)", valuation.total_cut_value),
        ("Value $/cwt Carcass", valuation.value_per_cwt_carcass),
        ("Value $/cwt Live", valuation.value_per_cwt_live),
    ]

    for r_idx, (label, val) in enumerate(rows, 1):
        ws.cell(row=r_idx, column=1, value=label).font = header_font if label.isupper() else Font()
        cell = ws.cell(row=r_idx, column=2, value=val)
        if isinstance(val, float):
            cell.number_format = currency_fmt

    # --- Sheet 2: Cut Detail ---
    ws2 = wb.create_sheet("Cut Detail")
    headers = ["IMPS", "Description", "Saddle", "$/cwt", "$/lb",
               "Yield %", "Weight (lb)", "Value ($)"]
    for c, h in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    ws2.column_dimensions['A'].width = 8
    ws2.column_dimensions['B'].width = 35
    ws2.column_dimensions['C'].width = 12
    for col in 'DEFGH':
        ws2.column_dimensions[col].width = 12

    for r_idx, cut in enumerate(valuation.cut_values, 2):
        ws2.cell(row=r_idx, column=1, value=cut['imps_code'])
        ws2.cell(row=r_idx, column=2, value=cut['description'])
        ws2.cell(row=r_idx, column=3, value=cut['saddle'])
        ws2.cell(row=r_idx, column=4, value=cut['fob_price_cwt']).number_format = currency_fmt
        ws2.cell(row=r_idx, column=5, value=cut['price_per_lb']).number_format = currency_fmt
        ws2.cell(row=r_idx, column=6, value=cut['yield_pct']).number_format = '0.0'
        ws2.cell(row=r_idx, column=7, value=cut['cut_weight_lbs']).number_format = '0.0'
        ws2.cell(row=r_idx, column=8, value=cut['cut_value']).number_format = currency_fmt

    if filename is None:
        filename = os.path.join(REPORTS_DIR, f"lamb_valuation_{datetime.now().strftime('%Y%m%d')}.xlsx")
    wb.save(filename)
    print(f"\nExcel workbook saved to: {filename}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Lamb Valuation Engine")
    parser.add_argument('--live-weight', type=float, default=DEFAULT_LAMB_LIVE_WEIGHT,
                        help=f'Live weight in lbs (default: {DEFAULT_LAMB_LIVE_WEIGHT})')
    parser.add_argument('--dress-pct', type=float, default=DEFAULT_LAMB_DRESS_PCT,
                        help=f'Dressing percentage (default: {DEFAULT_LAMB_DRESS_PCT})')
    parser.add_argument('--output', type=str, default=None,
                        help='Excel output filename')
    parser.add_argument('--save-db', action='store_true',
                        help='Save results to PostgreSQL')
    args = parser.parse_args()

    # Fetch data
    cutout_data = fetch_lamb_cutout()

    # Compute valuation
    valuation = compute_lamb_value(cutout_data, args.live_weight, args.dress_pct)
    print_lamb_valuation(valuation)

    # Purchase price analysis
    processor = LAMB_PROCESSORS.get('processor_a', {})
    purchase_price = compute_lamb_purchase_price(
        valuation, processor, args.live_weight)
    print_lamb_purchase_price(purchase_price)

    # Excel output
    write_lamb_excel(valuation, purchase_price, cutout_data, args.output)

    # DB persistence
    if args.save_db:
        try:
            from db import (init_schema, save_lamb_cutout, save_lamb_summary,
                            save_valuation)
            init_schema()

            # Save cut-level data
            cut_rows = [{'imps_code': c.imps_code, 'description': c.description,
                         'fob_price': c.fob_price,
                         'percentage_carcass': c.percentage_carcass,
                         'cut_weight': c.cut_weight, 'saddle': c.saddle}
                        for c in cutout_data['cuts'] if c.fob_price > 0]
            save_lamb_cutout(cutout_data['report_date'], REPORT_LAMB_CUTOUT,
                             cut_rows)

            # Save summary
            save_lamb_summary(cutout_data['report_date'], {
                'gross': cutout_data['gross_carcass_price'],
                'foresaddle': cutout_data['foresaddle_price'],
                'hindsaddle': cutout_data['hindsaddle_price'],
                'net': cutout_data['net_carcass_price'],
                'processing_cost': cutout_data['processing_cost'],
            })

            # Save valuation
            save_valuation('lamb', cutout_data['report_date'], {
                'live_weight': valuation.live_weight,
                'dressing_pct': valuation.dress_pct,
                'hot_carcass_weight': valuation.hot_carcass_weight,
                'total_cut_value': valuation.total_cut_value,
                'byproduct_value': 0,
                'gross_value': valuation.total_cut_value,
                'processing_cost': valuation.processing_cost,
                'net_value': valuation.total_cut_value - valuation.processing_cost,
                'value_per_lb_live': valuation.value_per_cwt_live / 100,
                'cut_detail': valuation.cut_values,
            })

            print("\nLamb data saved to PostgreSQL.")
        except Exception as e:
            print(f"\nDB save failed: {e}")


if __name__ == '__main__':
    main()
