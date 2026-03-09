"""Flexible Budget Workbook — Multi-Species Meat Processing.

Generates a 9-tab Excel workbook (.xlsx) with:
- DB-seeded assumption cells (latest USDA valuations, config defaults)
- Excel formulas so assumptions can be tweaked manually after generation
- Monthly columns + quarterly subtotals + annual total

Usage:
    python3 budget_workbook.py                          # default 12 months from next month
    python3 budget_workbook.py --start-month 2026-04    # start from April 2026
    python3 budget_workbook.py --output ~/Desktop/budget.xlsx
"""

import argparse
import os
from datetime import datetime, date
from calendar import month_abbr

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

import db
from config import (
    REPORTS_DIR,
    DEFAULT_LIVE_WEIGHT, DEFAULT_PORK_LIVE_WEIGHT, DEFAULT_LAMB_LIVE_WEIGHT,
    DEFAULT_CHICKEN_LIVE_WEIGHT, DEFAULT_GOAT_LIVE_WEIGHT,
    DRESS_PCT_BY_YG, DEFAULT_PORK_DRESS_PCT, DEFAULT_LAMB_DRESS_PCT,
    DEFAULT_CHICKEN_DRESS_PCT, DEFAULT_GOAT_DRESS_PCT,
    PROCESSORS, DEFAULT_PROCESSOR,
    PORK_PROCESSORS, LAMB_PROCESSORS,
)

# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------
BOLD = Font(bold=True)
BOLD_UNDERLINE = Font(bold=True, underline="single")
HEADER_FONT = Font(bold=True, size=12)
SECTION_FONT = Font(bold=True, size=11)
CURRENCY_FMT = '"$"#,##0.00'
CURRENCY_FMT_NEG = '"$"#,##0.00;[Red]("$"#,##0.00)'
PCT_FMT = '0.00%'
NUMBER_FMT = '#,##0'
THIN_BORDER = Border(bottom=Side(style="thin"))
DOUBLE_BORDER = Border(bottom=Side(style="double"))
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

SPECIES = ["Cattle", "Pork", "Lamb", "Chicken", "Goat"]


def _col(c):
    """Return column letter for 1-based index."""
    return get_column_letter(c)


def _month_headers(start_month, start_year, months):
    """Return list of (label, month_num, year) for each column including Q subtotals and annual."""
    headers = []
    m, y = start_month, start_year
    q_count = 0
    for i in range(months):
        headers.append(("month", f"{month_abbr[m]} {y}", m, y))
        q_count += 1
        if q_count == 3:
            q_num = ((i // 3) + 1)
            headers.append(("quarter", f"Q{q_num}", None, None))
            q_count = 0
        m += 1
        if m > 12:
            m = 1
            y += 1
    # If months not divisible by 3, flush remaining quarter
    if q_count > 0:
        q_num = (months // 3) + 1
        headers.append(("quarter", f"Q{q_num}", None, None))
    headers.append(("annual", "Annual Total", None, None))
    return headers


def _write_period_headers(ws, row, start_col, headers):
    """Write month/quarter/annual headers and return column mapping."""
    for i, (htype, label, _, _) in enumerate(headers):
        col = start_col + i
        cell = ws.cell(row=row, column=col, value=label)
        cell.font = BOLD
        cell.alignment = Alignment(horizontal="center")
        cell.fill = HEADER_FILL
    return len(headers)


def _quarter_sum_formula(col_idx, month_cols_in_quarter, row):
    """Build SUM formula for a quarter column referencing preceding month columns."""
    refs = [f"{_col(c)}{row}" for c in month_cols_in_quarter]
    return f"={'+'.join(refs)}"


def _annual_sum_formula(all_month_cols, row):
    """Build SUM formula across all month columns for annual total."""
    refs = [f"{_col(c)}{row}" for c in all_month_cols]
    return f"={'+'.join(refs)}"


def _get_period_col_info(headers, start_col):
    """Return (month_cols, quarter_cols_with_months, annual_col)."""
    month_cols = []
    quarter_groups = []
    current_q_months = []
    annual_col = None
    for i, (htype, label, _, _) in enumerate(headers):
        col = start_col + i
        if htype == "month":
            month_cols.append(col)
            current_q_months.append(col)
        elif htype == "quarter":
            quarter_groups.append((col, list(current_q_months)))
            current_q_months = []
        elif htype == "annual":
            annual_col = col
    return month_cols, quarter_groups, annual_col


def _set_col_widths(ws, start_col, num_cols, width=14):
    """Set column widths for period columns."""
    for i in range(num_cols):
        ws.column_dimensions[_col(start_col + i)].width = width


# ---------------------------------------------------------------------------
# DB data helpers
# ---------------------------------------------------------------------------

def _get_latest_prices_by_species():
    """Pull latest prices from DB for seeding assumptions."""
    prices = {}
    try:
        conn = db.get_connection()
        try:
            import psycopg2.extras
            with conn.cursor(cursor_factory=psycopg2.extras.DictCursor) as cur:
                # Latest cattle valuation value_per_lb_live
                cur.execute("""
                    SELECT species, value_per_lb_live, value_per_lb_carcass,
                           live_weight, dressing_pct, processing_cost, hot_carcass_weight
                    FROM valuations
                    WHERE species = 'cattle'
                    ORDER BY valuation_date DESC LIMIT 1
                """)
                row = cur.fetchone()
                if row:
                    prices["Cattle"] = {
                        "sale_price_per_lb": float(row["value_per_lb_live"] or 0),
                        "carcass_price_per_lb": float(row["value_per_lb_carcass"] or 0),
                    }

                # Latest slaughter cattle price (live cost)
                cur.execute("""
                    SELECT weighted_avg_price FROM slaughter_cattle_prices
                    ORDER BY report_date DESC LIMIT 1
                """)
                row = cur.fetchone()
                if row:
                    prices.setdefault("Cattle", {})["live_cost_cwt"] = float(row["weighted_avg_price"] or 0)

                # Pork valuation
                cur.execute("""
                    SELECT species, value_per_lb_live, value_per_lb_carcass
                    FROM valuations WHERE species = 'pork'
                    ORDER BY valuation_date DESC LIMIT 1
                """)
                row = cur.fetchone()
                if row:
                    prices["Pork"] = {
                        "sale_price_per_lb": float(row["value_per_lb_live"] or 0),
                        "carcass_price_per_lb": float(row["value_per_lb_carcass"] or 0),
                    }

                # Pork live cost
                cur.execute("""
                    SELECT weighted_avg_price FROM pork_live_prices
                    ORDER BY report_date DESC LIMIT 1
                """)
                row = cur.fetchone()
                if row:
                    prices.setdefault("Pork", {})["live_cost_cwt"] = float(row["weighted_avg_price"] or 0)

                # Lamb valuation
                cur.execute("""
                    SELECT species, value_per_lb_live, value_per_lb_carcass
                    FROM valuations WHERE species = 'lamb'
                    ORDER BY valuation_date DESC LIMIT 1
                """)
                row = cur.fetchone()
                if row:
                    prices["Lamb"] = {
                        "sale_price_per_lb": float(row["value_per_lb_live"] or 0),
                        "carcass_price_per_lb": float(row["value_per_lb_carcass"] or 0),
                    }
        finally:
            conn.close()
    except Exception as e:
        print(f"Warning: Could not fetch DB prices: {e}")
        print("Budget will be generated with placeholder values.")

    # Manual species (chicken, goat)
    for species in ("chicken", "goat"):
        try:
            rows = db.get_latest_manual_prices(species)
            if rows:
                avg_price = sum(float(r["price_per_lb"]) for r in rows) / len(rows)
                prices[species.title()] = {"sale_price_per_lb": avg_price}
        except Exception:
            pass

    return prices


def _species_defaults():
    """Return dict of default weights/dress pcts/processing costs per species."""
    beef_proc = PROCESSORS[DEFAULT_PROCESSOR]
    pork_proc = PORK_PROCESSORS.get("processor_a", {})
    lamb_proc = LAMB_PROCESSORS.get("processor_a", {})

    return {
        "Cattle": {
            "live_weight": DEFAULT_LIVE_WEIGHT,
            "dress_pct": DRESS_PCT_BY_YG.get(3, 0.60),
            "carcass_weight": DEFAULT_LIVE_WEIGHT * DRESS_PCT_BY_YG.get(3, 0.60),
            "kill_fee": beef_proc.get("kill_fee", 175),
            "fab_cost_per_lb": beef_proc.get("fab_cost_per_lb", 0.22),
            "shrink_pct": beef_proc.get("shrink_pct", 0.025),
        },
        "Pork": {
            "live_weight": DEFAULT_PORK_LIVE_WEIGHT,
            "dress_pct": DEFAULT_PORK_DRESS_PCT,
            "carcass_weight": DEFAULT_PORK_LIVE_WEIGHT * DEFAULT_PORK_DRESS_PCT,
            "kill_fee": pork_proc.get("kill_fee", 45),
            "fab_cost_per_lb": pork_proc.get("fab_cost_per_lb", 0.12),
            "shrink_pct": pork_proc.get("shrink_pct", 0.02),
        },
        "Lamb": {
            "live_weight": DEFAULT_LAMB_LIVE_WEIGHT,
            "dress_pct": DEFAULT_LAMB_DRESS_PCT,
            "carcass_weight": DEFAULT_LAMB_LIVE_WEIGHT * DEFAULT_LAMB_DRESS_PCT,
            "kill_fee": lamb_proc.get("kill_fee", 35),
            "fab_cost_per_lb": lamb_proc.get("fab_cost_per_lb", 0.15),
            "shrink_pct": lamb_proc.get("shrink_pct", 0.02),
        },
        "Chicken": {
            "live_weight": DEFAULT_CHICKEN_LIVE_WEIGHT,
            "dress_pct": DEFAULT_CHICKEN_DRESS_PCT,
            "carcass_weight": DEFAULT_CHICKEN_LIVE_WEIGHT * DEFAULT_CHICKEN_DRESS_PCT,
            "kill_fee": 2.50,
            "fab_cost_per_lb": 0.10,
            "shrink_pct": 0.015,
        },
        "Goat": {
            "live_weight": DEFAULT_GOAT_LIVE_WEIGHT,
            "dress_pct": DEFAULT_GOAT_DRESS_PCT,
            "carcass_weight": DEFAULT_GOAT_LIVE_WEIGHT * DEFAULT_GOAT_DRESS_PCT,
            "kill_fee": 30.00,
            "fab_cost_per_lb": 0.15,
            "shrink_pct": 0.02,
        },
    }


# ---------------------------------------------------------------------------
# Tab builders
# ---------------------------------------------------------------------------

def _create_beginning_balance_sheet(wb):
    """Tab 1: Beginning Balance Sheet — blank template with total formulas."""
    ws = wb.create_sheet("Beginning Balance Sheet")
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 18

    ws.cell(row=1, column=1, value="Beginning Balance Sheet").font = HEADER_FONT

    # Assets
    r = 3
    ws.cell(row=r, column=1, value="ASSETS").font = SECTION_FONT
    r += 1
    ws.cell(row=r, column=1, value="Current Assets").font = BOLD
    current_assets_start = r + 1
    items = ["Cash", "Accounts Receivable", "Inventory — Live Animals",
             "Inventory — Finished Goods", "Prepaid Expenses"]
    for item in items:
        r += 1
        ws.cell(row=r, column=1, value=f"  {item}")
        ws.cell(row=r, column=2).number_format = CURRENCY_FMT
    current_assets_end = r
    r += 1
    ws.cell(row=r, column=1, value="TOTAL CURRENT ASSETS").font = BOLD
    ws.cell(row=r, column=1).border = THIN_BORDER
    ws.cell(row=r, column=2, value=f"=SUM(B{current_assets_start}:B{current_assets_end})")
    ws.cell(row=r, column=2).number_format = CURRENCY_FMT
    ws.cell(row=r, column=2).font = BOLD
    total_ca_row = r

    r += 2
    ws.cell(row=r, column=1, value="Property, Plant & Equipment").font = BOLD
    ppe_start = r + 1
    items = ["Land", "Buildings", "Equipment", "Vehicles",
             "Less: Accumulated Depreciation"]
    for item in items:
        r += 1
        ws.cell(row=r, column=1, value=f"  {item}")
        ws.cell(row=r, column=2).number_format = CURRENCY_FMT
    ppe_end = r
    r += 1
    ws.cell(row=r, column=1, value="TOTAL PPE (Net)").font = BOLD
    ws.cell(row=r, column=1).border = THIN_BORDER
    ws.cell(row=r, column=2, value=f"=SUM(B{ppe_start}:B{ppe_end})")
    ws.cell(row=r, column=2).number_format = CURRENCY_FMT
    ws.cell(row=r, column=2).font = BOLD
    total_ppe_row = r

    r += 2
    ws.cell(row=r, column=1, value="TOTAL ASSETS").font = BOLD_UNDERLINE
    ws.cell(row=r, column=1).border = DOUBLE_BORDER
    ws.cell(row=r, column=2, value=f"=B{total_ca_row}+B{total_ppe_row}")
    ws.cell(row=r, column=2).number_format = CURRENCY_FMT
    ws.cell(row=r, column=2).font = BOLD
    total_assets_row = r

    # Liabilities
    r += 2
    ws.cell(row=r, column=1, value="LIABILITIES").font = SECTION_FONT
    r += 1
    ws.cell(row=r, column=1, value="Current Liabilities").font = BOLD
    cl_start = r + 1
    items = ["Accounts Payable", "Accrued Expenses", "Current Portion of Long-Term Debt",
             "Income Taxes Payable"]
    for item in items:
        r += 1
        ws.cell(row=r, column=1, value=f"  {item}")
        ws.cell(row=r, column=2).number_format = CURRENCY_FMT
    cl_end = r
    r += 1
    ws.cell(row=r, column=1, value="TOTAL CURRENT LIABILITIES").font = BOLD
    ws.cell(row=r, column=1).border = THIN_BORDER
    ws.cell(row=r, column=2, value=f"=SUM(B{cl_start}:B{cl_end})")
    ws.cell(row=r, column=2).number_format = CURRENCY_FMT
    ws.cell(row=r, column=2).font = BOLD
    total_cl_row = r

    r += 2
    ws.cell(row=r, column=1, value="Long-Term Liabilities").font = BOLD
    lt_start = r + 1
    items = ["Long-Term Debt", "Notes Payable"]
    for item in items:
        r += 1
        ws.cell(row=r, column=1, value=f"  {item}")
        ws.cell(row=r, column=2).number_format = CURRENCY_FMT
    lt_end = r
    r += 1
    ws.cell(row=r, column=1, value="TOTAL LONG-TERM LIABILITIES").font = BOLD
    ws.cell(row=r, column=1).border = THIN_BORDER
    ws.cell(row=r, column=2, value=f"=SUM(B{lt_start}:B{lt_end})")
    ws.cell(row=r, column=2).number_format = CURRENCY_FMT
    ws.cell(row=r, column=2).font = BOLD
    total_lt_row = r

    r += 2
    ws.cell(row=r, column=1, value="TOTAL LIABILITIES").font = BOLD
    ws.cell(row=r, column=2, value=f"=B{total_cl_row}+B{total_lt_row}")
    ws.cell(row=r, column=2).number_format = CURRENCY_FMT
    ws.cell(row=r, column=2).font = BOLD
    total_liab_row = r

    # Equity
    r += 2
    ws.cell(row=r, column=1, value="STOCKHOLDERS' EQUITY").font = SECTION_FONT
    eq_start = r + 1
    items = ["Common Stock", "Retained Earnings"]
    for item in items:
        r += 1
        ws.cell(row=r, column=1, value=f"  {item}")
        ws.cell(row=r, column=2).number_format = CURRENCY_FMT
    eq_end = r
    r += 1
    ws.cell(row=r, column=1, value="TOTAL STOCKHOLDERS' EQUITY").font = BOLD
    ws.cell(row=r, column=1).border = THIN_BORDER
    ws.cell(row=r, column=2, value=f"=SUM(B{eq_start}:B{eq_end})")
    ws.cell(row=r, column=2).number_format = CURRENCY_FMT
    ws.cell(row=r, column=2).font = BOLD
    total_eq_row = r

    r += 2
    ws.cell(row=r, column=1, value="TOTAL LIABILITIES + EQUITY").font = BOLD_UNDERLINE
    ws.cell(row=r, column=1).border = DOUBLE_BORDER
    ws.cell(row=r, column=2, value=f"=B{total_liab_row}+B{total_eq_row}")
    ws.cell(row=r, column=2).number_format = CURRENCY_FMT
    ws.cell(row=r, column=2).font = BOLD


def _create_budget_data(wb, db_prices, defaults):
    """Tab 2: Budget Data (Assumptions) — seeded from DB where available."""
    ws = wb.create_sheet("Budget Data")
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 16
    for i, sp in enumerate(SPECIES):
        ws.column_dimensions[_col(i + 2)].width = 16

    ws.cell(row=1, column=1, value="Budget Data — Assumptions").font = HEADER_FONT

    # Species headers
    for i, sp in enumerate(SPECIES):
        ws.cell(row=2, column=i + 2, value=sp).font = BOLD
        ws.cell(row=2, column=i + 2).alignment = Alignment(horizontal="center")
        ws.cell(row=2, column=i + 2).fill = HEADER_FILL

    r = 4
    # --- Revenue Assumptions ---
    ws.cell(row=r, column=1, value="REVENUE ASSUMPTIONS").font = SECTION_FONT
    r += 1
    ws.cell(row=r, column=1, value="Expected units sold/month (head)")
    for i, sp in enumerate(SPECIES):
        ws.cell(row=r, column=i + 2, value=0).number_format = NUMBER_FMT
    units_row = r  # row 5

    r += 1
    ws.cell(row=r, column=1, value="Avg sale price per lb (live)")
    for i, sp in enumerate(SPECIES):
        price = db_prices.get(sp, {}).get("sale_price_per_lb", 0)
        ws.cell(row=r, column=i + 2, value=round(price, 4)).number_format = CURRENCY_FMT
    sale_price_row = r  # row 6

    r += 1
    ws.cell(row=r, column=1, value="Avg carcass weight (lbs)")
    for i, sp in enumerate(SPECIES):
        ws.cell(row=r, column=i + 2, value=round(defaults[sp]["carcass_weight"], 1)).number_format = NUMBER_FMT
    carcass_wt_row = r  # row 7

    r += 1
    ws.cell(row=r, column=1, value="Avg live weight (lbs)")
    for i, sp in enumerate(SPECIES):
        ws.cell(row=r, column=i + 2, value=defaults[sp]["live_weight"]).number_format = NUMBER_FMT
    live_wt_row = r  # row 8

    r += 1
    ws.cell(row=r, column=1, value="Dressing %")
    for i, sp in enumerate(SPECIES):
        ws.cell(row=r, column=i + 2, value=defaults[sp]["dress_pct"]).number_format = PCT_FMT
    dress_pct_row = r  # row 9

    r += 2
    ws.cell(row=r, column=1, value="Cash sales %")
    ws.cell(row=r, column=2, value=0.30).number_format = PCT_FMT
    cash_pct_row = r  # row 11

    r += 1
    ws.cell(row=r, column=1, value="Credit sales %")
    ws.cell(row=r, column=2, value=f"=1-B{cash_pct_row}").number_format = PCT_FMT
    credit_pct_row = r

    r += 1
    ws.cell(row=r, column=1, value="Collections: month of sale %")
    ws.cell(row=r, column=2, value=0.60).number_format = PCT_FMT
    collect_m0_row = r  # row 13

    r += 1
    ws.cell(row=r, column=1, value="Collections: month + 1 %")
    ws.cell(row=r, column=2, value=0.30).number_format = PCT_FMT
    collect_m1_row = r

    r += 1
    ws.cell(row=r, column=1, value="Collections: month + 2 %")
    ws.cell(row=r, column=2, value=0.10).number_format = PCT_FMT
    collect_m2_row = r

    # --- COGS Assumptions ---
    r += 2
    ws.cell(row=r, column=1, value="COGS ASSUMPTIONS").font = SECTION_FONT
    r += 1
    ws.cell(row=r, column=1, value="Live animal cost per head ($)")
    for i, sp in enumerate(SPECIES):
        cost_cwt = db_prices.get(sp, {}).get("live_cost_cwt", 0)
        if cost_cwt:
            cost_per_head = cost_cwt * defaults[sp]["live_weight"] / 100
        else:
            cost_per_head = 0
        ws.cell(row=r, column=i + 2, value=round(cost_per_head, 2)).number_format = CURRENCY_FMT
    animal_cost_row = r  # row 19

    # --- Processing Cost Assumptions ---
    r += 2
    ws.cell(row=r, column=1, value="PROCESSING COST ASSUMPTIONS").font = SECTION_FONT
    r += 1
    ws.cell(row=r, column=1, value="Kill fee per head ($)")
    for i, sp in enumerate(SPECIES):
        ws.cell(row=r, column=i + 2, value=defaults[sp]["kill_fee"]).number_format = CURRENCY_FMT
    kill_fee_row = r  # row 22

    r += 1
    ws.cell(row=r, column=1, value="Fabrication cost per lb ($)")
    for i, sp in enumerate(SPECIES):
        ws.cell(row=r, column=i + 2, value=defaults[sp]["fab_cost_per_lb"]).number_format = CURRENCY_FMT
    fab_cost_row = r  # row 23

    r += 1
    ws.cell(row=r, column=1, value="Shrink %")
    for i, sp in enumerate(SPECIES):
        ws.cell(row=r, column=i + 2, value=defaults[sp]["shrink_pct"]).number_format = PCT_FMT
    shrink_row = r  # row 24

    # --- Selling & Admin Assumptions ---
    r += 2
    ws.cell(row=r, column=1, value="SELLING & ADMIN ASSUMPTIONS").font = SECTION_FONT
    r += 1
    ws.cell(row=r, column=1, value="Variable S&A rate (% of sales)")
    ws.cell(row=r, column=2, value=0.05).number_format = PCT_FMT
    var_sa_row = r  # row 27

    r += 1
    ws.cell(row=r, column=1, value="Commissions (% of sales)")
    ws.cell(row=r, column=2, value=0.02).number_format = PCT_FMT

    r += 1
    ws.cell(row=r, column=1, value="Delivery cost (% of sales)")
    ws.cell(row=r, column=2, value=0.02).number_format = PCT_FMT

    r += 1
    ws.cell(row=r, column=1, value="Packaging (% of sales)")
    ws.cell(row=r, column=2, value=0.01).number_format = PCT_FMT

    r += 2
    ws.cell(row=r, column=1, value="Fixed monthly costs:").font = BOLD
    r += 1
    ws.cell(row=r, column=1, value="  Rent")
    ws.cell(row=r, column=2, value=0).number_format = CURRENCY_FMT
    rent_row = r  # row 33

    r += 1
    ws.cell(row=r, column=1, value="  Insurance")
    ws.cell(row=r, column=2, value=0).number_format = CURRENCY_FMT
    insurance_row = r

    r += 1
    ws.cell(row=r, column=1, value="  Utilities")
    ws.cell(row=r, column=2, value=0).number_format = CURRENCY_FMT
    utilities_row = r

    r += 1
    ws.cell(row=r, column=1, value="  Salaries & Wages")
    ws.cell(row=r, column=2, value=0).number_format = CURRENCY_FMT
    salaries_row = r

    r += 1
    ws.cell(row=r, column=1, value="  Equipment Lease")
    ws.cell(row=r, column=2, value=0).number_format = CURRENCY_FMT
    equip_row = r

    r += 1
    ws.cell(row=r, column=1, value="  Other Fixed")
    ws.cell(row=r, column=2, value=0).number_format = CURRENCY_FMT
    other_fixed_row = r

    r += 1
    ws.cell(row=r, column=1, value="TOTAL FIXED MONTHLY COSTS").font = BOLD
    ws.cell(row=r, column=1).border = THIN_BORDER
    ws.cell(row=r, column=2, value=f"=SUM(B{rent_row}:B{other_fixed_row})").number_format = CURRENCY_FMT
    ws.cell(row=r, column=2).font = BOLD
    total_fixed_row = r  # row 41

    r += 1
    ws.cell(row=r, column=1, value="  Depreciation per month")
    ws.cell(row=r, column=2, value=0).number_format = CURRENCY_FMT
    depreciation_row = r  # row 42

    # --- Other Assumptions ---
    r += 2
    ws.cell(row=r, column=1, value="OTHER ASSUMPTIONS").font = SECTION_FONT
    r += 1
    ws.cell(row=r, column=1, value="Tax rate")
    ws.cell(row=r, column=2, value=0.21).number_format = PCT_FMT
    tax_rate_row = r  # row 45

    r += 1
    ws.cell(row=r, column=1, value="Interest rate on debt (annual)")
    ws.cell(row=r, column=2, value=0.06).number_format = PCT_FMT
    interest_rate_row = r

    r += 1
    ws.cell(row=r, column=1, value="Minimum cash balance")
    ws.cell(row=r, column=2, value=10000).number_format = CURRENCY_FMT
    min_cash_row = r

    r += 1
    ws.cell(row=r, column=1, value="Purchases paid current month %")
    ws.cell(row=r, column=2, value=0.50).number_format = PCT_FMT
    purch_pay_pct_row = r  # row 48

    r += 1
    ws.cell(row=r, column=1, value="Purchases paid next month %")
    ws.cell(row=r, column=2, value=f"=1-B{purch_pay_pct_row}").number_format = PCT_FMT

    r += 1
    ws.cell(row=r, column=1, value="Capital expenditures per month")
    ws.cell(row=r, column=2, value=0).number_format = CURRENCY_FMT
    capex_row = r  # row 50

    r += 1
    ws.cell(row=r, column=1, value="Dividends per month")
    ws.cell(row=r, column=2, value=0).number_format = CURRENCY_FMT
    dividends_row = r  # row 51

    # Store key row references for other tabs
    ws.sheet_properties.tabColor = "1F4E79"

    # Return a dict of row references for other tabs to use
    return {
        "units_row": units_row, "sale_price_row": sale_price_row,
        "carcass_wt_row": carcass_wt_row, "live_wt_row": live_wt_row,
        "dress_pct_row": dress_pct_row,
        "cash_pct_row": cash_pct_row,
        "collect_m0_row": collect_m0_row, "collect_m1_row": collect_m1_row,
        "collect_m2_row": collect_m2_row,
        "animal_cost_row": animal_cost_row,
        "kill_fee_row": kill_fee_row, "fab_cost_row": fab_cost_row,
        "shrink_row": shrink_row,
        "var_sa_row": var_sa_row, "total_fixed_row": total_fixed_row,
        "depreciation_row": depreciation_row,
        "tax_rate_row": tax_rate_row, "interest_rate_row": interest_rate_row,
        "min_cash_row": min_cash_row, "purch_pay_pct_row": purch_pay_pct_row,
        "capex_row": capex_row, "dividends_row": dividends_row,
        "rent_row": rent_row, "other_fixed_row": other_fixed_row,
    }


def _create_sales_budget(wb, headers, start_col, refs):
    """Tab 3: Sales Budget — revenue by species + cash collections."""
    ws = wb.create_sheet("Sales Budget")
    ws.column_dimensions["A"].width = 35
    BD = "'Budget Data'"

    num_cols = _write_period_headers(ws, 2, start_col, headers)
    _set_col_widths(ws, start_col, num_cols)
    month_cols, quarter_groups, annual_col = _get_period_col_info(headers, start_col)

    # Part 1: Sales Revenue
    ws.cell(row=1, column=1, value="Sales Budget").font = HEADER_FONT
    ws.cell(row=3, column=1, value="PART 1: SALES REVENUE").font = SECTION_FONT

    species_rev_rows = {}
    r = 4
    for si, sp in enumerate(SPECIES):
        ws.cell(row=r, column=1, value=f"  {sp} Revenue")
        sp_col = _col(si + 2)  # species column in Budget Data
        for mc in month_cols:
            # units * live_weight * sale_price_per_lb
            ws.cell(row=r, column=mc,
                    value=f"={BD}!{sp_col}{refs['units_row']}*{BD}!{sp_col}{refs['live_wt_row']}*{BD}!{sp_col}{refs['sale_price_row']}")
            ws.cell(row=r, column=mc).number_format = CURRENCY_FMT
        # Quarter sums
        for qcol, qmonths in quarter_groups:
            ws.cell(row=r, column=qcol, value=_quarter_sum_formula(qcol, qmonths, r))
            ws.cell(row=r, column=qcol).number_format = CURRENCY_FMT
            ws.cell(row=r, column=qcol).font = BOLD
        # Annual
        ws.cell(row=r, column=annual_col, value=_annual_sum_formula(month_cols, r))
        ws.cell(row=r, column=annual_col).number_format = CURRENCY_FMT
        ws.cell(row=r, column=annual_col).font = BOLD
        species_rev_rows[sp] = r
        r += 1

    # Total Revenue
    ws.cell(row=r, column=1, value="TOTAL REVENUE").font = BOLD
    ws.cell(row=r, column=1).border = THIN_BORDER
    first_sp_row = species_rev_rows[SPECIES[0]]
    last_sp_row = species_rev_rows[SPECIES[-1]]
    for c in month_cols + [qc for qc, _ in quarter_groups] + [annual_col]:
        ws.cell(row=r, column=c,
                value=f"=SUM({_col(c)}{first_sp_row}:{_col(c)}{last_sp_row})")
        ws.cell(row=r, column=c).number_format = CURRENCY_FMT
        ws.cell(row=r, column=c).font = BOLD
        ws.cell(row=r, column=c).border = THIN_BORDER
    total_rev_row = r

    # Part 2: Cash Collections
    r += 2
    ws.cell(row=r, column=1, value="PART 2: CASH COLLECTIONS").font = SECTION_FONT
    r += 1
    ws.cell(row=r, column=1, value="  Cash Sales")
    cash_sales_row = r
    for mc in month_cols:
        ws.cell(row=r, column=mc,
                value=f"={_col(mc)}{total_rev_row}*{BD}!B{refs['cash_pct_row']}")
        ws.cell(row=r, column=mc).number_format = CURRENCY_FMT

    r += 1
    ws.cell(row=r, column=1, value="  Credit — collected month of sale")
    credit_m0_row = r
    for mc in month_cols:
        ws.cell(row=r, column=mc,
                value=f"={_col(mc)}{total_rev_row}*(1-{BD}!B{refs['cash_pct_row']})*{BD}!B{refs['collect_m0_row']}")
        ws.cell(row=r, column=mc).number_format = CURRENCY_FMT

    r += 1
    ws.cell(row=r, column=1, value="  Credit — collected month + 1")
    credit_m1_row = r
    for idx, mc in enumerate(month_cols):
        if idx == 0:
            ws.cell(row=r, column=mc, value=0).number_format = CURRENCY_FMT
        else:
            prev_mc = month_cols[idx - 1]
            ws.cell(row=r, column=mc,
                    value=f"={_col(prev_mc)}{total_rev_row}*(1-{BD}!B{refs['cash_pct_row']})*{BD}!B{refs['collect_m1_row']}")
            ws.cell(row=r, column=mc).number_format = CURRENCY_FMT

    r += 1
    ws.cell(row=r, column=1, value="  Credit — collected month + 2")
    credit_m2_row = r
    for idx, mc in enumerate(month_cols):
        if idx <= 1:
            ws.cell(row=r, column=mc, value=0).number_format = CURRENCY_FMT
        else:
            prev2_mc = month_cols[idx - 2]
            ws.cell(row=r, column=mc,
                    value=f"={_col(prev2_mc)}{total_rev_row}*(1-{BD}!B{refs['cash_pct_row']})*{BD}!B{refs['collect_m2_row']}")
            ws.cell(row=r, column=mc).number_format = CURRENCY_FMT

    r += 1
    ws.cell(row=r, column=1, value="TOTAL CASH RECEIPTS").font = BOLD
    ws.cell(row=r, column=1).border = THIN_BORDER
    total_cash_receipts_row = r
    for mc in month_cols:
        ws.cell(row=r, column=mc,
                value=f"=SUM({_col(mc)}{cash_sales_row}:{_col(mc)}{credit_m2_row})")
        ws.cell(row=r, column=mc).number_format = CURRENCY_FMT
        ws.cell(row=r, column=mc).font = BOLD
    for qcol, qmonths in quarter_groups:
        ws.cell(row=r, column=qcol, value=_quarter_sum_formula(qcol, qmonths, r))
        ws.cell(row=r, column=qcol).number_format = CURRENCY_FMT
        ws.cell(row=r, column=qcol).font = BOLD
    ws.cell(row=r, column=annual_col, value=_annual_sum_formula(month_cols, r))
    ws.cell(row=r, column=annual_col).number_format = CURRENCY_FMT
    ws.cell(row=r, column=annual_col).font = BOLD

    return {"total_rev_row": total_rev_row, "total_cash_receipts_row": total_cash_receipts_row,
            "species_rev_rows": species_rev_rows}


def _create_purchases_budget(wb, headers, start_col, refs):
    """Tab 4: Purchases Budget (COGS) — animal + processing costs + cash payments."""
    ws = wb.create_sheet("Purchases Budget")
    ws.column_dimensions["A"].width = 35
    BD = "'Budget Data'"

    num_cols = _write_period_headers(ws, 2, start_col, headers)
    _set_col_widths(ws, start_col, num_cols)
    month_cols, quarter_groups, annual_col = _get_period_col_info(headers, start_col)

    ws.cell(row=1, column=1, value="Purchases Budget (COGS)").font = HEADER_FONT

    # Part 1: Animal Purchases
    ws.cell(row=3, column=1, value="PART 1: ANIMAL PURCHASES").font = SECTION_FONT
    r = 4
    animal_cost_rows = {}
    for si, sp in enumerate(SPECIES):
        ws.cell(row=r, column=1, value=f"  {sp} — animal cost")
        sp_col = _col(si + 2)
        for mc in month_cols:
            ws.cell(row=r, column=mc,
                    value=f"={BD}!{sp_col}{refs['units_row']}*{BD}!{sp_col}{refs['animal_cost_row']}")
            ws.cell(row=r, column=mc).number_format = CURRENCY_FMT
        for qcol, qmonths in quarter_groups:
            ws.cell(row=r, column=qcol, value=_quarter_sum_formula(qcol, qmonths, r))
            ws.cell(row=r, column=qcol).number_format = CURRENCY_FMT
            ws.cell(row=r, column=qcol).font = BOLD
        ws.cell(row=r, column=annual_col, value=_annual_sum_formula(month_cols, r))
        ws.cell(row=r, column=annual_col).number_format = CURRENCY_FMT
        ws.cell(row=r, column=annual_col).font = BOLD
        animal_cost_rows[sp] = r
        r += 1

    first_ac = animal_cost_rows[SPECIES[0]]
    last_ac = animal_cost_rows[SPECIES[-1]]
    ws.cell(row=r, column=1, value="Total Animal Purchases").font = BOLD
    ws.cell(row=r, column=1).border = THIN_BORDER
    for c in month_cols + [qc for qc, _ in quarter_groups] + [annual_col]:
        ws.cell(row=r, column=c, value=f"=SUM({_col(c)}{first_ac}:{_col(c)}{last_ac})")
        ws.cell(row=r, column=c).number_format = CURRENCY_FMT
        ws.cell(row=r, column=c).font = BOLD
    total_animal_row = r

    # Processing Costs
    r += 2
    ws.cell(row=r, column=1, value="PROCESSING COSTS").font = SECTION_FONT
    r += 1
    proc_cost_rows = {}
    for si, sp in enumerate(SPECIES):
        ws.cell(row=r, column=1, value=f"  {sp} — kill + fabrication")
        sp_col = _col(si + 2)
        for mc in month_cols:
            # kill_fee + (carcass_weight * fab_cost_per_lb)
            ws.cell(row=r, column=mc,
                    value=f"={BD}!{sp_col}{refs['units_row']}*({BD}!{sp_col}{refs['kill_fee_row']}+{BD}!{sp_col}{refs['carcass_wt_row']}*{BD}!{sp_col}{refs['fab_cost_row']})")
            ws.cell(row=r, column=mc).number_format = CURRENCY_FMT
        for qcol, qmonths in quarter_groups:
            ws.cell(row=r, column=qcol, value=_quarter_sum_formula(qcol, qmonths, r))
            ws.cell(row=r, column=qcol).number_format = CURRENCY_FMT
            ws.cell(row=r, column=qcol).font = BOLD
        ws.cell(row=r, column=annual_col, value=_annual_sum_formula(month_cols, r))
        ws.cell(row=r, column=annual_col).number_format = CURRENCY_FMT
        ws.cell(row=r, column=annual_col).font = BOLD
        proc_cost_rows[sp] = r
        r += 1

    first_pc = proc_cost_rows[SPECIES[0]]
    last_pc = proc_cost_rows[SPECIES[-1]]
    ws.cell(row=r, column=1, value="Total Processing Costs").font = BOLD
    ws.cell(row=r, column=1).border = THIN_BORDER
    for c in month_cols + [qc for qc, _ in quarter_groups] + [annual_col]:
        ws.cell(row=r, column=c, value=f"=SUM({_col(c)}{first_pc}:{_col(c)}{last_pc})")
        ws.cell(row=r, column=c).number_format = CURRENCY_FMT
        ws.cell(row=r, column=c).font = BOLD
    total_proc_row = r

    # Total COGS
    r += 1
    ws.cell(row=r, column=1, value="TOTAL COGS").font = BOLD_UNDERLINE
    ws.cell(row=r, column=1).border = DOUBLE_BORDER
    for c in month_cols + [qc for qc, _ in quarter_groups] + [annual_col]:
        ws.cell(row=r, column=c, value=f"={_col(c)}{total_animal_row}+{_col(c)}{total_proc_row}")
        ws.cell(row=r, column=c).number_format = CURRENCY_FMT
        ws.cell(row=r, column=c).font = BOLD
    total_cogs_row = r

    # Part 2: Cash Payments for Purchases
    r += 2
    ws.cell(row=r, column=1, value="PART 2: CASH PAYMENTS FOR PURCHASES").font = SECTION_FONT
    r += 1
    ws.cell(row=r, column=1, value="  Paid current month")
    paid_current_row = r
    for mc in month_cols:
        ws.cell(row=r, column=mc,
                value=f"={_col(mc)}{total_cogs_row}*{BD}!B{refs['purch_pay_pct_row']}")
        ws.cell(row=r, column=mc).number_format = CURRENCY_FMT

    r += 1
    ws.cell(row=r, column=1, value="  Paid from prior month")
    paid_prior_row = r
    for idx, mc in enumerate(month_cols):
        if idx == 0:
            ws.cell(row=r, column=mc, value=0).number_format = CURRENCY_FMT
        else:
            prev_mc = month_cols[idx - 1]
            ws.cell(row=r, column=mc,
                    value=f"={_col(prev_mc)}{total_cogs_row}*(1-{BD}!B{refs['purch_pay_pct_row']})")
            ws.cell(row=r, column=mc).number_format = CURRENCY_FMT

    r += 1
    ws.cell(row=r, column=1, value="TOTAL CASH PAID FOR PURCHASES").font = BOLD
    ws.cell(row=r, column=1).border = THIN_BORDER
    total_purch_cash_row = r
    for mc in month_cols:
        ws.cell(row=r, column=mc,
                value=f"={_col(mc)}{paid_current_row}+{_col(mc)}{paid_prior_row}")
        ws.cell(row=r, column=mc).number_format = CURRENCY_FMT
        ws.cell(row=r, column=mc).font = BOLD
    for qcol, qmonths in quarter_groups:
        ws.cell(row=r, column=qcol, value=_quarter_sum_formula(qcol, qmonths, r))
        ws.cell(row=r, column=qcol).number_format = CURRENCY_FMT
        ws.cell(row=r, column=qcol).font = BOLD
    ws.cell(row=r, column=annual_col, value=_annual_sum_formula(month_cols, r))
    ws.cell(row=r, column=annual_col).number_format = CURRENCY_FMT
    ws.cell(row=r, column=annual_col).font = BOLD

    return {"total_cogs_row": total_cogs_row, "total_purch_cash_row": total_purch_cash_row,
            "total_animal_row": total_animal_row, "total_proc_row": total_proc_row}


def _create_selling_admin_budget(wb, headers, start_col, refs, sales_refs):
    """Tab 5: Selling & Admin Budget."""
    ws = wb.create_sheet("Selling & Admin Budget")
    ws.column_dimensions["A"].width = 35
    BD = "'Budget Data'"
    SB = "'Sales Budget'"

    num_cols = _write_period_headers(ws, 2, start_col, headers)
    _set_col_widths(ws, start_col, num_cols)
    month_cols, quarter_groups, annual_col = _get_period_col_info(headers, start_col)

    ws.cell(row=1, column=1, value="Selling & Administrative Budget").font = HEADER_FONT

    # Part 1: Variable Costs
    ws.cell(row=3, column=1, value="PART 1: VARIABLE COSTS").font = SECTION_FONT
    r = 4
    ws.cell(row=r, column=1, value="  Variable S&A (% of sales)")
    var_sa_cost_row = r
    for mc in month_cols:
        ws.cell(row=r, column=mc,
                value=f"={SB}!{_col(mc)}{sales_refs['total_rev_row']}*{BD}!B{refs['var_sa_row']}")
        ws.cell(row=r, column=mc).number_format = CURRENCY_FMT
    for qcol, qmonths in quarter_groups:
        ws.cell(row=r, column=qcol, value=_quarter_sum_formula(qcol, qmonths, r))
        ws.cell(row=r, column=qcol).number_format = CURRENCY_FMT
        ws.cell(row=r, column=qcol).font = BOLD
    ws.cell(row=r, column=annual_col, value=_annual_sum_formula(month_cols, r))
    ws.cell(row=r, column=annual_col).number_format = CURRENCY_FMT
    ws.cell(row=r, column=annual_col).font = BOLD

    r += 1
    ws.cell(row=r, column=1, value="Total Variable Costs").font = BOLD
    ws.cell(row=r, column=1).border = THIN_BORDER
    total_var_row = r
    for c in month_cols + [qc for qc, _ in quarter_groups] + [annual_col]:
        ws.cell(row=r, column=c, value=f"={_col(c)}{var_sa_cost_row}")
        ws.cell(row=r, column=c).number_format = CURRENCY_FMT
        ws.cell(row=r, column=c).font = BOLD

    # Part 2: Fixed Costs
    r += 2
    ws.cell(row=r, column=1, value="PART 2: FIXED COSTS").font = SECTION_FONT
    r += 1
    ws.cell(row=r, column=1, value="  Fixed S&A costs")
    fixed_sa_row = r
    for mc in month_cols:
        ws.cell(row=r, column=mc, value=f"={BD}!B{refs['total_fixed_row']}")
        ws.cell(row=r, column=mc).number_format = CURRENCY_FMT
    for qcol, qmonths in quarter_groups:
        ws.cell(row=r, column=qcol, value=_quarter_sum_formula(qcol, qmonths, r))
        ws.cell(row=r, column=qcol).number_format = CURRENCY_FMT
        ws.cell(row=r, column=qcol).font = BOLD
    ws.cell(row=r, column=annual_col, value=_annual_sum_formula(month_cols, r))
    ws.cell(row=r, column=annual_col).number_format = CURRENCY_FMT
    ws.cell(row=r, column=annual_col).font = BOLD

    r += 1
    ws.cell(row=r, column=1, value="  Depreciation")
    depr_row = r
    for mc in month_cols:
        ws.cell(row=r, column=mc, value=f"={BD}!B{refs['depreciation_row']}")
        ws.cell(row=r, column=mc).number_format = CURRENCY_FMT
    for qcol, qmonths in quarter_groups:
        ws.cell(row=r, column=qcol, value=_quarter_sum_formula(qcol, qmonths, r))
        ws.cell(row=r, column=qcol).number_format = CURRENCY_FMT
        ws.cell(row=r, column=qcol).font = BOLD
    ws.cell(row=r, column=annual_col, value=_annual_sum_formula(month_cols, r))
    ws.cell(row=r, column=annual_col).number_format = CURRENCY_FMT
    ws.cell(row=r, column=annual_col).font = BOLD

    r += 1
    ws.cell(row=r, column=1, value="Total Fixed Costs").font = BOLD
    ws.cell(row=r, column=1).border = THIN_BORDER
    total_fixed_cost_row = r
    for c in month_cols + [qc for qc, _ in quarter_groups] + [annual_col]:
        ws.cell(row=r, column=c, value=f"={_col(c)}{fixed_sa_row}+{_col(c)}{depr_row}")
        ws.cell(row=r, column=c).number_format = CURRENCY_FMT
        ws.cell(row=r, column=c).font = BOLD

    # Total S&A
    r += 1
    ws.cell(row=r, column=1, value="TOTAL S&A EXPENSE").font = BOLD_UNDERLINE
    ws.cell(row=r, column=1).border = DOUBLE_BORDER
    total_sa_row = r
    for c in month_cols + [qc for qc, _ in quarter_groups] + [annual_col]:
        ws.cell(row=r, column=c, value=f"={_col(c)}{total_var_row}+{_col(c)}{total_fixed_cost_row}")
        ws.cell(row=r, column=c).number_format = CURRENCY_FMT
        ws.cell(row=r, column=c).font = BOLD

    # Part 3: Cash Payments (exclude depreciation)
    r += 2
    ws.cell(row=r, column=1, value="PART 3: CASH PAYMENTS FOR S&A").font = SECTION_FONT
    r += 1
    ws.cell(row=r, column=1, value="TOTAL S&A CASH PAYMENTS").font = BOLD
    ws.cell(row=r, column=1).border = THIN_BORDER
    total_sa_cash_row = r
    for mc in month_cols:
        ws.cell(row=r, column=mc,
                value=f"={_col(mc)}{total_sa_row}-{_col(mc)}{depr_row}")
        ws.cell(row=r, column=mc).number_format = CURRENCY_FMT
        ws.cell(row=r, column=mc).font = BOLD
    for qcol, qmonths in quarter_groups:
        ws.cell(row=r, column=qcol, value=_quarter_sum_formula(qcol, qmonths, r))
        ws.cell(row=r, column=qcol).number_format = CURRENCY_FMT
        ws.cell(row=r, column=qcol).font = BOLD
    ws.cell(row=r, column=annual_col, value=_annual_sum_formula(month_cols, r))
    ws.cell(row=r, column=annual_col).number_format = CURRENCY_FMT
    ws.cell(row=r, column=annual_col).font = BOLD

    return {"total_var_row": total_var_row, "fixed_sa_row": fixed_sa_row,
            "depr_row": depr_row, "total_fixed_cost_row": total_fixed_cost_row,
            "total_sa_row": total_sa_row, "total_sa_cash_row": total_sa_cash_row}


def _create_cash_budget(wb, headers, start_col, refs, sales_refs, purch_refs, sa_refs):
    """Tab 6: Cash Budget."""
    ws = wb.create_sheet("Cash Budget")
    ws.column_dimensions["A"].width = 35
    BD = "'Budget Data'"
    SB = "'Sales Budget'"
    PB = "'Purchases Budget'"
    SA = "'Selling & Admin Budget'"

    num_cols = _write_period_headers(ws, 2, start_col, headers)
    _set_col_widths(ws, start_col, num_cols)
    month_cols, quarter_groups, annual_col = _get_period_col_info(headers, start_col)

    ws.cell(row=1, column=1, value="Cash Budget").font = HEADER_FONT

    # Part 1: Cash Available
    ws.cell(row=3, column=1, value="PART 1: CASH AVAILABLE").font = SECTION_FONT
    r = 4
    ws.cell(row=r, column=1, value="  Beginning Cash Balance")
    beg_cash_row = r
    # First month: reference beginning balance sheet cash (user enters)
    ws.cell(row=r, column=month_cols[0], value=0).number_format = CURRENCY_FMT
    for idx in range(1, len(month_cols)):
        mc = month_cols[idx]
        # Will be set to ending cash from previous month later
        pass  # filled after ending cash row is known

    r += 1
    ws.cell(row=r, column=1, value="  Cash Receipts (from Sales Budget)")
    cash_receipts_row = r
    for mc in month_cols:
        ws.cell(row=r, column=mc,
                value=f"={SB}!{_col(mc)}{sales_refs['total_cash_receipts_row']}")
        ws.cell(row=r, column=mc).number_format = CURRENCY_FMT

    r += 1
    ws.cell(row=r, column=1, value="TOTAL CASH AVAILABLE").font = BOLD
    ws.cell(row=r, column=1).border = THIN_BORDER
    total_avail_row = r
    for mc in month_cols:
        ws.cell(row=r, column=mc,
                value=f"={_col(mc)}{beg_cash_row}+{_col(mc)}{cash_receipts_row}")
        ws.cell(row=r, column=mc).number_format = CURRENCY_FMT
        ws.cell(row=r, column=mc).font = BOLD

    # Part 2: Cash Payments
    r += 2
    ws.cell(row=r, column=1, value="PART 2: CASH PAYMENTS").font = SECTION_FONT
    r += 1
    ws.cell(row=r, column=1, value="  Inventory Purchases")
    inv_purch_row = r
    for mc in month_cols:
        ws.cell(row=r, column=mc,
                value=f"={PB}!{_col(mc)}{purch_refs['total_purch_cash_row']}")
        ws.cell(row=r, column=mc).number_format = CURRENCY_FMT

    r += 1
    ws.cell(row=r, column=1, value="  Selling & Admin")
    sa_pay_row = r
    for mc in month_cols:
        ws.cell(row=r, column=mc,
                value=f"={SA}!{_col(mc)}{sa_refs['total_sa_cash_row']}")
        ws.cell(row=r, column=mc).number_format = CURRENCY_FMT

    r += 1
    ws.cell(row=r, column=1, value="  Capital Expenditures")
    capex_pay_row = r
    for mc in month_cols:
        ws.cell(row=r, column=mc, value=f"={BD}!B{refs['capex_row']}")
        ws.cell(row=r, column=mc).number_format = CURRENCY_FMT

    r += 1
    ws.cell(row=r, column=1, value="  Income Tax Payments")
    tax_pay_row = r
    for mc in month_cols:
        ws.cell(row=r, column=mc, value=0).number_format = CURRENCY_FMT

    r += 1
    ws.cell(row=r, column=1, value="TOTAL CASH PAYMENTS").font = BOLD
    ws.cell(row=r, column=1).border = THIN_BORDER
    total_payments_row = r
    for mc in month_cols:
        ws.cell(row=r, column=mc,
                value=f"={_col(mc)}{inv_purch_row}+{_col(mc)}{sa_pay_row}+{_col(mc)}{capex_pay_row}+{_col(mc)}{tax_pay_row}")
        ws.cell(row=r, column=mc).number_format = CURRENCY_FMT
        ws.cell(row=r, column=mc).font = BOLD

    # Part 3: Financing
    r += 2
    ws.cell(row=r, column=1, value="PART 3: FINANCING").font = SECTION_FONT
    r += 1
    ws.cell(row=r, column=1, value="  Cash Surplus / (Deficit)")
    surplus_row = r
    for mc in month_cols:
        ws.cell(row=r, column=mc,
                value=f"={_col(mc)}{total_avail_row}-{_col(mc)}{total_payments_row}")
        ws.cell(row=r, column=mc).number_format = CURRENCY_FMT_NEG
    # Conditional formatting — red if negative
    for mc in month_cols:
        cell_ref = f"{_col(mc)}{r}"
        ws.conditional_formatting.add(cell_ref,
            CellIsRule(operator="lessThan", formula=["0"], fill=RED_FILL))

    r += 1
    ws.cell(row=r, column=1, value="  New Borrowing")
    borrowing_row = r
    for mc in month_cols:
        ws.cell(row=r, column=mc, value=0).number_format = CURRENCY_FMT

    r += 1
    ws.cell(row=r, column=1, value="  Loan Repayments")
    repayment_row = r
    for mc in month_cols:
        ws.cell(row=r, column=mc, value=0).number_format = CURRENCY_FMT

    r += 1
    ws.cell(row=r, column=1, value="  Interest Expense")
    interest_row = r
    for mc in month_cols:
        ws.cell(row=r, column=mc, value=0).number_format = CURRENCY_FMT

    r += 1
    ws.cell(row=r, column=1, value="  Dividends")
    dividends_pay_row = r
    for mc in month_cols:
        ws.cell(row=r, column=mc, value=f"={BD}!B{refs['dividends_row']}")
        ws.cell(row=r, column=mc).number_format = CURRENCY_FMT

    r += 1
    ws.cell(row=r, column=1, value="ENDING CASH BALANCE").font = BOLD_UNDERLINE
    ws.cell(row=r, column=1).border = DOUBLE_BORDER
    ending_cash_row = r
    for mc in month_cols:
        ws.cell(row=r, column=mc,
                value=f"={_col(mc)}{surplus_row}+{_col(mc)}{borrowing_row}-{_col(mc)}{repayment_row}-{_col(mc)}{interest_row}-{_col(mc)}{dividends_pay_row}")
        ws.cell(row=r, column=mc).number_format = CURRENCY_FMT
        ws.cell(row=r, column=mc).font = BOLD
    # Conditional formatting on ending cash
    for mc in month_cols:
        cell_ref = f"{_col(mc)}{r}"
        ws.conditional_formatting.add(cell_ref,
            CellIsRule(operator="lessThan", formula=["0"], fill=RED_FILL))

    # Quarter + annual for ending cash
    for qcol, qmonths in quarter_groups:
        last_month_in_q = qmonths[-1]
        ws.cell(row=r, column=qcol, value=f"={_col(last_month_in_q)}{ending_cash_row}")
        ws.cell(row=r, column=qcol).number_format = CURRENCY_FMT
        ws.cell(row=r, column=qcol).font = BOLD
    ws.cell(row=r, column=annual_col, value=f"={_col(month_cols[-1])}{ending_cash_row}")
    ws.cell(row=r, column=annual_col).number_format = CURRENCY_FMT
    ws.cell(row=r, column=annual_col).font = BOLD

    # Now fill beginning cash for months 2+
    for idx in range(1, len(month_cols)):
        mc = month_cols[idx]
        prev_mc = month_cols[idx - 1]
        ws.cell(row=beg_cash_row, column=mc,
                value=f"={_col(prev_mc)}{ending_cash_row}")
        ws.cell(row=beg_cash_row, column=mc).number_format = CURRENCY_FMT

    # Long-term debt schedule
    r += 2
    ws.cell(row=r, column=1, value="LONG-TERM DEBT SCHEDULE").font = SECTION_FONT
    r += 1
    ws.cell(row=r, column=1, value="  Beginning Debt Balance")
    beg_debt_row = r
    ws.cell(row=r, column=month_cols[0], value=0).number_format = CURRENCY_FMT
    for idx in range(1, len(month_cols)):
        mc = month_cols[idx]
        prev_mc = month_cols[idx - 1]
        ws.cell(row=r, column=mc, value=f"={_col(prev_mc)}{beg_debt_row + 3}")
        ws.cell(row=r, column=mc).number_format = CURRENCY_FMT

    r += 1
    ws.cell(row=r, column=1, value="  + New Borrowing")
    for mc in month_cols:
        ws.cell(row=r, column=mc, value=f"={_col(mc)}{borrowing_row}")
        ws.cell(row=r, column=mc).number_format = CURRENCY_FMT

    r += 1
    ws.cell(row=r, column=1, value="  - Repayments")
    for mc in month_cols:
        ws.cell(row=r, column=mc, value=f"={_col(mc)}{repayment_row}")
        ws.cell(row=r, column=mc).number_format = CURRENCY_FMT

    r += 1
    ws.cell(row=r, column=1, value="  Ending Debt Balance").font = BOLD
    ws.cell(row=r, column=1).border = THIN_BORDER
    for mc in month_cols:
        ws.cell(row=r, column=mc,
                value=f"={_col(mc)}{beg_debt_row}+{_col(mc)}{beg_debt_row+1}-{_col(mc)}{beg_debt_row+2}")
        ws.cell(row=r, column=mc).number_format = CURRENCY_FMT
        ws.cell(row=r, column=mc).font = BOLD

    return {"ending_cash_row": ending_cash_row, "beg_cash_row": beg_cash_row,
            "interest_row": interest_row, "borrowing_row": borrowing_row,
            "repayment_row": repayment_row, "capex_pay_row": capex_pay_row,
            "dividends_pay_row": dividends_pay_row, "tax_pay_row": tax_pay_row}


def _create_income_statement(wb, headers, start_col, refs, sales_refs, purch_refs, sa_refs, cash_refs):
    """Tab 7: Proforma Income Statement — contribution margin format."""
    ws = wb.create_sheet("Income Statement")
    ws.column_dimensions["A"].width = 35
    BD = "'Budget Data'"
    SB = "'Sales Budget'"
    PB = "'Purchases Budget'"
    SA = "'Selling & Admin Budget'"
    CB = "'Cash Budget'"

    num_cols = _write_period_headers(ws, 2, start_col, headers)
    _set_col_widths(ws, start_col, num_cols)
    month_cols, quarter_groups, annual_col = _get_period_col_info(headers, start_col)
    all_cols = month_cols + [qc for qc, _ in quarter_groups] + [annual_col]

    ws.cell(row=1, column=1, value="Proforma Income Statement").font = HEADER_FONT

    r = 3
    ws.cell(row=r, column=1, value="Revenue")
    rev_row = r
    for mc in month_cols:
        ws.cell(row=r, column=mc, value=f"={SB}!{_col(mc)}{sales_refs['total_rev_row']}")
        ws.cell(row=r, column=mc).number_format = CURRENCY_FMT
    for qcol, qmonths in quarter_groups:
        ws.cell(row=r, column=qcol, value=_quarter_sum_formula(qcol, qmonths, r))
        ws.cell(row=r, column=qcol).number_format = CURRENCY_FMT
        ws.cell(row=r, column=qcol).font = BOLD
    ws.cell(row=r, column=annual_col, value=_annual_sum_formula(month_cols, r))
    ws.cell(row=r, column=annual_col).number_format = CURRENCY_FMT
    ws.cell(row=r, column=annual_col).font = BOLD

    r += 1
    ws.cell(row=r, column=1, value="Less: Variable COGS")
    vcogs_row = r
    for mc in month_cols:
        ws.cell(row=r, column=mc, value=f"={PB}!{_col(mc)}{purch_refs['total_cogs_row']}")
        ws.cell(row=r, column=mc).number_format = CURRENCY_FMT
    for qcol, qmonths in quarter_groups:
        ws.cell(row=r, column=qcol, value=_quarter_sum_formula(qcol, qmonths, r))
        ws.cell(row=r, column=qcol).number_format = CURRENCY_FMT
        ws.cell(row=r, column=qcol).font = BOLD
    ws.cell(row=r, column=annual_col, value=_annual_sum_formula(month_cols, r))
    ws.cell(row=r, column=annual_col).number_format = CURRENCY_FMT
    ws.cell(row=r, column=annual_col).font = BOLD

    r += 1
    ws.cell(row=r, column=1, value="Less: Variable S&A")
    var_sa_row = r
    for mc in month_cols:
        ws.cell(row=r, column=mc, value=f"={SA}!{_col(mc)}{sa_refs['total_var_row']}")
        ws.cell(row=r, column=mc).number_format = CURRENCY_FMT
    for qcol, qmonths in quarter_groups:
        ws.cell(row=r, column=qcol, value=_quarter_sum_formula(qcol, qmonths, r))
        ws.cell(row=r, column=qcol).number_format = CURRENCY_FMT
        ws.cell(row=r, column=qcol).font = BOLD
    ws.cell(row=r, column=annual_col, value=_annual_sum_formula(month_cols, r))
    ws.cell(row=r, column=annual_col).number_format = CURRENCY_FMT
    ws.cell(row=r, column=annual_col).font = BOLD

    r += 1
    ws.cell(row=r, column=1, value="CONTRIBUTION MARGIN").font = BOLD
    ws.cell(row=r, column=1).border = THIN_BORDER
    cm_row = r
    for c in all_cols:
        ws.cell(row=r, column=c, value=f"={_col(c)}{rev_row}-{_col(c)}{vcogs_row}-{_col(c)}{var_sa_row}")
        ws.cell(row=r, column=c).number_format = CURRENCY_FMT
        ws.cell(row=r, column=c).font = BOLD

    r += 1
    ws.cell(row=r, column=1, value="Less: Fixed S&A")
    fixed_sa_row = r
    for mc in month_cols:
        ws.cell(row=r, column=mc, value=f"={SA}!{_col(mc)}{sa_refs['total_fixed_cost_row']}")
        ws.cell(row=r, column=mc).number_format = CURRENCY_FMT
    for qcol, qmonths in quarter_groups:
        ws.cell(row=r, column=qcol, value=_quarter_sum_formula(qcol, qmonths, r))
        ws.cell(row=r, column=qcol).number_format = CURRENCY_FMT
        ws.cell(row=r, column=qcol).font = BOLD
    ws.cell(row=r, column=annual_col, value=_annual_sum_formula(month_cols, r))
    ws.cell(row=r, column=annual_col).number_format = CURRENCY_FMT
    ws.cell(row=r, column=annual_col).font = BOLD

    r += 1
    ws.cell(row=r, column=1, value="OPERATING INCOME").font = BOLD
    ws.cell(row=r, column=1).border = THIN_BORDER
    op_income_row = r
    for c in all_cols:
        ws.cell(row=r, column=c, value=f"={_col(c)}{cm_row}-{_col(c)}{fixed_sa_row}")
        ws.cell(row=r, column=c).number_format = CURRENCY_FMT_NEG
        ws.cell(row=r, column=c).font = BOLD

    r += 1
    ws.cell(row=r, column=1, value="Less: Interest Expense")
    int_exp_row = r
    for mc in month_cols:
        ws.cell(row=r, column=mc, value=f"={CB}!{_col(mc)}{cash_refs['interest_row']}")
        ws.cell(row=r, column=mc).number_format = CURRENCY_FMT
    for qcol, qmonths in quarter_groups:
        ws.cell(row=r, column=qcol, value=_quarter_sum_formula(qcol, qmonths, r))
        ws.cell(row=r, column=qcol).number_format = CURRENCY_FMT
        ws.cell(row=r, column=qcol).font = BOLD
    ws.cell(row=r, column=annual_col, value=_annual_sum_formula(month_cols, r))
    ws.cell(row=r, column=annual_col).number_format = CURRENCY_FMT
    ws.cell(row=r, column=annual_col).font = BOLD

    r += 1
    ws.cell(row=r, column=1, value="INCOME BEFORE TAXES").font = BOLD
    ws.cell(row=r, column=1).border = THIN_BORDER
    ibt_row = r
    for c in all_cols:
        ws.cell(row=r, column=c, value=f"={_col(c)}{op_income_row}-{_col(c)}{int_exp_row}")
        ws.cell(row=r, column=c).number_format = CURRENCY_FMT_NEG
        ws.cell(row=r, column=c).font = BOLD

    r += 1
    ws.cell(row=r, column=1, value="Less: Income Tax Expense")
    tax_exp_row = r
    for mc in month_cols:
        # MAX(0, ibt * tax_rate) — no tax benefit if loss
        ws.cell(row=r, column=mc,
                value=f"=MAX(0,{_col(mc)}{ibt_row}*{BD}!B{refs['tax_rate_row']})")
        ws.cell(row=r, column=mc).number_format = CURRENCY_FMT
    for qcol, qmonths in quarter_groups:
        ws.cell(row=r, column=qcol, value=_quarter_sum_formula(qcol, qmonths, r))
        ws.cell(row=r, column=qcol).number_format = CURRENCY_FMT
        ws.cell(row=r, column=qcol).font = BOLD
    ws.cell(row=r, column=annual_col, value=_annual_sum_formula(month_cols, r))
    ws.cell(row=r, column=annual_col).number_format = CURRENCY_FMT
    ws.cell(row=r, column=annual_col).font = BOLD

    r += 1
    ws.cell(row=r, column=1, value="NET INCOME").font = BOLD_UNDERLINE
    ws.cell(row=r, column=1).border = DOUBLE_BORDER
    ni_row = r
    for c in all_cols:
        ws.cell(row=r, column=c, value=f"={_col(c)}{ibt_row}-{_col(c)}{tax_exp_row}")
        ws.cell(row=r, column=c).number_format = CURRENCY_FMT_NEG
        ws.cell(row=r, column=c).font = BOLD

    return {"rev_row": rev_row, "vcogs_row": vcogs_row, "cm_row": cm_row,
            "op_income_row": op_income_row, "ni_row": ni_row,
            "tax_exp_row": tax_exp_row, "int_exp_row": int_exp_row}


def _create_ending_balance_sheet(wb, headers, start_col, refs, sales_refs, purch_refs, cash_refs, is_refs):
    """Tab 8: Balance Sheet (Ending) — monthly ending balances."""
    ws = wb.create_sheet("Balance Sheet (Ending)")
    ws.column_dimensions["A"].width = 40
    CB = "'Cash Budget'"
    SB = "'Sales Budget'"
    PB = "'Purchases Budget'"
    IS_ = "'Income Statement'"
    BD = "'Budget Data'"

    num_cols = _write_period_headers(ws, 2, start_col, headers)
    _set_col_widths(ws, start_col, num_cols)
    month_cols, quarter_groups, annual_col = _get_period_col_info(headers, start_col)

    ws.cell(row=1, column=1, value="Balance Sheet (Ending)").font = HEADER_FONT

    r = 3
    ws.cell(row=r, column=1, value="ASSETS").font = SECTION_FONT
    r += 1
    ws.cell(row=r, column=1, value="Current Assets").font = BOLD

    r += 1
    ws.cell(row=r, column=1, value="  Cash")
    cash_row = r
    for mc in month_cols:
        ws.cell(row=r, column=mc, value=f"={CB}!{_col(mc)}{cash_refs['ending_cash_row']}")
        ws.cell(row=r, column=mc).number_format = CURRENCY_FMT

    r += 1
    ws.cell(row=r, column=1, value="  Accounts Receivable")
    ar_row = r
    # A/R = total revenue - total cash receipts (cumulative uncollected)
    for mc in month_cols:
        ws.cell(row=r, column=mc,
                value=f"={SB}!{_col(mc)}{sales_refs['total_rev_row']}-{SB}!{_col(mc)}{sales_refs['total_cash_receipts_row']}")
        ws.cell(row=r, column=mc).number_format = CURRENCY_FMT

    r += 1
    ws.cell(row=r, column=1, value="  Inventory")
    inv_row = r
    for mc in month_cols:
        ws.cell(row=r, column=mc, value=0).number_format = CURRENCY_FMT

    r += 1
    ws.cell(row=r, column=1, value="  Prepaid Expenses")
    prepaid_row = r
    for mc in month_cols:
        ws.cell(row=r, column=mc, value=0).number_format = CURRENCY_FMT

    r += 1
    ws.cell(row=r, column=1, value="TOTAL CURRENT ASSETS").font = BOLD
    ws.cell(row=r, column=1).border = THIN_BORDER
    total_ca_row = r
    for mc in month_cols:
        ws.cell(row=r, column=mc,
                value=f"=SUM({_col(mc)}{cash_row}:{_col(mc)}{prepaid_row})")
        ws.cell(row=r, column=mc).number_format = CURRENCY_FMT
        ws.cell(row=r, column=mc).font = BOLD

    r += 2
    ws.cell(row=r, column=1, value="  PPE (Net)")
    ppe_row = r
    for mc in month_cols:
        ws.cell(row=r, column=mc, value=0).number_format = CURRENCY_FMT

    r += 2
    ws.cell(row=r, column=1, value="TOTAL ASSETS").font = BOLD_UNDERLINE
    ws.cell(row=r, column=1).border = DOUBLE_BORDER
    total_assets_row = r
    for mc in month_cols:
        ws.cell(row=r, column=mc, value=f"={_col(mc)}{total_ca_row}+{_col(mc)}{ppe_row}")
        ws.cell(row=r, column=mc).number_format = CURRENCY_FMT
        ws.cell(row=r, column=mc).font = BOLD

    # Liabilities
    r += 2
    ws.cell(row=r, column=1, value="LIABILITIES").font = SECTION_FONT
    r += 1
    ws.cell(row=r, column=1, value="  Accounts Payable")
    ap_row = r
    # A/P = total COGS - total cash paid for purchases
    for mc in month_cols:
        ws.cell(row=r, column=mc,
                value=f"={PB}!{_col(mc)}{purch_refs['total_cogs_row']}-{PB}!{_col(mc)}{purch_refs['total_purch_cash_row']}")
        ws.cell(row=r, column=mc).number_format = CURRENCY_FMT

    r += 1
    ws.cell(row=r, column=1, value="  Accrued Liabilities")
    accrued_row = r
    for mc in month_cols:
        ws.cell(row=r, column=mc, value=0).number_format = CURRENCY_FMT

    r += 1
    ws.cell(row=r, column=1, value="TOTAL CURRENT LIABILITIES").font = BOLD
    ws.cell(row=r, column=1).border = THIN_BORDER
    total_cl_row = r
    for mc in month_cols:
        ws.cell(row=r, column=mc, value=f"={_col(mc)}{ap_row}+{_col(mc)}{accrued_row}")
        ws.cell(row=r, column=mc).number_format = CURRENCY_FMT
        ws.cell(row=r, column=mc).font = BOLD

    r += 1
    ws.cell(row=r, column=1, value="  Long-Term Debt")
    ltd_row = r
    for mc in month_cols:
        ws.cell(row=r, column=mc, value=0).number_format = CURRENCY_FMT

    r += 1
    ws.cell(row=r, column=1, value="TOTAL LIABILITIES").font = BOLD
    total_liab_row = r
    for mc in month_cols:
        ws.cell(row=r, column=mc, value=f"={_col(mc)}{total_cl_row}+{_col(mc)}{ltd_row}")
        ws.cell(row=r, column=mc).number_format = CURRENCY_FMT
        ws.cell(row=r, column=mc).font = BOLD

    # Equity
    r += 2
    ws.cell(row=r, column=1, value="STOCKHOLDERS' EQUITY").font = SECTION_FONT
    r += 1
    ws.cell(row=r, column=1, value="  Common Stock")
    cs_row = r
    for mc in month_cols:
        ws.cell(row=r, column=mc, value=0).number_format = CURRENCY_FMT

    r += 1
    ws.cell(row=r, column=1, value="  Retained Earnings")
    re_row = r
    # RE = beginning RE + net income - dividends (cumulative)
    for idx, mc in enumerate(month_cols):
        if idx == 0:
            ws.cell(row=r, column=mc,
                    value=f"={IS_}!{_col(mc)}{is_refs['ni_row']}-{BD}!B{refs['dividends_row']}")
        else:
            prev_mc = month_cols[idx - 1]
            ws.cell(row=r, column=mc,
                    value=f"={_col(prev_mc)}{re_row}+{IS_}!{_col(mc)}{is_refs['ni_row']}-{BD}!B{refs['dividends_row']}")
        ws.cell(row=r, column=mc).number_format = CURRENCY_FMT

    r += 1
    ws.cell(row=r, column=1, value="TOTAL EQUITY").font = BOLD
    ws.cell(row=r, column=1).border = THIN_BORDER
    total_eq_row = r
    for mc in month_cols:
        ws.cell(row=r, column=mc, value=f"={_col(mc)}{cs_row}+{_col(mc)}{re_row}")
        ws.cell(row=r, column=mc).number_format = CURRENCY_FMT
        ws.cell(row=r, column=mc).font = BOLD

    r += 1
    ws.cell(row=r, column=1, value="TOTAL LIABILITIES + EQUITY").font = BOLD_UNDERLINE
    ws.cell(row=r, column=1).border = DOUBLE_BORDER
    for mc in month_cols:
        ws.cell(row=r, column=mc, value=f"={_col(mc)}{total_liab_row}+{_col(mc)}{total_eq_row}")
        ws.cell(row=r, column=mc).number_format = CURRENCY_FMT
        ws.cell(row=r, column=mc).font = BOLD

    return {"cash_row": cash_row, "ar_row": ar_row, "inv_row": inv_row,
            "ap_row": ap_row, "accrued_row": accrued_row,
            "re_row": re_row, "ppe_row": ppe_row}


def _create_cash_flow_statement(wb, headers, start_col, refs, is_refs, sa_refs, bs_refs, cash_refs):
    """Tab 9: Cash Flow Statement — operating, investing, financing."""
    ws = wb.create_sheet("Cash Flow Statement")
    ws.column_dimensions["A"].width = 40
    IS_ = "'Income Statement'"
    SA = "'Selling & Admin Budget'"
    BS = "'Balance Sheet (Ending)'"
    CB = "'Cash Budget'"
    BD = "'Budget Data'"

    num_cols = _write_period_headers(ws, 2, start_col, headers)
    _set_col_widths(ws, start_col, num_cols)
    month_cols, quarter_groups, annual_col = _get_period_col_info(headers, start_col)

    ws.cell(row=1, column=1, value="Cash Flow Statement").font = HEADER_FONT

    # Operating Activities
    r = 3
    ws.cell(row=r, column=1, value="OPERATING ACTIVITIES").font = SECTION_FONT
    r += 1
    ws.cell(row=r, column=1, value="  Net Income")
    ni_row = r
    for mc in month_cols:
        ws.cell(row=r, column=mc, value=f"={IS_}!{_col(mc)}{is_refs['ni_row']}")
        ws.cell(row=r, column=mc).number_format = CURRENCY_FMT

    r += 1
    ws.cell(row=r, column=1, value="  + Depreciation")
    depr_row = r
    for mc in month_cols:
        ws.cell(row=r, column=mc, value=f"={SA}!{_col(mc)}{sa_refs['depr_row']}")
        ws.cell(row=r, column=mc).number_format = CURRENCY_FMT

    r += 1
    ws.cell(row=r, column=1, value="  Change in A/R")
    chg_ar_row = r
    for idx, mc in enumerate(month_cols):
        if idx == 0:
            # Assume beginning A/R = 0
            ws.cell(row=r, column=mc, value=f"=-{BS}!{_col(mc)}{bs_refs['ar_row']}")
        else:
            prev_mc = month_cols[idx - 1]
            ws.cell(row=r, column=mc,
                    value=f"=-({BS}!{_col(mc)}{bs_refs['ar_row']}-{BS}!{_col(prev_mc)}{bs_refs['ar_row']})")
        ws.cell(row=r, column=mc).number_format = CURRENCY_FMT

    r += 1
    ws.cell(row=r, column=1, value="  Change in Inventory")
    chg_inv_row = r
    for idx, mc in enumerate(month_cols):
        if idx == 0:
            ws.cell(row=r, column=mc, value=f"=-{BS}!{_col(mc)}{bs_refs['inv_row']}")
        else:
            prev_mc = month_cols[idx - 1]
            ws.cell(row=r, column=mc,
                    value=f"=-({BS}!{_col(mc)}{bs_refs['inv_row']}-{BS}!{_col(prev_mc)}{bs_refs['inv_row']})")
        ws.cell(row=r, column=mc).number_format = CURRENCY_FMT

    r += 1
    ws.cell(row=r, column=1, value="  Change in A/P")
    chg_ap_row = r
    for idx, mc in enumerate(month_cols):
        if idx == 0:
            ws.cell(row=r, column=mc, value=f"={BS}!{_col(mc)}{bs_refs['ap_row']}")
        else:
            prev_mc = month_cols[idx - 1]
            ws.cell(row=r, column=mc,
                    value=f"={BS}!{_col(mc)}{bs_refs['ap_row']}-{BS}!{_col(prev_mc)}{bs_refs['ap_row']}")
        ws.cell(row=r, column=mc).number_format = CURRENCY_FMT

    r += 1
    ws.cell(row=r, column=1, value="  Change in Accrued Liabilities")
    chg_accrued_row = r
    for idx, mc in enumerate(month_cols):
        if idx == 0:
            ws.cell(row=r, column=mc, value=f"={BS}!{_col(mc)}{bs_refs['accrued_row']}")
        else:
            prev_mc = month_cols[idx - 1]
            ws.cell(row=r, column=mc,
                    value=f"={BS}!{_col(mc)}{bs_refs['accrued_row']}-{BS}!{_col(prev_mc)}{bs_refs['accrued_row']}")
        ws.cell(row=r, column=mc).number_format = CURRENCY_FMT

    r += 1
    ws.cell(row=r, column=1, value="NET CASH FROM OPERATIONS").font = BOLD
    ws.cell(row=r, column=1).border = THIN_BORDER
    ops_row = r
    for mc in month_cols:
        ws.cell(row=r, column=mc,
                value=f"=SUM({_col(mc)}{ni_row}:{_col(mc)}{chg_accrued_row})")
        ws.cell(row=r, column=mc).number_format = CURRENCY_FMT
        ws.cell(row=r, column=mc).font = BOLD
    for qcol, qmonths in quarter_groups:
        ws.cell(row=r, column=qcol, value=_quarter_sum_formula(qcol, qmonths, r))
        ws.cell(row=r, column=qcol).number_format = CURRENCY_FMT
        ws.cell(row=r, column=qcol).font = BOLD
    ws.cell(row=r, column=annual_col, value=_annual_sum_formula(month_cols, r))
    ws.cell(row=r, column=annual_col).number_format = CURRENCY_FMT
    ws.cell(row=r, column=annual_col).font = BOLD

    # Investing Activities
    r += 2
    ws.cell(row=r, column=1, value="INVESTING ACTIVITIES").font = SECTION_FONT
    r += 1
    ws.cell(row=r, column=1, value="  Capital Expenditures")
    capex_row = r
    for mc in month_cols:
        ws.cell(row=r, column=mc, value=f"=-{CB}!{_col(mc)}{cash_refs['capex_pay_row']}")
        ws.cell(row=r, column=mc).number_format = CURRENCY_FMT

    r += 1
    ws.cell(row=r, column=1, value="NET CASH FROM INVESTING").font = BOLD
    ws.cell(row=r, column=1).border = THIN_BORDER
    invest_row = r
    for mc in month_cols:
        ws.cell(row=r, column=mc, value=f"={_col(mc)}{capex_row}")
        ws.cell(row=r, column=mc).number_format = CURRENCY_FMT
        ws.cell(row=r, column=mc).font = BOLD
    for qcol, qmonths in quarter_groups:
        ws.cell(row=r, column=qcol, value=_quarter_sum_formula(qcol, qmonths, r))
        ws.cell(row=r, column=qcol).number_format = CURRENCY_FMT
        ws.cell(row=r, column=qcol).font = BOLD
    ws.cell(row=r, column=annual_col, value=_annual_sum_formula(month_cols, r))
    ws.cell(row=r, column=annual_col).number_format = CURRENCY_FMT
    ws.cell(row=r, column=annual_col).font = BOLD

    # Financing Activities
    r += 2
    ws.cell(row=r, column=1, value="FINANCING ACTIVITIES").font = SECTION_FONT
    r += 1
    ws.cell(row=r, column=1, value="  Borrowings")
    borrow_row = r
    for mc in month_cols:
        ws.cell(row=r, column=mc, value=f"={CB}!{_col(mc)}{cash_refs['borrowing_row']}")
        ws.cell(row=r, column=mc).number_format = CURRENCY_FMT

    r += 1
    ws.cell(row=r, column=1, value="  Repayments")
    repay_row = r
    for mc in month_cols:
        ws.cell(row=r, column=mc, value=f"=-{CB}!{_col(mc)}{cash_refs['repayment_row']}")
        ws.cell(row=r, column=mc).number_format = CURRENCY_FMT

    r += 1
    ws.cell(row=r, column=1, value="  Interest Paid")
    int_paid_row = r
    for mc in month_cols:
        ws.cell(row=r, column=mc, value=f"=-{CB}!{_col(mc)}{cash_refs['interest_row']}")
        ws.cell(row=r, column=mc).number_format = CURRENCY_FMT

    r += 1
    ws.cell(row=r, column=1, value="  Dividends Paid")
    div_paid_row = r
    for mc in month_cols:
        ws.cell(row=r, column=mc, value=f"=-{CB}!{_col(mc)}{cash_refs['dividends_pay_row']}")
        ws.cell(row=r, column=mc).number_format = CURRENCY_FMT

    r += 1
    ws.cell(row=r, column=1, value="NET CASH FROM FINANCING").font = BOLD
    ws.cell(row=r, column=1).border = THIN_BORDER
    fin_row = r
    for mc in month_cols:
        ws.cell(row=r, column=mc,
                value=f"=SUM({_col(mc)}{borrow_row}:{_col(mc)}{div_paid_row})")
        ws.cell(row=r, column=mc).number_format = CURRENCY_FMT
        ws.cell(row=r, column=mc).font = BOLD
    for qcol, qmonths in quarter_groups:
        ws.cell(row=r, column=qcol, value=_quarter_sum_formula(qcol, qmonths, r))
        ws.cell(row=r, column=qcol).number_format = CURRENCY_FMT
        ws.cell(row=r, column=qcol).font = BOLD
    ws.cell(row=r, column=annual_col, value=_annual_sum_formula(month_cols, r))
    ws.cell(row=r, column=annual_col).number_format = CURRENCY_FMT
    ws.cell(row=r, column=annual_col).font = BOLD

    # Summary
    r += 2
    ws.cell(row=r, column=1, value="NET CHANGE IN CASH").font = BOLD
    ws.cell(row=r, column=1).border = THIN_BORDER
    net_chg_row = r
    for mc in month_cols:
        ws.cell(row=r, column=mc,
                value=f"={_col(mc)}{ops_row}+{_col(mc)}{invest_row}+{_col(mc)}{fin_row}")
        ws.cell(row=r, column=mc).number_format = CURRENCY_FMT
        ws.cell(row=r, column=mc).font = BOLD
    for qcol, qmonths in quarter_groups:
        ws.cell(row=r, column=qcol, value=_quarter_sum_formula(qcol, qmonths, r))
        ws.cell(row=r, column=qcol).number_format = CURRENCY_FMT
        ws.cell(row=r, column=qcol).font = BOLD
    ws.cell(row=r, column=annual_col, value=_annual_sum_formula(month_cols, r))
    ws.cell(row=r, column=annual_col).number_format = CURRENCY_FMT
    ws.cell(row=r, column=annual_col).font = BOLD

    r += 1
    ws.cell(row=r, column=1, value="Beginning Cash")
    beg_cash_cf_row = r
    for mc in month_cols:
        ws.cell(row=r, column=mc, value=f"={CB}!{_col(mc)}{cash_refs['beg_cash_row']}")
        ws.cell(row=r, column=mc).number_format = CURRENCY_FMT

    r += 1
    ws.cell(row=r, column=1, value="ENDING CASH").font = BOLD_UNDERLINE
    ws.cell(row=r, column=1).border = DOUBLE_BORDER
    for mc in month_cols:
        ws.cell(row=r, column=mc,
                value=f"={_col(mc)}{beg_cash_cf_row}+{_col(mc)}{net_chg_row}")
        ws.cell(row=r, column=mc).number_format = CURRENCY_FMT
        ws.cell(row=r, column=mc).font = BOLD

    r += 1
    ws.cell(row=r, column=1, value="  Check: ties to Cash Budget?")
    for mc in month_cols:
        ws.cell(row=r, column=mc,
                value=f"={_col(mc)}{r-1}-{CB}!{_col(mc)}{cash_refs['ending_cash_row']}")
        ws.cell(row=r, column=mc).number_format = CURRENCY_FMT
        # Should be 0
        ws.conditional_formatting.add(f"{_col(mc)}{r}",
            CellIsRule(operator="notEqual", formula=["0"], fill=RED_FILL))


# ---------------------------------------------------------------------------
# Main generator
# ---------------------------------------------------------------------------

def generate_budget_workbook(output_path=None, months=12, start_month=None):
    """Generate the flexible budget workbook."""
    today = date.today()
    if start_month:
        sm_year, sm_month = start_month
    else:
        sm_month = today.month + 1
        sm_year = today.year
        if sm_month > 12:
            sm_month = 1
            sm_year += 1

    if not output_path:
        output_path = os.path.join(REPORTS_DIR, f"budget_workbook_{today.strftime('%Y%m%d')}.xlsx")

    print(f"Generating budget workbook: {output_path}")
    print(f"  Period: {month_abbr[sm_month]} {sm_year} — {months} months")

    # Fetch DB data
    print("  Fetching latest prices from DB...")
    db_prices = _get_latest_prices_by_species()
    defaults = _species_defaults()

    # Log what we found
    for sp in SPECIES:
        if sp in db_prices:
            price = db_prices[sp].get("sale_price_per_lb", 0)
            print(f"    {sp}: ${price:.4f}/lb (from DB)")
        else:
            print(f"    {sp}: no DB data — using placeholder")

    # Build workbook
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    headers = _month_headers(sm_month, sm_year, months)
    start_col = 2  # Column B

    # Tab 1
    _create_beginning_balance_sheet(wb)

    # Tab 2
    refs = _create_budget_data(wb, db_prices, defaults)

    # Tab 3
    sales_refs = _create_sales_budget(wb, headers, start_col, refs)

    # Tab 4
    purch_refs = _create_purchases_budget(wb, headers, start_col, refs)

    # Tab 5
    sa_refs = _create_selling_admin_budget(wb, headers, start_col, refs, sales_refs)

    # Tab 6
    cash_refs = _create_cash_budget(wb, headers, start_col, refs, sales_refs, purch_refs, sa_refs)

    # Tab 7
    is_refs = _create_income_statement(wb, headers, start_col, refs, sales_refs, purch_refs, sa_refs, cash_refs)

    # Tab 8
    bs_refs = _create_ending_balance_sheet(wb, headers, start_col, refs, sales_refs, purch_refs, cash_refs, is_refs)

    # Tab 9
    _create_cash_flow_statement(wb, headers, start_col, refs, is_refs, sa_refs, bs_refs, cash_refs)

    wb.save(output_path)
    print(f"Budget workbook saved: {output_path}")
    return output_path


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Generate Flexible Budget Workbook")
    parser.add_argument("--start-month", type=str, default=None,
                        help="Start month in YYYY-MM format (default: next month)")
    parser.add_argument("--output", type=str, default=None,
                        help="Output file path (default: reports/budget_workbook_YYYYMMDD.xlsx)")
    parser.add_argument("--months", type=int, default=12,
                        help="Number of months (default: 12)")
    args = parser.parse_args()

    start_month = None
    if args.start_month:
        parts = args.start_month.split("-")
        start_month = (int(parts[0]), int(parts[1]))

    generate_budget_workbook(
        output_path=args.output,
        months=args.months,
        start_month=start_month,
    )
