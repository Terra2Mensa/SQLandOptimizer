#!/usr/bin/env python3
"""
Cattle Valuation Engine v2
==========================
Auto-pulls USDA boxed beef prices (Prime, Choice, Select, Grassfed) and computes
carcass value from live weight. Includes Indiana auction data and 5 Area slaughter
cattle prices for market comparison.

Data sources:
  - USDA DataMart 2461 (Weekly Boxed Beef - Choice & Select cuts)
  - USDA DataMart 2460 (Weekly Boxed Beef - Prime cuts)
  - USDA DataMart 2477 (5 Area Weekly Slaughter Cattle)
  - USDA DataMart 2672 (IA-MN Daily Slaughter Cattle)
  - USDA DataMart 2482 (Premiums & Discounts)
  - USDA MARS 1976 (Indiana Weekly Auction Summary)

Usage:
  python3 cattle_valuation.py --all-grades
  python3 cattle_valuation.py --quality-grade prime --output valuation.xlsx
  python3 cattle_valuation.py --live-weight 1400 --yield-grade 2 --all-grades --save-db
"""

import argparse
import base64
import json
import re
import urllib.request
from datetime import datetime
from dataclasses import dataclass, field

from config import (
    MARS_API_KEY, MARS_BASE_URL, DATAMART_BASE_URL,
    REPORT_CHOICE_SELECT, REPORT_PRIME, REPORT_5AREA_WEEKLY,
    REPORT_IAMN_DAILY, REPORT_PREMIUMS, REPORT_IN_AUCTION,
    DRESS_PCT_BY_YG, SUBPRIMAL_YIELDS, PRIMAL_ORDER,
    DEFAULT_LIVE_WEIGHT, DEFAULT_YIELD_GRADE, DEFAULT_QUALITY_GRADE,
    DEFAULT_BROKER_FEE_PCT, DEFAULT_BYPRODUCT_PCT,
    DEFAULT_BYPRODUCT_VALUE_PER_LB, DEFAULT_GRASSFED_PREMIUM_CWT,
    PROCESSORS, DEFAULT_PROCESSOR,
)


# ---------------------------------------------------------------------------
# Data classes
# ---------------------------------------------------------------------------

@dataclass
class SubprimalPrice:
    imps_code: str
    description: str
    weighted_avg_cwt: float
    price_range_low: float
    price_range_high: float
    number_trades: int
    total_pounds: int
    grade: str

    @property
    def price_per_lb(self) -> float:
        return self.weighted_avg_cwt / 100.0


@dataclass
class CarcassValuation:
    live_weight: float
    yield_grade: int
    quality_grade: str
    dressing_pct: float
    hot_carcass_weight: float
    report_date: str
    cut_values: list = field(default_factory=list)
    total_subprimal_value: float = 0.0
    byproduct_value: float = 0.0
    total_carcass_value: float = 0.0
    value_per_cwt_carcass: float = 0.0
    value_per_cwt_live: float = 0.0
    broker_fee: float = 0.0
    net_value: float = 0.0


@dataclass
class PurchasePriceResult:
    quality_grade: str
    live_weight: float
    dressing_pct: float
    processor_name: str
    processor_costs: dict = field(default_factory=dict)
    # Method results ($/cwt live)
    live_basis_cwt: float = 0.0
    dressed_basis_cwt: float = 0.0
    grid_formula_cwt: float = 0.0
    cutout_minus_margin_cwt: float = 0.0
    # Supporting detail
    live_market_ref_cwt: float = 0.0
    dressed_market_ref_cwt: float = 0.0
    grid_base_cwt: float = 0.0
    grid_quality_adj: float = 0.0
    grid_yg_adj: float = 0.0
    grid_wt_adj: float = 0.0
    cutout_net_value: float = 0.0
    cutout_kill_fee: float = 0.0
    cutout_fab_total: float = 0.0
    cutout_shrink_cost: float = 0.0
    auction_ref_cwt: float = 0.0


# ---------------------------------------------------------------------------
# Parsing helpers
# ---------------------------------------------------------------------------

def parse_number(s) -> float:
    if s is None or str(s) in ('', 'None', '.00'):
        return 0.0
    return float(str(s).replace(',', ''))


def parse_imps_code(description: str) -> str:
    match = re.search(r'\((\d+[A-Z]?)\s', description)
    return match.group(1) if match else ""


# ---------------------------------------------------------------------------
# DataMart API helpers
# ---------------------------------------------------------------------------

def fetch_datamart(report_id: int, last_reports: int = 1, all_sections: bool = True) -> list:
    params = f"lastReports={last_reports}"
    if all_sections:
        params += "&allSections=True"
    url = f"{DATAMART_BASE_URL}/{report_id}?{params}"
    req = urllib.request.Request(url)
    req.add_header('Accept', 'application/json')
    resp = urllib.request.urlopen(req, timeout=60)
    return json.loads(resp.read().decode())


def fetch_mars(report_id: int, api_key: str, last_reports: int = 1) -> dict:
    url = f"{MARS_BASE_URL}/{report_id}?lastReports={last_reports}"
    req = urllib.request.Request(url)
    req.add_header('Accept', 'application/json')
    encoded = base64.b64encode(f":{api_key}".encode()).decode()
    req.add_header('Authorization', f'Basic {encoded}')
    resp = urllib.request.urlopen(req, timeout=60)
    return json.loads(resp.read().decode())


def parse_cuts_from_section(items: list, grade: str) -> list:
    cuts = []
    for item in items:
        code = parse_imps_code(item.get('item_description', ''))
        if not code:
            continue
        avg = parse_number(item.get('weighted_average', '0'))
        if avg <= 0:
            continue
        cuts.append(SubprimalPrice(
            imps_code=code,
            description=item['item_description'],
            weighted_avg_cwt=avg,
            price_range_low=parse_number(item.get('price_range_low', '0')),
            price_range_high=parse_number(item.get('price_range_high', '0')),
            number_trades=int(parse_number(item.get('number_trades', '0'))),
            total_pounds=int(parse_number(item.get('total_pounds', '0'))),
            grade=grade,
        ))
    return cuts


# ---------------------------------------------------------------------------
# Data fetching functions
# ---------------------------------------------------------------------------

def fetch_boxed_beef() -> dict:
    """Fetch Choice/Select (2461) and Prime (2460) boxed beef cuts."""
    print("Fetching USDA boxed beef reports...")

    # Choice & Select
    data_2461 = fetch_datamart(REPORT_CHOICE_SELECT)
    result = {
        'choice_cuts': [], 'select_cuts': [], 'prime_cuts': [],
        'composites': {}, 'ground_beef': [], 'trimmings': [],
        'weekly_avg': {}, 'report_date': '',
    }

    for section in data_2461:
        sec = section.get('reportSection', '')
        items = section.get('results', [])
        if not items:
            continue
        result['report_date'] = items[0].get('report_date', '')

        if sec == 'Choice Cuts':
            result['choice_cuts'] = parse_cuts_from_section(items, 'Choice')
        elif sec == 'Select Cuts':
            result['select_cuts'] = parse_cuts_from_section(items, 'Select')
        elif sec == 'Weekly Composite Primal Values':
            for item in items:
                result['composites'][item.get('primal_desc', '')] = {
                    'choice': parse_number(item.get('choice_600_900', '0')),
                    'select': parse_number(item.get('select_600_900', '0')),
                }
        elif sec == 'Weekly Average Cutout Values':
            for item in items:
                result['weekly_avg'] = {
                    'choice': parse_number(item.get('choice_600_900_simple_avg', '0')),
                    'select': parse_number(item.get('select_600_900_simple_avg', '0')),
                }
        elif sec == 'Ground Beef':
            for item in items:
                avg = parse_number(item.get('weighted_average', '0'))
                if avg > 0:
                    result['ground_beef'].append({
                        'description': item.get('trim_description', ''),
                        'weighted_avg_cwt': avg,
                        'total_pounds': int(parse_number(item.get('total_pounds', '0'))),
                    })
        elif sec == 'Beef Trimmings':
            for item in items:
                avg = parse_number(item.get('weighted_average', '0'))
                if avg > 0:
                    result['trimmings'].append({
                        'description': item.get('trim_description', ''),
                        'weighted_avg_cwt': avg,
                    })

    # Prime
    data_2460 = fetch_datamart(REPORT_PRIME)
    for section in data_2460:
        if section.get('reportSection') == 'Prime Cuts':
            result['prime_cuts'] = parse_cuts_from_section(
                section.get('results', []), 'Prime'
            )

    print(f"  Report date: {result['report_date']}")
    print(f"  Prime: {len(result['prime_cuts'])} cuts | "
          f"Choice: {len(result['choice_cuts'])} cuts | "
          f"Select: {len(result['select_cuts'])} cuts")
    wavg = result.get('weekly_avg', {})
    print(f"  Weekly avg cutout: Choice ${wavg.get('choice', 0):.2f}/cwt, "
          f"Select ${wavg.get('select', 0):.2f}/cwt")

    return result


def fetch_slaughter_cattle() -> dict:
    """Fetch 5 Area Weekly (2477) and IA-MN Daily (2672) slaughter cattle."""
    print("Fetching slaughter cattle reports...")
    result = {'five_area': [], 'iamn': [], 'iamn_narrative': ''}

    # 5 Area Weekly
    data_2477 = fetch_datamart(REPORT_5AREA_WEEKLY)
    for section in data_2477:
        sec = section.get('reportSection', '')
        items = section.get('results', [])
        if sec == 'Detail':
            for item in items:
                price = item.get('weighted_avg_price', '')
                if not price or str(price) == 'None':
                    continue
                result['five_area'].append({
                    'report_date': item.get('report_date', ''),
                    'class': item.get('class_description', ''),
                    'basis': item.get('selling_basis_description', ''),
                    'grade': item.get('grade_description', ''),
                    'head_count': int(parse_number(item.get('head_count', '0'))),
                    'avg_weight': parse_number(item.get('weight_range_avg', '0')),
                    'price_low': parse_number(item.get('price_range_low', '0')),
                    'price_high': parse_number(item.get('price_range_high', '0')),
                    'avg_price': parse_number(item.get('weighted_avg_price', '0')),
                })
        elif sec == 'History':
            for item in items:
                price = item.get('weighted_avg_price', '')
                if not price or str(price) == 'None':
                    continue
                period = item.get('current_period', '')
                if 'WEEKLY WEIGHTED' in str(period).upper():
                    result['five_area'].append({
                        'report_date': item.get('report_date', ''),
                        'class': item.get('class_description', ''),
                        'basis': item.get('selling_basis_desc', ''),
                        'grade': 'Weekly Avg',
                        'head_count': int(parse_number(item.get('head_count', '0'))),
                        'avg_weight': parse_number(item.get('weight_range_avg', '0')),
                        'price_low': 0,
                        'price_high': 0,
                        'avg_price': parse_number(item.get('weighted_avg_price', '0')),
                    })

    # IA-MN Daily
    data_2672 = fetch_datamart(REPORT_IAMN_DAILY)
    for section in data_2672:
        sec = section.get('reportSection', '')
        items = section.get('results', [])
        if sec == 'Summary' and items:
            result['iamn_narrative'] = items[0].get('trend', '')
        elif sec == 'Detail':
            for item in items:
                price = item.get('wtd_avg_price', '')
                if not price or str(price) == 'None':
                    continue
                result['iamn'].append({
                    'report_date': item.get('report_date', ''),
                    'purchase_type': item.get('purchase_type_code', ''),
                    'class': item.get('class_desc', ''),
                    'basis': item.get('selling_basis_desc', ''),
                    'grade': item.get('grade_desc', ''),
                    'head_count': int(parse_number(item.get('head_count', '0'))),
                    'avg_weight': parse_number(item.get('wtd_avg_weight', '0')),
                    'dress_pct': parse_number(item.get('wtd_avg_dress_pct', '0')),
                    'price_low': parse_number(item.get('price_range_low', '0')),
                    'price_high': parse_number(item.get('price_range_high', '0')),
                    'avg_price': parse_number(item.get('wtd_avg_price', '0')),
                })

    print(f"  5 Area: {len(result['five_area'])} price rows")
    print(f"  IA-MN:  {len(result['iamn'])} price rows")

    return result


def fetch_premiums_discounts() -> dict:
    """Fetch National Weekly Premiums & Discounts (2482)."""
    print("Fetching premiums & discounts...")
    result = {'quality': [], 'yield_grade': [], 'weight': [], 'report_date': ''}

    data = fetch_datamart(REPORT_PREMIUMS)
    for section in data:
        if section.get('reportSection') != 'Detail':
            continue
        for item in section.get('results', []):
            row = {
                'type': item.get('type', ''),
                'class': item.get('class_description', ''),
                'avg_price': parse_number(item.get('avg_price', '0')),
                'price_low': parse_number(item.get('price_range_low', '0')),
                'price_high': parse_number(item.get('price_range_high', '0')),
                'price_change': parse_number(item.get('price_change', '0')),
            }
            result['report_date'] = item.get('report_date', '')
            if row['type'] == 'Quality':
                result['quality'].append(row)
            elif row['type'] == 'Yield Grade':
                result['yield_grade'].append(row)
            elif row['type'] == 'Weight':
                result['weight'].append(row)

    print(f"  Quality: {len(result['quality'])} rows | "
          f"YG: {len(result['yield_grade'])} rows | "
          f"Weight: {len(result['weight'])} rows")

    return result


def fetch_indiana_auction(api_key: str) -> dict:
    """Fetch Indiana Weekly Auction (MARS 1976)."""
    print("Fetching Indiana auction (MARS 1976)...")
    result = {
        'slaughter': [], 'feeder': [], 'replacement': [],
        'report_date': '', 'receipts': 0, 'receipts_week_ago': 0,
        'receipts_year_ago': 0,
    }

    if not api_key:
        print("  MARS API key not configured — skipping")
        return result

    try:
        data = fetch_mars(REPORT_IN_AUCTION, api_key)
    except Exception as e:
        print(f"  MARS fetch failed: {e}")
        return result

    rows = data.get('results', [])
    cattle = [r for r in rows if r.get('category') == 'Cattle']

    if cattle:
        result['report_date'] = cattle[0].get('report_date', '')
        result['receipts'] = int(parse_number(cattle[0].get('receipts', '0')))
        result['receipts_week_ago'] = int(parse_number(cattle[0].get('receipts_week_ago', '0')))
        result['receipts_year_ago'] = int(parse_number(cattle[0].get('receipts_year_ago', '0')))

    for r in cattle:
        commodity = r.get('commodity', '') or ''
        row = {
            'commodity': commodity,
            'class': r.get('class', '') or '',
            'quality_grade': r.get('quality_grade_name', '') or '',
            'frame': r.get('frame', '') or '',
            'dressing': r.get('dressing', '') or '',
            'yield_grade': r.get('yield_grade', '') or '',
            'head_count': int(parse_number(r.get('head_count', '0'))),
            'avg_weight': parse_number(r.get('avg_weight', '0')),
            'avg_price_min': parse_number(r.get('avg_price_min', '0')),
            'avg_price_max': parse_number(r.get('avg_price_max', '0')),
            'avg_price': parse_number(r.get('avg_price', '0')),
            'receipts': int(parse_number(r.get('receipts', '0'))),
        }

        if 'Slaughter' in commodity:
            result['slaughter'].append(row)
        elif 'Feeder' in commodity:
            result['feeder'].append(row)
        elif 'Replacement' in commodity:
            result['replacement'].append(row)

    print(f"  Report date: {result['report_date']} | "
          f"Receipts: {result['receipts']} head")
    print(f"  Slaughter: {len(result['slaughter'])} | "
          f"Feeder: {len(result['feeder'])} | "
          f"Replacement: {len(result['replacement'])}")

    return result


# ---------------------------------------------------------------------------
# Deduplication & valuation
# ---------------------------------------------------------------------------

def deduplicate_cuts(cuts: list) -> dict:
    grouped = {}
    for cut in cuts:
        code = cut.imps_code
        if code not in grouped:
            grouped[code] = []
        grouped[code].append(cut)

    result = {}
    for code, items in grouped.items():
        if len(items) == 1:
            result[code] = items[0]
        else:
            total_lbs = sum(c.total_pounds for c in items)
            if total_lbs > 0:
                wavg = sum(c.weighted_avg_cwt * c.total_pounds for c in items) / total_lbs
            else:
                wavg = sum(c.weighted_avg_cwt for c in items) / len(items)
            result[code] = SubprimalPrice(
                imps_code=code,
                description=items[0].description + " (combined)",
                weighted_avg_cwt=wavg,
                price_range_low=min(c.price_range_low for c in items),
                price_range_high=max(c.price_range_high for c in items),
                number_trades=sum(c.number_trades for c in items),
                total_pounds=total_lbs,
                grade=items[0].grade,
            )
    return result


def build_grade_cuts(usda_data: dict, grade: str, grassfed_premium_cwt: float = 0.0) -> dict:
    """
    Build a deduplicated IMPS->SubprimalPrice dict for the given grade.
    For Prime: use actual Prime prices where available, fall back to Choice.
    For Grassfed: use Choice prices + premium.
    """
    grade_lower = grade.lower()

    if grade_lower == 'choice':
        return deduplicate_cuts(usda_data['choice_cuts'])

    elif grade_lower == 'select':
        return deduplicate_cuts(usda_data['select_cuts'])

    elif grade_lower == 'prime':
        prime_by_code = deduplicate_cuts(usda_data['prime_cuts'])
        choice_by_code = deduplicate_cuts(usda_data['choice_cuts'])
        # Merge: Prime where available, Choice fallback
        merged = {}
        for code in set(list(prime_by_code.keys()) + list(choice_by_code.keys())):
            if code in prime_by_code:
                merged[code] = prime_by_code[code]
            elif code in choice_by_code:
                # Use Choice price as fallback for Prime
                cut = choice_by_code[code]
                merged[code] = SubprimalPrice(
                    imps_code=cut.imps_code,
                    description=cut.description + " (Choice fallback)",
                    weighted_avg_cwt=cut.weighted_avg_cwt,
                    price_range_low=cut.price_range_low,
                    price_range_high=cut.price_range_high,
                    number_trades=cut.number_trades,
                    total_pounds=cut.total_pounds,
                    grade='Prime*',
                )
        return merged

    elif grade_lower == 'grassfed':
        choice_by_code = deduplicate_cuts(usda_data['choice_cuts'])
        grassfed = {}
        for code, cut in choice_by_code.items():
            grassfed[code] = SubprimalPrice(
                imps_code=cut.imps_code,
                description=cut.description,
                weighted_avg_cwt=cut.weighted_avg_cwt + grassfed_premium_cwt,
                price_range_low=cut.price_range_low + grassfed_premium_cwt,
                price_range_high=cut.price_range_high + grassfed_premium_cwt,
                number_trades=cut.number_trades,
                total_pounds=cut.total_pounds,
                grade='Grassfed',
            )
        return grassfed

    else:
        raise ValueError(f"Unknown grade: {grade}")


def compute_carcass_value(
    usda_data: dict,
    live_weight: float = DEFAULT_LIVE_WEIGHT,
    yield_grade: int = DEFAULT_YIELD_GRADE,
    quality_grade: str = DEFAULT_QUALITY_GRADE,
    broker_fee_pct: float = DEFAULT_BROKER_FEE_PCT,
    byproduct_pct: float = DEFAULT_BYPRODUCT_PCT,
    byproduct_value_per_lb: float = DEFAULT_BYPRODUCT_VALUE_PER_LB,
    grassfed_premium_cwt: float = DEFAULT_GRASSFED_PREMIUM_CWT,
) -> CarcassValuation:
    cuts_by_code = build_grade_cuts(usda_data, quality_grade, grassfed_premium_cwt)

    dress_pct = DRESS_PCT_BY_YG.get(yield_grade, 0.60)
    hot_carcass_wt = live_weight * dress_pct

    cut_values = []
    total_subprimal_value = 0.0

    for imps_code, (desc, yield_pct, primal) in SUBPRIMAL_YIELDS.items():
        cut_data = cuts_by_code.get(imps_code)
        if cut_data is None:
            continue

        cut_weight = hot_carcass_wt * (yield_pct / 100.0)
        price_per_lb = cut_data.price_per_lb
        cut_value = cut_weight * price_per_lb
        total_subprimal_value += cut_value

        cut_values.append({
            'imps_code': imps_code,
            'description': desc,
            'primal': primal,
            'yield_pct': yield_pct,
            'cut_weight_lbs': round(cut_weight, 1),
            'usda_price_cwt': round(cut_data.weighted_avg_cwt, 2),
            'price_per_lb': round(price_per_lb, 2),
            'cut_value': round(cut_value, 2),
            'trades': cut_data.number_trades,
            'volume_lbs': cut_data.total_pounds,
        })

    byproduct_value = live_weight * byproduct_pct * byproduct_value_per_lb
    total_carcass_value = total_subprimal_value + byproduct_value
    broker_fee = total_carcass_value * broker_fee_pct
    net_value = total_carcass_value - broker_fee
    value_per_cwt_carcass = (total_subprimal_value / hot_carcass_wt) * 100 if hot_carcass_wt > 0 else 0
    value_per_cwt_live = (net_value / live_weight) * 100 if live_weight > 0 else 0

    return CarcassValuation(
        live_weight=live_weight,
        yield_grade=yield_grade,
        quality_grade=quality_grade.title(),
        dressing_pct=dress_pct,
        hot_carcass_weight=hot_carcass_wt,
        report_date=usda_data['report_date'],
        cut_values=sorted(cut_values, key=lambda x: -x['cut_value']),
        total_subprimal_value=round(total_subprimal_value, 2),
        byproduct_value=round(byproduct_value, 2),
        total_carcass_value=round(total_carcass_value, 2),
        value_per_cwt_carcass=round(value_per_cwt_carcass, 2),
        value_per_cwt_live=round(value_per_cwt_live, 2),
        broker_fee=round(broker_fee, 2),
        net_value=round(net_value, 2),
    )


# ---------------------------------------------------------------------------
# Purchase price calculation
# ---------------------------------------------------------------------------

def match_premium(premiums_list: list, search_term: str) -> float:
    search_upper = search_term.upper()
    for row in premiums_list:
        cls = (row.get('class') or '').upper()
        if search_upper in cls:
            return row.get('avg_price', 0.0)
    return 0.0


def weighted_avg_price(rows: list) -> float:
    total_head = sum(r['head_count'] for r in rows)
    if total_head == 0:
        return 0.0
    return sum(r['avg_price'] * r['head_count'] for r in rows) / total_head


def compute_purchase_prices(
    valuation: CarcassValuation,
    slaughter_data: dict,
    premiums_data: dict,
    auction_data: dict,
    processor: dict,
    live_weight: float,
    yield_grade: int,
    quality_grade: str,
) -> PurchasePriceResult:

    dressing_pct = valuation.dressing_pct
    hcw = valuation.hot_carcass_weight

    # ── Method 1: Live Weight Basis ──
    live_rows = [r for r in slaughter_data.get('five_area', [])
                 if 'Live' in str(r.get('basis', '')) and r['head_count'] > 0]
    live_market_ref = weighted_avg_price(live_rows) if live_rows else 0.0

    # Blend Indiana auction slaughter prices if available
    auction_slaughter = auction_data.get('slaughter', [])
    auction_ref = 0.0
    if auction_slaughter:
        auction_with_price = [r for r in auction_slaughter if r.get('avg_price', 0) > 0]
        if auction_with_price:
            auction_ref = weighted_avg_price(auction_with_price)

    if live_market_ref > 0 and auction_ref > 0:
        # Weight 5-Area more heavily (larger sample)
        live_head = sum(r['head_count'] for r in live_rows)
        auction_head = sum(r['head_count'] for r in auction_with_price)
        total = live_head + auction_head
        live_basis = (live_market_ref * live_head + auction_ref * auction_head) / total if total > 0 else live_market_ref
    elif live_market_ref > 0:
        live_basis = live_market_ref
    elif auction_ref > 0:
        live_basis = auction_ref
    else:
        live_basis = 0.0

    # ── Method 2: Dressed/Carcass Weight Basis ──
    dressed_rows = [r for r in slaughter_data.get('five_area', [])
                    if 'Dressed' in str(r.get('basis', '')) and r['head_count'] > 0]
    dressed_market_ref = weighted_avg_price(dressed_rows) if dressed_rows else 0.0
    dressed_basis = dressed_market_ref * dressing_pct if dressed_market_ref > 0 else 0.0

    # ── Method 3: Grid/Formula ──
    grid_base = dressed_market_ref  # dressed $/cwt as base

    grade_lower = quality_grade.lower()
    quality_adj = 0.0
    if grade_lower == 'prime':
        quality_adj = match_premium(premiums_data.get('quality', []), 'Prime')
    elif grade_lower == 'select':
        quality_adj = match_premium(premiums_data.get('quality', []), 'Select')
    elif grade_lower == 'grassfed':
        # Grassfed typically commands Choice+ premium; use Choice as base (0 adj)
        quality_adj = 0.0

    yg_adj = 0.0
    yg_search = f'YG {yield_grade}' if yield_grade <= 3 else f'YG 4'
    yg_adj = match_premium(premiums_data.get('yield_grade', []), yg_search)

    wt_adj = 0.0
    carcass_wt = hcw
    if carcass_wt < 550:
        wt_adj = match_premium(premiums_data.get('weight', []), 'Light')
    elif carcass_wt > 950:
        wt_adj = match_premium(premiums_data.get('weight', []), 'Heavy')

    grid_adjusted = grid_base + quality_adj + yg_adj + wt_adj
    grid_formula = grid_adjusted * dressing_pct if grid_adjusted > 0 else 0.0

    # ── Method 4: Cutout-Minus-Margin ──
    kill_fee = processor.get('kill_fee', 0)
    fab_cost_per_lb = processor.get('fab_cost_per_lb', 0)
    shrink_pct = processor.get('shrink_pct', 0)

    fab_total = fab_cost_per_lb * hcw
    shrink_cost = shrink_pct * valuation.total_subprimal_value
    remainder = valuation.net_value - kill_fee - fab_total - shrink_cost
    cutout_minus = (remainder / live_weight) * 100 if live_weight > 0 else 0.0

    return PurchasePriceResult(
        quality_grade=quality_grade.title(),
        live_weight=live_weight,
        dressing_pct=dressing_pct,
        processor_name=processor.get('name', 'Unknown'),
        processor_costs={
            'kill_fee': kill_fee,
            'fab_cost_per_lb': fab_cost_per_lb,
            'fab_total': round(fab_total, 2),
            'shrink_pct': shrink_pct,
            'shrink_cost': round(shrink_cost, 2),
            'payment_terms_days': processor.get('payment_terms_days', 0),
        },
        live_basis_cwt=round(live_basis, 2),
        dressed_basis_cwt=round(dressed_basis, 2),
        grid_formula_cwt=round(grid_formula, 2),
        cutout_minus_margin_cwt=round(cutout_minus, 2),
        live_market_ref_cwt=round(live_market_ref, 2),
        dressed_market_ref_cwt=round(dressed_market_ref, 2),
        grid_base_cwt=round(grid_base, 2),
        grid_quality_adj=round(quality_adj, 2),
        grid_yg_adj=round(yg_adj, 2),
        grid_wt_adj=round(wt_adj, 2),
        cutout_net_value=round(valuation.net_value, 2),
        cutout_kill_fee=round(kill_fee, 2),
        cutout_fab_total=round(fab_total, 2),
        cutout_shrink_cost=round(shrink_cost, 2),
        auction_ref_cwt=round(auction_ref, 2),
    )


# ---------------------------------------------------------------------------
# Console output
# ---------------------------------------------------------------------------

def print_valuation(val: CarcassValuation, usda_data: dict):
    print("\n" + "=" * 80)
    print("  CATTLE CARCASS VALUATION REPORT")
    print("=" * 80)
    print(f"  USDA Report Date:   {val.report_date}")
    print(f"  Valuation Date:     {datetime.now().strftime('%m/%d/%Y %H:%M')}")
    print(f"  Quality Grade:      {val.quality_grade}")
    print(f"  Yield Grade:        {val.yield_grade}")
    print(f"  Live Weight:        {val.live_weight:,.0f} lbs")
    print(f"  Dressing %:         {val.dressing_pct:.1%}")
    print(f"  Hot Carcass Wt:     {val.hot_carcass_weight:,.0f} lbs")

    wavg = usda_data.get('weekly_avg', {})
    if wavg:
        print(f"\n  USDA Weekly Avg Cutout:")
        print(f"    Choice: ${wavg.get('choice', 0):,.2f}/cwt")
        print(f"    Select: ${wavg.get('select', 0):,.2f}/cwt")

    print(f"\n{'─' * 80}")
    print(f"  {'IMPS':<6} {'Cut Description':<35} {'Wt(lb)':<8} {'$/lb':<8} {'Value':>10}")
    print(f"{'─' * 80}")

    by_primal = {}
    for cv in val.cut_values:
        by_primal.setdefault(cv['primal'], []).append(cv)

    for primal in PRIMAL_ORDER:
        cuts = by_primal.get(primal, [])
        if not cuts:
            continue
        primal_total = sum(c['cut_value'] for c in cuts)
        print(f"\n  {primal.upper()} (subtotal: ${primal_total:,.2f})")
        for cv in sorted(cuts, key=lambda x: -x['cut_value']):
            print(f"  {cv['imps_code']:<6} {cv['description']:<35} {cv['cut_weight_lbs']:>6.1f}  "
                  f"${cv['price_per_lb']:>5.2f}   ${cv['cut_value']:>9,.2f}")

    print(f"\n{'─' * 80}")
    total_cut_wt = sum(c['cut_weight_lbs'] for c in val.cut_values)
    total_yield = (total_cut_wt / val.hot_carcass_weight * 100) if val.hot_carcass_weight > 0 else 0
    print(f"  Total subprimal weight:  {total_cut_wt:,.1f} lbs ({total_yield:.1f}% of carcass)")
    print(f"  Total subprimal value:   ${val.total_subprimal_value:>12,.2f}")
    print(f"  Byproduct value:         ${val.byproduct_value:>12,.2f}")
    print(f"  GROSS CARCASS VALUE:     ${val.total_carcass_value:>12,.2f}")
    print(f"  Broker fee (2%):        -${val.broker_fee:>12,.2f}")
    print(f"  NET CARCASS VALUE:       ${val.net_value:>12,.2f}")
    print(f"\n  Value per cwt (carcass): ${val.value_per_cwt_carcass:>8,.2f}/cwt")
    print(f"  Value per cwt (live):    ${val.value_per_cwt_live:>8,.2f}/cwt")
    print(f"  Value per lb (live):     ${val.value_per_cwt_live / 100:>8,.4f}/lb")
    print("=" * 80)


def print_all_grades_summary(valuations: dict):
    print("\n" + "=" * 100)
    print("  ALL-GRADES COMPARISON")
    print("=" * 100)
    print(f"\n  {'Grade':<12} {'Subprimal $':>14} {'Byprod $':>10} {'Gross $':>12} {'Broker 2%':>10} "
          f"{'NET VALUE':>12} {'$/cwt Live':>12} {'$/cwt Carc':>12}")
    print("  " + "─" * 94)
    for grade in ['Prime', 'Choice', 'Select', 'Grassfed']:
        v = valuations.get(grade)
        if not v:
            continue
        print(f"  {grade:<12} ${v.total_subprimal_value:>12,.2f} ${v.byproduct_value:>8,.2f} "
              f"${v.total_carcass_value:>10,.2f} ${v.broker_fee:>8,.2f} "
              f"${v.net_value:>10,.2f} ${v.value_per_cwt_live:>10,.2f} "
              f"${v.value_per_cwt_carcass:>10,.2f}")
    print("=" * 100)


def print_cut_matrix(usda_data: dict, grassfed_premium_cwt: float):
    """Print the full cut-level price matrix across all grades."""
    prime_cuts = deduplicate_cuts(usda_data['prime_cuts'])
    choice_cuts = deduplicate_cuts(usda_data['choice_cuts'])
    select_cuts = deduplicate_cuts(usda_data['select_cuts'])

    print("\n" + "=" * 130)
    print("  CUT-LEVEL PRICE MATRIX ($/cwt)")
    print("=" * 130)
    print(f"  {'IMPS':<6} {'Description':<30} {'Primal':<9} {'Yld%':<6} "
          f"{'PRIME':>10} {'CHOICE':>10} {'SELECT':>10} {'GRASSFED':>10} "
          f"{'P-C':>8} {'C-S':>8}")
    print("─" * 130)

    current_primal = None
    for imps in sorted(SUBPRIMAL_YIELDS.keys(),
                       key=lambda x: (PRIMAL_ORDER.index(SUBPRIMAL_YIELDS[x][2])
                                      if SUBPRIMAL_YIELDS[x][2] in PRIMAL_ORDER else 99, x)):
        desc, yld, primal = SUBPRIMAL_YIELDS[imps]

        if primal != current_primal:
            current_primal = primal
            print(f"\n  {primal.upper()}")

        p = prime_cuts.get(imps)
        c = choice_cuts.get(imps)
        s = select_cuts.get(imps)

        p_val = p.weighted_avg_cwt if p else None
        c_val = c.weighted_avg_cwt if c else None
        s_val = s.weighted_avg_cwt if s else None
        g_val = c_val + grassfed_premium_cwt if c_val else None

        def fmt(v):
            return f"${v:,.2f}" if v else "   ---"

        pc = f"${p_val - c_val:+,.0f}" if (p_val and c_val) else ""
        cs = f"${c_val - s_val:+,.0f}" if (c_val and s_val) else ""

        print(f"  {imps:<6} {desc:<30} {primal:<9} {yld:>4.1f}%  "
              f"{fmt(p_val):>10} {fmt(c_val):>10} {fmt(s_val):>10} {fmt(g_val):>10} "
              f"{pc:>8} {cs:>8}")

    print("=" * 130)


def print_live_market(slaughter_data: dict, premiums_data: dict, auction_data: dict):
    """Print live cattle market comparison."""
    print("\n" + "=" * 100)
    print("  LIVE CATTLE MARKET")
    print("=" * 100)

    # 5 Area
    five_area = slaughter_data.get('five_area', [])
    if five_area:
        print(f"\n  5 AREA WEEKLY ({five_area[0].get('report_date', '')})")
        print(f"  {'Class':<18} {'Basis':<18} {'Grade':<22} {'Head':>8} {'AvgWt':>7} {'Avg $/cwt':>10}")
        print(f"  {'─' * 85}")
        for r in five_area:
            if r['head_count'] > 0:
                print(f"  {r['class']:<18} {r['basis']:<18} {r['grade']:<22} "
                      f"{r['head_count']:>8,} {r['avg_weight']:>7,.0f} ${r['avg_price']:>8,.2f}")

    # IA-MN
    iamn = slaughter_data.get('iamn', [])
    narrative = slaughter_data.get('iamn_narrative', '')
    if narrative:
        print(f"\n  IA-MN MARKET COMMENTARY:")
        # Wrap narrative
        words = narrative.split()
        line = "    "
        for w in words:
            if len(line) + len(w) > 95:
                print(line)
                line = "    "
            line += w + " "
        if line.strip():
            print(line)

    if iamn:
        print(f"\n  IA-MN PRICED TRANSACTIONS:")
        for r in iamn:
            print(f"    {r['class']:<22} {r['basis']:<16} {r['grade']:<20} "
                  f"head={r['head_count']:<5} ${r['avg_price']:,.2f}/cwt")

    # Premiums
    quality = premiums_data.get('quality', [])
    yg = premiums_data.get('yield_grade', [])
    wt = premiums_data.get('weight', [])
    if quality:
        print(f"\n  PREMIUMS & DISCOUNTS ($/cwt dressed, {premiums_data.get('report_date', '')})")
        print(f"  {'Type':<14} {'Class':<24} {'Avg':>8} {'Range':>20} {'Chg':>6}")
        print(f"  {'─' * 75}")
        for r in quality + yg + wt:
            rng = f"${r['price_low']:,.2f} to ${r['price_high']:,.2f}"
            print(f"  {r['type']:<14} {r['class']:<24} ${r['avg_price']:>6,.2f} {rng:>20} {r['price_change']:>+5,.2f}")

    # Indiana Auction
    if auction_data.get('slaughter') or auction_data.get('feeder'):
        print(f"\n  INDIANA AUCTION (MARS 1976, {auction_data.get('report_date', '')})")
        print(f"  Receipts: {auction_data['receipts']:,} head "
              f"(week ago: {auction_data['receipts_week_ago']:,}, "
              f"year ago: {auction_data['receipts_year_ago']:,})")

        for category, label in [('slaughter', 'SLAUGHTER'), ('feeder', 'FEEDER')]:
            rows = auction_data.get(category, [])
            if not rows:
                continue
            print(f"\n  {label} CATTLE:")
            print(f"  {'Class':<18} {'Grade':<22} {'Dress':<8} {'YG':<6} {'Head':>5} "
                  f"{'AvgWt':>7} {'Lo':>8} {'Hi':>8} {'Avg':>8}")
            print(f"  {'─' * 95}")
            for r in rows:
                print(f"  {r['class']:<18} {r['quality_grade']:<22} {r['dressing']:<8} "
                      f"{r['yield_grade']:<6} {r['head_count']:>5} {r['avg_weight']:>7,.0f} "
                      f"${r['avg_price_min']:>6,.0f} ${r['avg_price_max']:>6,.0f} "
                      f"${r['avg_price']:>7,.2f}")

    print("=" * 100)


def print_purchase_prices(pp: PurchasePriceResult):
    print("\n" + "=" * 80)
    print(f"  PURCHASE PRICE ANALYSIS — {pp.quality_grade} | Processor: {pp.processor_name}")
    print("=" * 80)

    costs = pp.processor_costs
    print(f"  Processor Costs:")
    print(f"    Kill fee:       ${costs['kill_fee']:,.2f}/head")
    print(f"    Fab cost:       ${costs['fab_cost_per_lb']:.2f}/lb (${costs['fab_total']:,.2f} total)")
    print(f"    Cooler shrink:  {costs['shrink_pct']:.1%} (${costs['shrink_cost']:,.2f})")
    print(f"    Payment terms:  Net {costs['payment_terms_days']} days")

    def fmt_method(label, cwt, lw):
        if cwt > 0:
            per_head = cwt * lw / 100
            return f"  {label:<30} ${cwt:>8,.2f}   ${per_head:>10,.2f}"
        return f"  {label:<30} {'N/A':>9}   {'N/A':>11}"

    print(f"\n  {'METHOD':<30} {'$/cwt Live':>9}   {'$/head':>11}")
    print(f"  {'─' * 54}")
    print(fmt_method("1. Live Weight Basis", pp.live_basis_cwt, pp.live_weight))
    print(fmt_method("2. Dressed Basis (→ live)", pp.dressed_basis_cwt, pp.live_weight))
    print(fmt_method("3. Grid/Formula (→ live)", pp.grid_formula_cwt, pp.live_weight))
    print(fmt_method("4. Cutout-Minus-Margin", pp.cutout_minus_margin_cwt, pp.live_weight))
    print(f"  {'─' * 54}")

    methods = [pp.live_basis_cwt, pp.dressed_basis_cwt, pp.grid_formula_cwt, pp.cutout_minus_margin_cwt]
    valid = [m for m in methods if m > 0]
    if valid:
        hi, lo = max(valid), min(valid)
        avg = sum(valid) / len(valid)
        avg_head = avg * pp.live_weight / 100
        print(f"  {'Spread (high - low):':<30} ${hi - lo:>8,.2f}   ${(hi - lo) * pp.live_weight / 100:>10,.2f}")
        print(f"  {'Average of methods:':<30} ${avg:>8,.2f}   ${avg_head:>10,.2f}")

    # Derivation notes
    print(f"\n  Derivations:")
    if pp.live_basis_cwt > 0:
        note = f"5 Area live ${pp.live_market_ref_cwt:,.2f}"
        if pp.auction_ref_cwt > 0:
            note += f", IN auction ${pp.auction_ref_cwt:,.2f}"
        print(f"    M1: {note}")
    if pp.dressed_basis_cwt > 0:
        print(f"    M2: 5 Area dressed ${pp.dressed_market_ref_cwt:,.2f}/cwt × {pp.dressing_pct:.1%} dress")
    if pp.grid_formula_cwt > 0:
        adj_parts = []
        if pp.grid_quality_adj != 0:
            adj_parts.append(f"quality {pp.grid_quality_adj:+,.2f}")
        if pp.grid_yg_adj != 0:
            adj_parts.append(f"YG {pp.grid_yg_adj:+,.2f}")
        if pp.grid_wt_adj != 0:
            adj_parts.append(f"wt {pp.grid_wt_adj:+,.2f}")
        adj_str = ", ".join(adj_parts) if adj_parts else "no adjustments"
        print(f"    M3: base ${pp.grid_base_cwt:,.2f} + {adj_str} → ${pp.grid_base_cwt + pp.grid_quality_adj + pp.grid_yg_adj + pp.grid_wt_adj:,.2f} dressed × {pp.dressing_pct:.1%}")
    if pp.cutout_minus_margin_cwt > 0:
        print(f"    M4: net ${pp.cutout_net_value:,.2f} - kill ${pp.cutout_kill_fee:,.2f}"
              f" - fab ${pp.cutout_fab_total:,.2f} - shrink ${pp.cutout_shrink_cost:,.2f}"
              f" = ${pp.cutout_net_value - pp.cutout_kill_fee - pp.cutout_fab_total - pp.cutout_shrink_cost:,.2f}")
    print("=" * 80)


def print_purchase_summary(purchase_results: dict):
    print("\n" + "=" * 100)
    print("  PURCHASE PRICE COMPARISON — ALL GRADES ($/cwt live)")
    print("=" * 100)
    print(f"\n  {'Grade':<12} {'Live Basis':>12} {'Dressed→Live':>14} {'Grid/Formula':>14} {'Cutout-Margin':>15} {'Average':>12}")
    print("  " + "─" * 81)

    def fmt(v):
        return f"${v:>8,.2f}" if v > 0 else "     N/A"

    for grade in ['Prime', 'Choice', 'Select', 'Grassfed']:
        pp = purchase_results.get(grade)
        if not pp:
            continue
        methods = [pp.live_basis_cwt, pp.dressed_basis_cwt, pp.grid_formula_cwt, pp.cutout_minus_margin_cwt]
        valid = [m for m in methods if m > 0]
        avg = sum(valid) / len(valid) if valid else 0.0
        print(f"  {grade:<12} {fmt(pp.live_basis_cwt):>12} {fmt(pp.dressed_basis_cwt):>14} "
              f"{fmt(pp.grid_formula_cwt):>14} {fmt(pp.cutout_minus_margin_cwt):>15} {fmt(avg):>12}")
    print("=" * 100)


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------

def write_excel(valuations: dict, usda_data: dict, slaughter_data: dict,
                premiums_data: dict, auction_data: dict, grassfed_premium_cwt: float,
                filepath: str, purchase_results: dict = None):
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        print("\nopenpyxl not installed. Run: pip3 install openpyxl")
        return

    wb = openpyxl.Workbook()
    hdr_font = Font(bold=True, size=14)
    sub_font = Font(bold=True, size=11)
    hdr_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    hdr_text = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(bottom=Side(style='thin', color='CCCCCC'))

    report_date = usda_data.get('report_date', '')

    # ── Sheet 1: Valuation Summary (all grades side by side) ──
    ws = wb.active
    ws.title = "Valuation Summary"
    ws['A1'] = 'CATTLE CARCASS VALUATION — ALL GRADES'
    ws['A1'].font = hdr_font
    ws['A2'] = f'USDA Report: {report_date} | Generated: {datetime.now().strftime("%m/%d/%Y")}'
    ws['A2'].font = Font(italic=True, size=10, color="666666")

    # Pick a reference valuation for shared params
    ref = next(iter(valuations.values()))
    row = 4
    ws.cell(row=row, column=1, value='INPUT PARAMETERS').font = sub_font
    for label, val in [('Live Weight', f"{ref.live_weight:,.0f} lbs"),
                       ('Yield Grade', ref.yield_grade),
                       ('Dressing %', f"{ref.dressing_pct:.1%}"),
                       ('Hot Carcass Wt', f"{ref.hot_carcass_weight:,.0f} lbs")]:
        row += 1
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=val)

    row += 2
    ws.cell(row=row, column=1, value='GRADE COMPARISON').font = sub_font
    row += 1
    grade_order = ['Prime', 'Choice', 'Select', 'Grassfed']
    headers = ['Metric'] + [g for g in grade_order if g in valuations]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = hdr_text
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center')

    metrics = [
        ('Total Subprimal Value', 'total_subprimal_value'),
        ('Byproduct Value', 'byproduct_value'),
        ('GROSS CARCASS VALUE', 'total_carcass_value'),
        ('Broker Fee (2%)', 'broker_fee'),
        ('NET CARCASS VALUE', 'net_value'),
        ('$/cwt (carcass)', 'value_per_cwt_carcass'),
        ('$/cwt (live)', 'value_per_cwt_live'),
    ]
    for label, attr in metrics:
        row += 1
        ws.cell(row=row, column=1, value=label)
        is_bold = 'NET' in label or 'GROSS' in label
        if is_bold:
            ws.cell(row=row, column=1).font = Font(bold=True)
        for c, grade in enumerate([g for g in grade_order if g in valuations], 2):
            v = getattr(valuations[grade], attr)
            if attr == 'broker_fee':
                v = -v
            cell = ws.cell(row=row, column=c, value=v)
            cell.number_format = '$#,##0.00'
            if is_bold:
                cell.font = Font(bold=True, size=12)
            cell.border = thin_border

    ws.column_dimensions['A'].width = 26
    for c in range(2, 6):
        ws.column_dimensions[chr(64 + c)].width = 16

    # ── Sheet 2: Cut Matrix ──
    ws2 = wb.create_sheet("Cut Matrix")
    ws2['A1'] = 'CUT-LEVEL PRICE MATRIX ($/cwt)'
    ws2['A1'].font = hdr_font
    ws2['A2'] = f'Report Date: {report_date}'

    prime_cuts = deduplicate_cuts(usda_data['prime_cuts'])
    choice_cuts = deduplicate_cuts(usda_data['choice_cuts'])
    select_cuts = deduplicate_cuts(usda_data['select_cuts'])

    row = 4
    matrix_headers = ['IMPS', 'Description', 'Primal', 'Yield %',
                      'Prime $/cwt', 'Choice $/cwt', 'Select $/cwt', 'Grassfed $/cwt',
                      'P-C Spread', 'C-S Spread']
    for c, h in enumerate(matrix_headers, 1):
        cell = ws2.cell(row=row, column=c, value=h)
        cell.font = hdr_text
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center')

    for imps in sorted(SUBPRIMAL_YIELDS.keys(),
                       key=lambda x: (PRIMAL_ORDER.index(SUBPRIMAL_YIELDS[x][2])
                                      if SUBPRIMAL_YIELDS[x][2] in PRIMAL_ORDER else 99, x)):
        desc, yld, primal = SUBPRIMAL_YIELDS[imps]
        row += 1

        p = prime_cuts.get(imps)
        ch = choice_cuts.get(imps)
        s = select_cuts.get(imps)
        p_val = p.weighted_avg_cwt if p else None
        c_val = ch.weighted_avg_cwt if ch else None
        s_val = s.weighted_avg_cwt if s else None
        g_val = c_val + grassfed_premium_cwt if c_val else None

        ws2.cell(row=row, column=1, value=imps)
        ws2.cell(row=row, column=2, value=desc)
        ws2.cell(row=row, column=3, value=primal)
        ws2.cell(row=row, column=4, value=yld / 100).number_format = '0.00%'

        for col, val in [(5, p_val), (6, c_val), (7, s_val), (8, g_val)]:
            if val:
                ws2.cell(row=row, column=col, value=val).number_format = '$#,##0.00'

        if p_val and c_val:
            ws2.cell(row=row, column=9, value=p_val - c_val).number_format = '$#,##0.00'
        if c_val and s_val:
            ws2.cell(row=row, column=10, value=c_val - s_val).number_format = '$#,##0.00'

        for c in range(1, 11):
            ws2.cell(row=row, column=c).border = thin_border

    ws2.column_dimensions['A'].width = 8
    ws2.column_dimensions['B'].width = 32
    ws2.column_dimensions['C'].width = 10
    ws2.column_dimensions['D'].width = 10
    for c in ['E', 'F', 'G', 'H', 'I', 'J']:
        ws2.column_dimensions[c].width = 14

    # ── Sheet 3: Live Market ──
    ws3 = wb.create_sheet("Live Market")
    ws3['A1'] = '5 AREA & IA-MN SLAUGHTER CATTLE PRICES'
    ws3['A1'].font = hdr_font

    row = 3
    five_area = slaughter_data.get('five_area', [])
    if five_area:
        headers_3 = ['Class', 'Basis', 'Grade', 'Head', 'Avg Wt', 'Avg $/cwt']
        for c, h in enumerate(headers_3, 1):
            cell = ws3.cell(row=row, column=c, value=h)
            cell.font = hdr_text
            cell.fill = hdr_fill
        for r in five_area:
            if r['head_count'] > 0:
                row += 1
                ws3.cell(row=row, column=1, value=r['class'])
                ws3.cell(row=row, column=2, value=r['basis'])
                ws3.cell(row=row, column=3, value=r['grade'])
                ws3.cell(row=row, column=4, value=r['head_count']).number_format = '#,##0'
                ws3.cell(row=row, column=5, value=r['avg_weight']).number_format = '#,##0'
                ws3.cell(row=row, column=6, value=r['avg_price']).number_format = '$#,##0.00'

    ws3.column_dimensions['A'].width = 20
    ws3.column_dimensions['B'].width = 20
    ws3.column_dimensions['C'].width = 24
    ws3.column_dimensions['D'].width = 10
    ws3.column_dimensions['E'].width = 10
    ws3.column_dimensions['F'].width = 14

    # ── Sheet 4: Premiums & Discounts ──
    ws4 = wb.create_sheet("Premiums & Discounts")
    ws4['A1'] = 'NATIONAL WEEKLY PREMIUMS & DISCOUNTS ($/cwt dressed)'
    ws4['A1'].font = hdr_font
    ws4['A2'] = f'Report Date: {premiums_data.get("report_date", "")}'

    row = 4
    headers_4 = ['Type', 'Class', 'Avg $/cwt', 'Range Low', 'Range High', 'Change']
    for c, h in enumerate(headers_4, 1):
        cell = ws4.cell(row=row, column=c, value=h)
        cell.font = hdr_text
        cell.fill = hdr_fill

    all_premiums = premiums_data.get('quality', []) + premiums_data.get('yield_grade', []) + premiums_data.get('weight', [])
    for r in all_premiums:
        row += 1
        ws4.cell(row=row, column=1, value=r['type'])
        ws4.cell(row=row, column=2, value=r['class'])
        ws4.cell(row=row, column=3, value=r['avg_price']).number_format = '$#,##0.00'
        ws4.cell(row=row, column=4, value=r['price_low']).number_format = '$#,##0.00'
        ws4.cell(row=row, column=5, value=r['price_high']).number_format = '$#,##0.00'
        ws4.cell(row=row, column=6, value=r['price_change']).number_format = '+#,##0.00;-#,##0.00'

    ws4.column_dimensions['A'].width = 14
    ws4.column_dimensions['B'].width = 26
    for c in ['C', 'D', 'E', 'F']:
        ws4.column_dimensions[c].width = 14

    # ── Sheet 5: USDA Composites ──
    ws5 = wb.create_sheet("USDA Composites")
    ws5['A1'] = 'Weekly Composite Primal Values ($/cwt, 600-900 lb carcass)'
    ws5['A1'].font = sub_font
    ws5['A2'] = f'Report Date: {report_date}'

    row = 4
    for c, h in enumerate(['Primal', 'Choice $/cwt', 'Select $/cwt', 'Spread'], 1):
        cell = ws5.cell(row=row, column=c, value=h)
        cell.font = hdr_text
        cell.fill = hdr_fill

    for primal, vals in usda_data.get('composites', {}).items():
        row += 1
        ws5.cell(row=row, column=1, value=primal)
        ws5.cell(row=row, column=2, value=vals['choice']).number_format = '$#,##0.00'
        ws5.cell(row=row, column=3, value=vals['select']).number_format = '$#,##0.00'
        ws5.cell(row=row, column=4, value=vals['choice'] - vals['select']).number_format = '$#,##0.00'

    row += 2
    wavg = usda_data.get('weekly_avg', {})
    ws5.cell(row=row, column=1, value='Weekly Average Cutout').font = Font(bold=True)
    row += 1
    ws5.cell(row=row, column=1, value='Choice')
    ws5.cell(row=row, column=2, value=wavg.get('choice', 0)).number_format = '$#,##0.00'
    row += 1
    ws5.cell(row=row, column=1, value='Select')
    ws5.cell(row=row, column=2, value=wavg.get('select', 0)).number_format = '$#,##0.00'
    row += 1
    ws5.cell(row=row, column=1, value='Choice-Select Spread').font = Font(bold=True)
    ws5.cell(row=row, column=2, value=wavg.get('choice', 0) - wavg.get('select', 0)).number_format = '$#,##0.00'

    ws5.column_dimensions['A'].width = 28
    ws5.column_dimensions['B'].width = 16
    ws5.column_dimensions['C'].width = 16
    ws5.column_dimensions['D'].width = 14

    # ── Sheet 6: Ground Beef ──
    ws6 = wb.create_sheet("Ground Beef")
    ws6['A1'] = 'Ground Beef Prices'
    ws6['A1'].font = sub_font

    row = 3
    for c, h in enumerate(['Product', '$/cwt', 'Volume (lbs)'], 1):
        cell = ws6.cell(row=row, column=c, value=h)
        cell.font = hdr_text
        cell.fill = hdr_fill

    for gb in usda_data.get('ground_beef', []):
        row += 1
        ws6.cell(row=row, column=1, value=gb['description'])
        ws6.cell(row=row, column=2, value=gb['weighted_avg_cwt']).number_format = '$#,##0.00'
        ws6.cell(row=row, column=3, value=gb['total_pounds']).number_format = '#,##0'

    ws6.column_dimensions['A'].width = 25
    ws6.column_dimensions['B'].width = 14
    ws6.column_dimensions['C'].width = 16

    # ── Sheet 7: Indiana Auction ──
    ws7 = wb.create_sheet("Indiana Auction")
    ws7['A1'] = 'INDIANA WEEKLY AUCTION SUMMARY (MARS 1976)'
    ws7['A1'].font = hdr_font
    ws7['A2'] = f'Report Date: {auction_data.get("report_date", "N/A")}'
    ws7['A3'] = (f'Receipts: {auction_data.get("receipts", 0):,} head | '
                 f'Week ago: {auction_data.get("receipts_week_ago", 0):,} | '
                 f'Year ago: {auction_data.get("receipts_year_ago", 0):,}')

    for category, label, start_row in [('slaughter', 'SLAUGHTER CATTLE', 5),
                                        ('feeder', 'FEEDER CATTLE', None),
                                        ('replacement', 'REPLACEMENT DAIRY', None)]:
        rows_data = auction_data.get(category, [])
        if not rows_data:
            continue

        if start_row is None:
            start_row = row + 2
        row = start_row

        ws7.cell(row=row, column=1, value=label).font = sub_font
        row += 1
        auction_headers = ['Class', 'Grade', 'Frame', 'Dressing', 'YG', 'Head',
                           'Avg Wt', 'Price Lo', 'Price Hi', 'Price Avg']
        for c, h in enumerate(auction_headers, 1):
            cell = ws7.cell(row=row, column=c, value=h)
            cell.font = hdr_text
            cell.fill = hdr_fill

        for r in rows_data:
            row += 1
            ws7.cell(row=row, column=1, value=r['class'])
            ws7.cell(row=row, column=2, value=r['quality_grade'])
            ws7.cell(row=row, column=3, value=r['frame'])
            ws7.cell(row=row, column=4, value=r['dressing'])
            ws7.cell(row=row, column=5, value=r['yield_grade'])
            ws7.cell(row=row, column=6, value=r['head_count']).number_format = '#,##0'
            ws7.cell(row=row, column=7, value=r['avg_weight']).number_format = '#,##0'
            ws7.cell(row=row, column=8, value=r['avg_price_min']).number_format = '$#,##0.00'
            ws7.cell(row=row, column=9, value=r['avg_price_max']).number_format = '$#,##0.00'
            ws7.cell(row=row, column=10, value=r['avg_price']).number_format = '$#,##0.00'

    ws7.column_dimensions['A'].width = 20
    ws7.column_dimensions['B'].width = 24
    ws7.column_dimensions['C'].width = 18
    ws7.column_dimensions['D'].width = 10
    ws7.column_dimensions['E'].width = 8
    ws7.column_dimensions['F'].width = 8
    ws7.column_dimensions['G'].width = 10
    for c in ['H', 'I', 'J']:
        ws7.column_dimensions[c].width = 12

    # ── Sheet 8: Purchase Price Analysis ──
    if purchase_results:
        ws8 = wb.create_sheet("Purchase Price")
        ws8['A1'] = 'PURCHASE PRICE ANALYSIS'
        ws8['A1'].font = hdr_font

        # Get processor info from first result
        first_pp = next(iter(purchase_results.values()))
        costs = first_pp.processor_costs
        ws8['A2'] = f'Processor: {first_pp.processor_name} | Report: {report_date}'
        ws8['A2'].font = Font(italic=True, size=10, color="666666")

        row = 4
        ws8.cell(row=row, column=1, value='PROCESSOR COSTS').font = sub_font
        for label, val_str in [
            ('Kill Fee', f"${costs['kill_fee']:,.2f}/head"),
            ('Fab Cost', f"${costs['fab_cost_per_lb']:.2f}/lb (${costs['fab_total']:,.2f} total)"),
            ('Cooler Shrink', f"{costs['shrink_pct']:.1%} (${costs['shrink_cost']:,.2f})"),
            ('Payment Terms', f"Net {costs['payment_terms_days']} days"),
        ]:
            row += 1
            ws8.cell(row=row, column=1, value=label)
            ws8.cell(row=row, column=2, value=val_str)

        row += 2
        ws8.cell(row=row, column=1, value='PURCHASE PRICE BY METHOD ($/cwt live)').font = sub_font
        row += 1
        grade_order = ['Prime', 'Choice', 'Select', 'Grassfed']
        pp_headers = ['Method'] + [g for g in grade_order if g in purchase_results]
        for c, h in enumerate(pp_headers, 1):
            cell = ws8.cell(row=row, column=c, value=h)
            cell.font = hdr_text
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal='center')

        method_attrs = [
            ('1. Live Weight Basis', 'live_basis_cwt'),
            ('2. Dressed Basis (→ live)', 'dressed_basis_cwt'),
            ('3. Grid/Formula (→ live)', 'grid_formula_cwt'),
            ('4. Cutout-Minus-Margin', 'cutout_minus_margin_cwt'),
        ]
        for label, attr in method_attrs:
            row += 1
            ws8.cell(row=row, column=1, value=label)
            for c, grade in enumerate([g for g in grade_order if g in purchase_results], 2):
                pp = purchase_results[grade]
                val = getattr(pp, attr)
                if val > 0:
                    cell = ws8.cell(row=row, column=c, value=val)
                    cell.number_format = '$#,##0.00'
                else:
                    ws8.cell(row=row, column=c, value='N/A')
                ws8.cell(row=row, column=c).border = thin_border

        # Average row
        row += 1
        ws8.cell(row=row, column=1, value='Average of Methods').font = Font(bold=True)
        for c, grade in enumerate([g for g in grade_order if g in purchase_results], 2):
            pp = purchase_results[grade]
            methods = [pp.live_basis_cwt, pp.dressed_basis_cwt, pp.grid_formula_cwt, pp.cutout_minus_margin_cwt]
            valid = [m for m in methods if m > 0]
            if valid:
                cell = ws8.cell(row=row, column=c, value=sum(valid) / len(valid))
                cell.number_format = '$#,##0.00'
                cell.font = Font(bold=True, size=12)

        # $/head section
        row += 2
        ws8.cell(row=row, column=1, value='PURCHASE PRICE BY METHOD ($/head)').font = sub_font
        row += 1
        for c, h in enumerate(pp_headers, 1):
            cell = ws8.cell(row=row, column=c, value=h)
            cell.font = hdr_text
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal='center')

        for label, attr in method_attrs:
            row += 1
            ws8.cell(row=row, column=1, value=label)
            for c, grade in enumerate([g for g in grade_order if g in purchase_results], 2):
                pp = purchase_results[grade]
                cwt_val = getattr(pp, attr)
                if cwt_val > 0:
                    per_head = cwt_val * pp.live_weight / 100
                    cell = ws8.cell(row=row, column=c, value=per_head)
                    cell.number_format = '$#,##0.00'
                else:
                    ws8.cell(row=row, column=c, value='N/A')
                ws8.cell(row=row, column=c).border = thin_border

        ws8.column_dimensions['A'].width = 30
        for c in range(2, 7):
            ws8.column_dimensions[chr(64 + c)].width = 16

    wb.save(filepath)
    print(f"\nExcel workbook saved to: {filepath}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description='Cattle Carcass Valuation Engine v2')
    parser.add_argument('--live-weight', type=float, default=DEFAULT_LIVE_WEIGHT,
                        help=f'Live weight in lbs (default: {DEFAULT_LIVE_WEIGHT})')
    parser.add_argument('--yield-grade', type=int, default=DEFAULT_YIELD_GRADE,
                        choices=[1, 2, 3, 4, 5],
                        help=f'USDA Yield Grade 1-5 (default: {DEFAULT_YIELD_GRADE})')
    parser.add_argument('--quality-grade', type=str, default=DEFAULT_QUALITY_GRADE,
                        choices=['prime', 'choice', 'select', 'grassfed'],
                        help=f'Quality grade (default: {DEFAULT_QUALITY_GRADE})')
    parser.add_argument('--all-grades', action='store_true',
                        help='Compute all 4 grades side by side')
    parser.add_argument('--grassfed-premium', type=float, default=DEFAULT_GRASSFED_PREMIUM_CWT,
                        help=f'Grassfed premium $/cwt over Choice (default: {DEFAULT_GRASSFED_PREMIUM_CWT})')
    parser.add_argument('--broker-fee', type=float, default=DEFAULT_BROKER_FEE_PCT,
                        help=f'Broker fee as decimal (default: {DEFAULT_BROKER_FEE_PCT})')
    parser.add_argument('--output', type=str, default=None,
                        help='Output Excel file path')
    parser.add_argument('--no-excel', action='store_true',
                        help='Skip Excel output')
    parser.add_argument('--no-mars', action='store_true',
                        help='Skip MARS Indiana auction pull')
    parser.add_argument('--save-db', action='store_true',
                        help='Save data to PostgreSQL')
    parser.add_argument('--processor', type=str, default=None,
                        help=f'Processor profile name (available: {", ".join(PROCESSORS.keys())})')
    parser.add_argument('--kill-fee', type=float, default=None,
                        help='Override kill fee $/head')
    parser.add_argument('--fab-cost', type=float, default=None,
                        help='Override fabrication cost $/lb carcass')
    parser.add_argument('--shrink-pct', type=float, default=None,
                        help='Override cooler shrink %% (e.g., 0.025 for 2.5%%)')
    args = parser.parse_args()

    # ── Fetch all data ──
    usda_data = fetch_boxed_beef()
    slaughter_data = fetch_slaughter_cattle()
    premiums_data = fetch_premiums_discounts()

    auction_data = {'slaughter': [], 'feeder': [], 'replacement': [],
                    'report_date': '', 'receipts': 0,
                    'receipts_week_ago': 0, 'receipts_year_ago': 0}
    if not args.no_mars:
        auction_data = fetch_indiana_auction(MARS_API_KEY)

    # ── Compute valuations ──
    grades = ['prime', 'choice', 'select', 'grassfed'] if args.all_grades else [args.quality_grade]
    valuations = {}
    for grade in grades:
        val = compute_carcass_value(
            usda_data,
            live_weight=args.live_weight,
            yield_grade=args.yield_grade,
            quality_grade=grade,
            broker_fee_pct=args.broker_fee,
            grassfed_premium_cwt=args.grassfed_premium,
        )
        valuations[val.quality_grade] = val

    # ── Resolve processor and compute purchase prices ──
    processor_key = args.processor or DEFAULT_PROCESSOR
    if processor_key not in PROCESSORS:
        print(f"\nUnknown processor '{processor_key}'. Available: {list(PROCESSORS.keys())}")
        return
    processor = dict(PROCESSORS[processor_key])
    if args.kill_fee is not None:
        processor['kill_fee'] = args.kill_fee
    if args.fab_cost is not None:
        processor['fab_cost_per_lb'] = args.fab_cost
    if args.shrink_pct is not None:
        processor['shrink_pct'] = args.shrink_pct

    purchase_results = {}
    for grade, val in valuations.items():
        purchase_results[grade] = compute_purchase_prices(
            valuation=val,
            slaughter_data=slaughter_data,
            premiums_data=premiums_data,
            auction_data=auction_data,
            processor=processor,
            live_weight=args.live_weight,
            yield_grade=args.yield_grade,
            quality_grade=grade.lower(),
        )

    # ── Console output ──
    if args.all_grades:
        # Print cut matrix
        print_cut_matrix(usda_data, args.grassfed_premium)
        # Print each grade's detail
        for grade in ['Prime', 'Choice', 'Select', 'Grassfed']:
            if grade in valuations:
                print_valuation(valuations[grade], usda_data)
        # Print comparison summary
        print_all_grades_summary(valuations)
    else:
        val = next(iter(valuations.values()))
        print_valuation(val, usda_data)

    # Print live market
    print_live_market(slaughter_data, premiums_data, auction_data)

    # Print purchase price analysis
    if args.all_grades:
        for grade in ['Prime', 'Choice', 'Select', 'Grassfed']:
            if grade in purchase_results:
                print_purchase_prices(purchase_results[grade])
        print_purchase_summary(purchase_results)
    else:
        pp = next(iter(purchase_results.values()))
        print_purchase_prices(pp)

    # ── Excel output ──
    if not args.no_excel:
        filepath = args.output or f'valuation_{datetime.now().strftime("%Y%m%d")}.xlsx'
        write_excel(valuations, usda_data, slaughter_data, premiums_data,
                    auction_data, args.grassfed_premium, filepath, purchase_results)

    # ── Database persistence ──
    if args.save_db:
        try:
            import db
            db.init_schema()
            rd = usda_data['report_date']
            db.save_subprimal_prices(rd, REPORT_CHOICE_SELECT, 'Choice', usda_data['choice_cuts'])
            db.save_subprimal_prices(rd, REPORT_CHOICE_SELECT, 'Select', usda_data['select_cuts'])
            db.save_subprimal_prices(rd, REPORT_PRIME, 'Prime', usda_data['prime_cuts'])
            db.save_composites(rd, usda_data.get('composites', {}))
            # Save 5 Area
            if slaughter_data['five_area']:
                db.save_slaughter_cattle(
                    slaughter_data['five_area'][0].get('report_date', rd),
                    REPORT_5AREA_WEEKLY, '5 Area Weekly', slaughter_data['five_area'])
            # Save premiums
            if premiums_data.get('quality'):
                all_prem = premiums_data['quality'] + premiums_data['yield_grade'] + premiums_data['weight']
                db.save_premiums_discounts(premiums_data['report_date'], all_prem)
            # Save auction
            if auction_data.get('slaughter'):
                all_auction = auction_data['slaughter'] + auction_data['feeder'] + auction_data['replacement']
                db.save_indiana_auction(auction_data['report_date'], all_auction)
            # Save valuations
            for val in valuations.values():
                db.save_valuation(val)
            # Save purchase prices
            for pp in purchase_results.values():
                db.save_purchase_prices(rd, pp, args.yield_grade)
            print("\nData saved to PostgreSQL.")
        except ImportError:
            print("\npsycopg2 not installed — skipping DB save. Run: pip3 install psycopg2-binary")
        except Exception as e:
            print(f"\nDB save failed: {e}")


if __name__ == '__main__':
    main()
