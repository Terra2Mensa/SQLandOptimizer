"""Demand aggregation, carcass allocation, margin calculation, and output."""
from dataclasses import dataclass, field
from typing import List, Optional

from config import (
    SUBPRIMAL_YIELDS, GROUND_BEEF_PRODUCTS, PRIMAL_ORDER,
    REGIONS, GRADE_RANK, TRIM_YIELD_PCT, PROCESSORS,
    DEFAULT_REGION, DEFAULT_PROCESSOR,
)
from buyers import (
    BuyerProfile, compute_buyer_price, _get_base_price_cwt,
)


# ---------------------------------------------------------------------------
# Dataclasses
# ---------------------------------------------------------------------------

@dataclass
class DemandLine:
    cut_code: str
    description: str
    primal: str
    total_weekly_demand_lbs: float
    num_buyers: int
    highest_price_lb: float
    lowest_price_lb: float
    avg_price_lb: float
    carcass_yield_lbs: float
    supply_ratio: float  # demand / yield (how many animals needed)
    buyer_details: list = field(default_factory=list)


@dataclass
class AllocationLine:
    buyer_id: str
    buyer_name: str
    cut_code: str
    description: str
    lbs_allocated: float
    price_per_lb: float
    line_revenue: float


@dataclass
class AllocationResult:
    quality_grade: str
    hcw: float
    live_weight: float
    allocations: List[AllocationLine] = field(default_factory=list)
    unallocated: list = field(default_factory=list)
    total_revenue: float = 0.0
    unallocated_value: float = 0.0
    farmer_cost: float = 0.0
    processing_cost: float = 0.0
    gross_margin: float = 0.0
    margin_pct: float = 0.0


# ---------------------------------------------------------------------------
# Demand aggregation
# ---------------------------------------------------------------------------

def _cut_description(code: str) -> tuple:
    if code in SUBPRIMAL_YIELDS:
        desc, _, primal = SUBPRIMAL_YIELDS[code]
        return desc, primal
    if code in GROUND_BEEF_PRODUCTS:
        return GROUND_BEEF_PRODUCTS[code]["description"], "Trim"
    return code, "Other"


def aggregate_demand(
    buyers: List[BuyerProfile],
    usda_prices: dict,
    ground_beef_prices: dict,
    hcw: float,
    region: str = DEFAULT_REGION,
    quality_grade: str = "choice",
) -> List[DemandLine]:
    regional_adj = REGIONS.get(region, REGIONS[DEFAULT_REGION])["pricing_adjustment"]
    grade_rank = GRADE_RANK.get(quality_grade.lower(), 2)

    # Collect demand per cut across all active, grade-eligible buyers
    cut_demand = {}  # code -> list of {buyer_id, name, volume, price}
    for buyer in buyers:
        if not buyer.active:
            continue
        buyer_grade_rank = GRADE_RANK.get(buyer.min_quality_grade.lower(), 1)
        if grade_rank < buyer_grade_rank:
            continue  # this grade doesn't meet buyer's minimum

        for pref in buyer.cut_preferences:
            base_cwt = _get_base_price_cwt(pref.cut_code, usda_prices, ground_beef_prices)
            if base_cwt <= 0:
                continue
            price_lb = compute_buyer_price(pref.cut_code, pref, base_cwt, regional_adj)

            if pref.cut_code not in cut_demand:
                cut_demand[pref.cut_code] = []
            cut_demand[pref.cut_code].append({
                "buyer_id": buyer.buyer_id,
                "buyer_name": buyer.name,
                "volume_lbs_week": pref.volume_lbs_week,
                "price_lb": price_lb,
            })

    # Build DemandLine objects
    lines = []
    for code, details in cut_demand.items():
        desc, primal = _cut_description(code)
        total_demand = sum(d["volume_lbs_week"] for d in details)
        prices = [d["price_lb"] for d in details]

        # Carcass yield for this cut
        if code in SUBPRIMAL_YIELDS:
            yield_pct = SUBPRIMAL_YIELDS[code][1] / 100.0
            carcass_yield = hcw * yield_pct
        elif code in GROUND_BEEF_PRODUCTS:
            carcass_yield = hcw * TRIM_YIELD_PCT
        else:
            carcass_yield = 0.0

        supply_ratio = total_demand / carcass_yield if carcass_yield > 0 else 999.0

        lines.append(DemandLine(
            cut_code=code,
            description=desc,
            primal=primal,
            total_weekly_demand_lbs=round(total_demand, 1),
            num_buyers=len(details),
            highest_price_lb=round(max(prices), 2),
            lowest_price_lb=round(min(prices), 2),
            avg_price_lb=round(sum(prices) / len(prices), 2),
            carcass_yield_lbs=round(carcass_yield, 1),
            supply_ratio=round(supply_ratio, 1),
            buyer_details=details,
        ))

    lines.sort(key=lambda x: -x.supply_ratio)
    return lines


def compute_animals_needed(demand_lines: List[DemandLine]) -> dict:
    if not demand_lines:
        return {"animals_needed": 0, "bottleneck": "N/A", "excess": {}}
    worst = max(demand_lines, key=lambda x: x.supply_ratio)
    animals = worst.supply_ratio
    excess = {}
    for line in demand_lines:
        if line.carcass_yield_lbs > 0 and animals > 0:
            produced = line.carcass_yield_lbs * animals
            surplus = produced - line.total_weekly_demand_lbs
            if surplus > 0:
                excess[line.cut_code] = round(surplus, 1)
    return {
        "animals_needed": round(animals, 1),
        "bottleneck": worst.cut_code,
        "bottleneck_desc": worst.description,
        "excess": excess,
    }


# ---------------------------------------------------------------------------
# Carcass allocation
# ---------------------------------------------------------------------------

def allocate_carcass(
    buyers: List[BuyerProfile],
    usda_prices: dict,
    ground_beef_prices: dict,
    valuation,
    purchase_price_cwt: float,
    processor_key: str = DEFAULT_PROCESSOR,
    region: str = DEFAULT_REGION,
) -> AllocationResult:
    processor = PROCESSORS[processor_key]
    regional_adj = REGIONS.get(region, REGIONS[DEFAULT_REGION])["pricing_adjustment"]
    hcw = valuation.hot_carcass_weight
    live_weight = valuation.live_weight
    quality_grade = valuation.quality_grade.lower()
    grade_rank = GRADE_RANK.get(quality_grade, 2)

    allocations = []
    unallocated = []
    total_revenue = 0.0
    unallocated_value = 0.0

    # Build all allocatable cuts: IMPS subprimals + ground products
    all_cuts = {}
    for code, (desc, yield_pct, primal) in SUBPRIMAL_YIELDS.items():
        cut_weight = hcw * (yield_pct / 100.0)
        base_cwt = _get_base_price_cwt(code, usda_prices, ground_beef_prices)
        all_cuts[code] = {
            "desc": desc, "weight": cut_weight, "base_cwt": base_cwt, "primal": primal,
        }

    # Ground beef: split trim proportionally among ground products with demand
    ground_total_weight = hcw * TRIM_YIELD_PCT
    ground_codes_with_demand = []
    for gcode in GROUND_BEEF_PRODUCTS:
        if _get_base_price_cwt(gcode, usda_prices, ground_beef_prices) > 0:
            ground_codes_with_demand.append(gcode)
    if not ground_codes_with_demand:
        ground_codes_with_demand = list(GROUND_BEEF_PRODUCTS.keys())

    per_ground = ground_total_weight / len(ground_codes_with_demand) if ground_codes_with_demand else 0
    for gcode in ground_codes_with_demand:
        base_cwt = _get_base_price_cwt(gcode, usda_prices, ground_beef_prices)
        all_cuts[gcode] = {
            "desc": GROUND_BEEF_PRODUCTS[gcode]["description"],
            "weight": per_ground,
            "base_cwt": base_cwt,
            "primal": "Trim",
        }

    # For each cut, allocate to buyers sorted by price (highest first)
    for code, info in all_cuts.items():
        remaining = info["weight"]
        base_cwt = info["base_cwt"]
        desc = info["desc"]
        if remaining <= 0 or base_cwt <= 0:
            continue

        # Find eligible buyers for this cut
        eligible = []
        for buyer in buyers:
            if not buyer.active:
                continue
            buyer_grade_rank = GRADE_RANK.get(buyer.min_quality_grade.lower(), 1)
            if grade_rank < buyer_grade_rank:
                continue
            for pref in buyer.cut_preferences:
                if pref.cut_code == code:
                    price_lb = compute_buyer_price(code, pref, base_cwt, regional_adj)
                    eligible.append({
                        "buyer": buyer,
                        "pref": pref,
                        "price_lb": price_lb,
                    })
                    break

        eligible.sort(key=lambda x: -x["price_lb"])

        for e in eligible:
            if remaining <= 0:
                break
            alloc_lbs = min(e["pref"].volume_lbs_week, remaining)
            line_rev = alloc_lbs * e["price_lb"]
            allocations.append(AllocationLine(
                buyer_id=e["buyer"].buyer_id,
                buyer_name=e["buyer"].name,
                cut_code=code,
                description=desc,
                lbs_allocated=round(alloc_lbs, 1),
                price_per_lb=round(e["price_lb"], 2),
                line_revenue=round(line_rev, 2),
            ))
            total_revenue += line_rev
            remaining -= alloc_lbs

        # Remainder at wholesale
        if remaining > 0.1:
            wholesale_lb = (base_cwt / 100.0) * regional_adj
            uv = remaining * wholesale_lb
            unallocated.append({
                "cut_code": code,
                "description": desc,
                "lbs": round(remaining, 1),
                "wholesale_lb": round(wholesale_lb, 2),
                "value": round(uv, 2),
            })
            unallocated_value += uv

    # Margin calculation
    farmer_cost = purchase_price_cwt * live_weight / 100.0
    kill_fee = processor["kill_fee"]
    fab_total = processor["fab_cost_per_lb"] * hcw
    shrink_cost = processor["shrink_pct"] * valuation.total_subprimal_value
    processing_cost = kill_fee + fab_total + shrink_cost
    gross_margin = total_revenue + unallocated_value - farmer_cost - processing_cost
    margin_pct = (gross_margin / (total_revenue + unallocated_value) * 100) if (total_revenue + unallocated_value) > 0 else 0

    return AllocationResult(
        quality_grade=valuation.quality_grade,
        hcw=hcw,
        live_weight=live_weight,
        allocations=allocations,
        unallocated=unallocated,
        total_revenue=round(total_revenue, 2),
        unallocated_value=round(unallocated_value, 2),
        farmer_cost=round(farmer_cost, 2),
        processing_cost=round(processing_cost, 2),
        gross_margin=round(gross_margin, 2),
        margin_pct=round(margin_pct, 1),
    )


# ---------------------------------------------------------------------------
# Console output
# ---------------------------------------------------------------------------

def print_demand_report(demand_lines: List[DemandLine], animals_needed: dict, quality_grade: str, region: str):
    region_label = REGIONS.get(region, REGIONS[DEFAULT_REGION])["label"]
    print(f"\nDEMAND ANALYSIS — {quality_grade.title()} | Region: {region_label}")
    print("=" * 90)
    print(f"{'IMPS':<14} {'Cut':<26} {'Primal':<8} {'Demand/wk':>10} {'Buyers':>7} {'Avg$/lb':>8} {'Yield':>7} {'Ratio':>7}")
    print("-" * 90)

    for line in demand_lines:
        flag = " \u25b2" if line.supply_ratio > 1.0 else ""
        print(f"{line.cut_code:<14} {line.description:<26} {line.primal:<8} "
              f"{line.total_weekly_demand_lbs:>8,.0f} lb {line.num_buyers:>6} "
              f"${line.avg_price_lb:>6.2f} {line.carcass_yield_lbs:>6.1f} "
              f"{line.supply_ratio:>5.1f}x{flag}")

    print("-" * 90)
    an = animals_needed
    print(f"Animals/week to meet all demand: {an['animals_needed']} "
          f"(bottleneck: {an.get('bottleneck_desc', an['bottleneck'])})")
    if an.get("excess"):
        top_excess = sorted(an["excess"].items(), key=lambda x: -x[1])[:5]
        excess_str = ", ".join(f"{c}: {lbs:.0f} lb" for c, lbs in top_excess)
        print(f"Top excess production: {excess_str}")
    print()


def print_allocation_report(result: AllocationResult):
    print(f"\nCARCASS ALLOCATION — {result.quality_grade}, {result.hcw:.0f} lb carcass")
    print("=" * 78)
    print(f"{'Buyer':<22} {'Cut':<18} {'Lbs':>6} {'$/lb':>7} {'Revenue':>10}")
    print("-" * 78)

    for a in sorted(result.allocations, key=lambda x: -x.line_revenue):
        code_desc = f"{a.cut_code} {a.description}"
        if len(code_desc) > 17:
            code_desc = code_desc[:17]
        print(f"{a.buyer_name:<22} {code_desc:<18} {a.lbs_allocated:>5.1f} "
              f"${a.price_per_lb:>6.2f} ${a.line_revenue:>9,.2f}")

    if result.unallocated:
        print("-" * 78)
        for u in result.unallocated:
            print(f"{'Unallocated':<22} {u['cut_code']:<18} {u['lbs']:>5.1f} "
                  f"${u['wholesale_lb']:>6.2f} ${u['value']:>9,.2f}")

    total_rev = result.total_revenue + result.unallocated_value
    print("-" * 78)
    print(f"{'Total Revenue:':<42} ${total_rev:>12,.2f}")
    print(f"{'Farmer Cost:':<42} -${result.farmer_cost:>11,.2f}")
    print(f"{'Processing:':<42} -${result.processing_cost:>11,.2f}")
    margin_sign = "" if result.gross_margin >= 0 else "-"
    print(f"{'MARGIN:':<42} {margin_sign}${abs(result.gross_margin):>11,.2f}  ({result.margin_pct:.1f}%)")
    print()


def print_buyer_list(buyers: List[BuyerProfile]):
    print(f"\nBUYER DIRECTORY — {len(buyers)} profiles")
    print("=" * 95)
    print(f"{'ID':<16} {'Name':<24} {'Type':<18} {'Region':<16} {'Grade':<8} {'Active':<7} {'Cuts':>5}")
    print("-" * 95)
    for b in buyers:
        region_label = REGIONS.get(b.region, {}).get("label", b.region)
        print(f"{b.buyer_id:<16} {b.name:<24} {b.buyer_type:<18} {region_label:<16} "
              f"{b.min_quality_grade:<8} {'Yes' if b.active else 'No':<7} {len(b.cut_preferences):>5}")
    print()


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------

def write_demand_excel(
    demand_lines: List[DemandLine],
    animals_needed: dict,
    allocation: AllocationResult,
    buyers: List[BuyerProfile],
    filepath: str = "demand_report.xlsx",
):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("openpyxl not installed. Run: pip3 install openpyxl")
        return

    wb = Workbook()
    hdr_font = Font(bold=True, size=11)
    hdr_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    hdr_font_white = Font(bold=True, color="FFFFFF", size=11)
    money_fmt = '#,##0.00'
    pct_fmt = '0.0%'
    thin_border = Border(
        bottom=Side(style="thin", color="CCCCCC"),
    )

    def style_header(ws, row, ncols):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = hdr_font_white
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center")

    # --- Sheet 1: Demand Analysis ---
    ws1 = wb.active
    ws1.title = "Demand Analysis"
    headers = ["IMPS", "Cut", "Primal", "Demand/wk (lb)", "Buyers",
               "Avg $/lb", "High $/lb", "Low $/lb", "Yield (lb)", "Supply Ratio"]
    for c, h in enumerate(headers, 1):
        ws1.cell(row=1, column=c, value=h)
    style_header(ws1, 1, len(headers))

    for i, line in enumerate(demand_lines, 2):
        ws1.cell(row=i, column=1, value=line.cut_code)
        ws1.cell(row=i, column=2, value=line.description)
        ws1.cell(row=i, column=3, value=line.primal)
        ws1.cell(row=i, column=4, value=line.total_weekly_demand_lbs)
        ws1.cell(row=i, column=5, value=line.num_buyers)
        ws1.cell(row=i, column=6, value=line.avg_price_lb).number_format = money_fmt
        ws1.cell(row=i, column=7, value=line.highest_price_lb).number_format = money_fmt
        ws1.cell(row=i, column=8, value=line.lowest_price_lb).number_format = money_fmt
        ws1.cell(row=i, column=9, value=line.carcass_yield_lbs)
        ws1.cell(row=i, column=10, value=line.supply_ratio)

    r = len(demand_lines) + 3
    ws1.cell(row=r, column=1, value="Animals/week needed:").font = hdr_font
    ws1.cell(row=r, column=3, value=animals_needed["animals_needed"])
    ws1.cell(row=r + 1, column=1, value="Bottleneck cut:").font = hdr_font
    ws1.cell(row=r + 1, column=3, value=animals_needed.get("bottleneck_desc", animals_needed["bottleneck"]))

    for col in ws1.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws1.column_dimensions[col[0].column_letter].width = min(max_len + 3, 22)

    # --- Sheet 2: Allocation ---
    ws2 = wb.create_sheet("Allocation")
    headers2 = ["Buyer ID", "Buyer Name", "Cut Code", "Cut", "Lbs", "$/lb", "Revenue"]
    for c, h in enumerate(headers2, 1):
        ws2.cell(row=1, column=c, value=h)
    style_header(ws2, 1, len(headers2))

    for i, a in enumerate(sorted(allocation.allocations, key=lambda x: -x.line_revenue), 2):
        ws2.cell(row=i, column=1, value=a.buyer_id)
        ws2.cell(row=i, column=2, value=a.buyer_name)
        ws2.cell(row=i, column=3, value=a.cut_code)
        ws2.cell(row=i, column=4, value=a.description)
        ws2.cell(row=i, column=5, value=a.lbs_allocated)
        ws2.cell(row=i, column=6, value=a.price_per_lb).number_format = money_fmt
        ws2.cell(row=i, column=7, value=a.line_revenue).number_format = money_fmt

    r2 = len(allocation.allocations) + 3
    if allocation.unallocated:
        ws2.cell(row=r2, column=1, value="UNALLOCATED").font = hdr_font
        r2 += 1
        for u in allocation.unallocated:
            ws2.cell(row=r2, column=3, value=u["cut_code"])
            ws2.cell(row=r2, column=4, value=u["description"])
            ws2.cell(row=r2, column=5, value=u["lbs"])
            ws2.cell(row=r2, column=6, value=u["wholesale_lb"]).number_format = money_fmt
            ws2.cell(row=r2, column=7, value=u["value"]).number_format = money_fmt
            r2 += 1

    for col in ws2.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws2.column_dimensions[col[0].column_letter].width = min(max_len + 3, 22)

    # --- Sheet 3: Margin Analysis ---
    ws3 = wb.create_sheet("Margin Analysis")
    margin_data = [
        ("Quality Grade", allocation.quality_grade),
        ("Hot Carcass Weight (lb)", allocation.hcw),
        ("Live Weight (lb)", allocation.live_weight),
        ("", ""),
        ("Allocated Revenue", allocation.total_revenue),
        ("Unallocated (wholesale)", allocation.unallocated_value),
        ("Total Revenue", allocation.total_revenue + allocation.unallocated_value),
        ("", ""),
        ("Farmer Cost", allocation.farmer_cost),
        ("Processing Cost", allocation.processing_cost),
        ("Total Cost", allocation.farmer_cost + allocation.processing_cost),
        ("", ""),
        ("Gross Margin", allocation.gross_margin),
        ("Margin %", allocation.margin_pct / 100.0),
    ]
    ws3.cell(row=1, column=1, value="Metric").font = hdr_font
    ws3.cell(row=1, column=2, value="Value").font = hdr_font
    style_header(ws3, 1, 2)
    for i, (label, val) in enumerate(margin_data, 2):
        ws3.cell(row=i, column=1, value=label)
        cell = ws3.cell(row=i, column=2, value=val)
        if isinstance(val, float) and label and label != "Margin %":
            cell.number_format = money_fmt
        elif label == "Margin %":
            cell.number_format = pct_fmt
    ws3.column_dimensions["A"].width = 25
    ws3.column_dimensions["B"].width = 18

    # --- Sheet 4: Buyer Directory ---
    ws4 = wb.create_sheet("Buyer Directory")
    headers4 = ["Buyer ID", "Name", "Type", "City", "State", "Region",
                "Min Grade", "Payment Terms", "Active", "Contact", "Email", "Phone", "Num Cuts"]
    for c, h in enumerate(headers4, 1):
        ws4.cell(row=1, column=c, value=h)
    style_header(ws4, 1, len(headers4))

    for i, b in enumerate(buyers, 2):
        ws4.cell(row=i, column=1, value=b.buyer_id)
        ws4.cell(row=i, column=2, value=b.name)
        ws4.cell(row=i, column=3, value=b.buyer_type)
        ws4.cell(row=i, column=4, value=b.city)
        ws4.cell(row=i, column=5, value=b.state)
        ws4.cell(row=i, column=6, value=REGIONS.get(b.region, {}).get("label", b.region))
        ws4.cell(row=i, column=7, value=b.min_quality_grade)
        ws4.cell(row=i, column=8, value=b.payment_terms_days)
        ws4.cell(row=i, column=9, value="Yes" if b.active else "No")
        ws4.cell(row=i, column=10, value=b.contact_name)
        ws4.cell(row=i, column=11, value=b.contact_email)
        ws4.cell(row=i, column=12, value=b.contact_phone)
        ws4.cell(row=i, column=13, value=len(b.cut_preferences))

    for col in ws4.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws4.column_dimensions[col[0].column_letter].width = min(max_len + 3, 22)

    wb.save(filepath)
    print(f"Demand report saved to {filepath}")
